import sys
import tkinter
from tkinter import messagebox
import mysql.connector
import dbConnect
from dbConnect import DBConnect
import mysql.connector
import time
import smtplib
import os
from openpyxl import Workbook

try:
    import Tkinter as tk
except ImportError:
    import tkinter as tk

try:
    import ttk
    py3 = False
except ImportError:
    import tkinter.ttk as ttk
    py3 = True

import unknown_support
from tkinter import*






















def main():                                       # here the main application starts

    try:
        import Tkinter as tk
    except ImportError:
        import tkinter as tk

    try:
        import ttk
        py3 = False
    except ImportError:
        import tkinter.ttk as ttk
        py3 = True

    import unknown_support

    def vp_start_gui():
        '''Starting point when module is the main routine.'''
        global val, w, root
        root = tk.Tk()
        top = Main (root)
        unknown_support.init(root, top)
        root.mainloop()

    w = None
    def create_Main(root, *args, **kwargs):
        '''Starting point when module is imported by another program.'''
        global w, w_win, rt
        rt = root
        w = tk.Toplevel (root)
        top = Main (w)
        unknown_support.init(w, top, *args, **kwargs)
        return (w, top)

    def destroy_Main():
        global w
        w.destroy()
        w = None

    






























    def newdata1():
        try:
            import Tkinter as tk
        except ImportError:
            import tkinter as tk

        try:
            import ttk
            py3 = False
        except ImportError:
            import tkinter.ttk as ttk
            py3 = True

        import new_support

        def vp_start_gui():
            '''Starting point when module is the main routine.'''
            global val, w, root
            root = tk.Tk()
            new_support.set_Tk_var()
            top = Toplevel1 (root)
            new_support.init(root, top)
            root.mainloop()

        w = None
        def create_Toplevel1(root, *args, **kwargs):
            '''Starting point when module is imported by another program.'''
            global w, w_win, rt
            rt = root
            w = tk.Toplevel (root)
            new_support.set_Tk_var()
            top = Toplevel1 (w)
            new_support.init(w, top, *args, **kwargs)
            return (w, top)

        def destroy_Toplevel1():
            global w
            w.destroy()
            w = None

        class Toplevel1:
            def exits(self):
              msg=tkinter.messagebox.askyesno('etax-2019','Do You Want To Exit ?');
              if(msg):
                os._exit(1)


            def continues(self):
              a=str(self.txtid.get());
              b=str(self.txtname.get());
              c=str(self.txtmeter.get());
              d=str(self.txtward.get());
              e=str(self.txthousetax.get());
              f=str(self.txthealthtax.get());
              g=str(self.txtlighttax.get());
              h=str(self.txtwatertax.get());
              j=float(e)
              k=float(f)
              l=float(g)
              m=float(h)
              i=j+k+l+m
              i=str(i)
              k=i

              database = DBConnect(host='remotemysql.com',user='NhpDbfVMT3',password='NXWSVJNu55',database='NhpDbfVMT3',port='3306')
              new_user = {'idnumber': a,'meternumber': c,'wardnumber': d,'name': b,'housetax': e,'healthtax': f,'lighttax': g,'watertax': h,'total': i,'rest': k}
              database.insert(new_user,'kalamwadiinfo')
              tkinter.messagebox.showinfo('etax-2019','Data was entered successfully')
              root.destroy()
              main()
            def backs(self):
                root.destroy()
                main()
            def __init__(self, top=None):
                _bgcolor = '#d9d9d9'  # X11 color: 'gray85'
                _fgcolor = '#000000'  # X11 color: 'black'
                _compcolor = '#d9d9d9' # X11 color: 'gray85'
                _ana1color = '#d9d9d9' # X11 color: 'gray85' 
                _ana2color = '#ececec' # Closest X11 color: 'gray92' 
                font10 = "-family {Segoe Script} -size 12 -weight normal "  \
                    "-slant italic -underline 0 -overstrike 0"
                font11 = "-family {Rockwell Extra Bold} -size 12 -weight bold "  \
                    "-slant roman -underline 0 -overstrike 0"
                font12 = "-family {Sakkal Majalla} -size 17 -weight bold "  \
                    "-slant roman -underline 0 -overstrike 0"
                font13 = "-family {Segoe UI Semibold} -size 14 -weight bold "  \
                    "-slant roman -underline 0 -overstrike 0"
                font14 = "-family {Bookman Old Style} -size 17 -weight bold "  \
                    "-slant roman -underline 0 -overstrike 0"
                font15 = "-family {Segoe UI Semibold} -size 12 -weight bold "  \
                    "-slant roman -underline 0 -overstrike 0"
                font9 = "-family {Britannic Bold} -size 48 -weight bold -slant"  \
                    " roman -underline 0 -overstrike 0"
                self.style = ttk.Style()
                if sys.platform == "win32":
                    self.style.theme_use('winnative')
                self.style.configure('.',background=_bgcolor)
                self.style.configure('.',foreground=_fgcolor)
                self.style.configure('.',font="TkDefaultFont")
                self.style.map('.',background=
                    [('selected', _compcolor), ('active',_ana2color)])

                top.geometry("1538x878+-109+36")
                top.title("New Toplevel")
                top.configure(background="#ffff24")
                top.configure(highlightbackground="#d9d9d9")
                top.configure(highlightcolor="black")

                self.Label1 = tk.Label(top)
                self.Label1.place(relx=0.013, rely=0.023, height=81, width=156)
                self.Label1.configure(activebackground="#f9f9f9")
                self.Label1.configure(activeforeground="black")
                self.Label1.configure(background="#ffff24")
                self.Label1.configure(disabledforeground="#a3a3a3")
                self.Label1.configure(font=font9)
                self.Label1.configure(foreground="#ff250d")
                self.Label1.configure(highlightbackground="#d9d9d9")
                self.Label1.configure(highlightcolor="black")
                self.Label1.configure(text='''eTAX''')

                self.menubar = tk.Menu(top,font="TkMenuFont",bg=_bgcolor,fg=_fgcolor)
                top.configure(menu = self.menubar)

                self.Label2 = tk.Label(top)
                self.Label2.place(relx=0.072, rely=0.103, height=31, width=141)
                self.Label2.configure(activebackground="#f9f9f9")
                self.Label2.configure(activeforeground="black")
                self.Label2.configure(background="#ffff24")
                self.Label2.configure(disabledforeground="#a3a3a3")
                self.Label2.configure(font=font10)
                self.Label2.configure(foreground="#13c12a")
                self.Label2.configure(highlightbackground="#d9d9d9")
                self.Label2.configure(highlightcolor="black")
                self.Label2.configure(text='''working for you''')

                self.backbutton = tk.Button(top)
                self.backbutton.place(relx=0.033, rely=0.194, height=44, width=97)
                self.backbutton.configure(activebackground="#ececec")
                self.backbutton.configure(activeforeground="#000000")
                self.backbutton.configure(background="#120bd8")
                self.backbutton.configure(disabledforeground="#a3a3a3")
                self.backbutton.configure(font=font11)
                self.backbutton.configure(foreground="#fcffff")
                self.backbutton.configure(highlightbackground="#d9d9d9")
                self.backbutton.configure(highlightcolor="black")
                self.backbutton.configure(pady="0")
                self.backbutton.configure(command = self.backs,text='''Back''')

                self.exit = tk.Button(top)
                self.exit.place(relx=0.033, rely=0.285, height=44, width=97)
                self.exit.configure(activebackground="#ececec")
                self.exit.configure(activeforeground="#000000")
                self.exit.configure(background="#120bd8")
                self.exit.configure(disabledforeground="#a3a3a3")
                self.exit.configure(font=font11)
                self.exit.configure(foreground="#fcffff")
                self.exit.configure(highlightbackground="#d9d9d9")
                self.exit.configure(highlightcolor="black")
                self.exit.configure(pady="0")
                self.exit.configure(command = self.exits, text='''Exit''')

                self.Label3 = tk.Label(top)
                self.Label3.place(relx=0.013, rely=0.9, height=21, width=56)
                self.Label3.configure(activebackground="#f9f9f9")
                self.Label3.configure(activeforeground="black")
                self.Label3.configure(background="#ffff24")
                self.Label3.configure(disabledforeground="#a3a3a3")
                self.Label3.configure(foreground="#000000")
                self.Label3.configure(highlightbackground="#d9d9d9")
                self.Label3.configure(highlightcolor="black")
                self.Label3.configure(text='''etax-2019''')

                self.Label3_3 = tk.Label(top)
                self.Label3_3.place(relx=0.013, rely=0.923, height=21, width=34)
                self.Label3_3.configure(activebackground="#f9f9f9")
                self.Label3_3.configure(activeforeground="black")
                self.Label3_3.configure(background="#ffff24")
                self.Label3_3.configure(disabledforeground="#a3a3a3")
                self.Label3_3.configure(foreground="#000000")
                self.Label3_3.configure(highlightbackground="#d9d9d9")
                self.Label3_3.configure(highlightcolor="black")
                self.Label3_3.configure(text='''v 1.0.2''')

                self.Label3_4 = tk.Label(top)
                self.Label3_4.place(relx=0.007, rely=0.968, height=21, width=134)
                self.Label3_4.configure(activebackground="#f9f9f9")
                self.Label3_4.configure(activeforeground="black")
                self.Label3_4.configure(background="#ffff24")
                self.Label3_4.configure(disabledforeground="#a3a3a3")
                self.Label3_4.configure(foreground="#000000")
                self.Label3_4.configure(highlightbackground="#d9d9d9")
                self.Label3_4.configure(highlightcolor="black")
                self.Label3_4.configure(text='''Working On Windows''')

                self.Label3_1 = tk.Label(top)
                self.Label3_1.place(relx=0.013, rely=0.945, height=21, width=164)
                self.Label3_1.configure(activebackground="#f9f9f9")
                self.Label3_1.configure(activeforeground="black")
                self.Label3_1.configure(background="#ffff24")
                self.Label3_1.configure(disabledforeground="#a3a3a3")
                self.Label3_1.configure(foreground="#000000")
                self.Label3_1.configure(highlightbackground="#d9d9d9")
                self.Label3_1.configure(highlightcolor="black")
                self.Label3_1.configure(text='''Connected to MySQL server 8.0''')

                self.Frame1 = tk.Frame(top)
                self.Frame1.place(relx=0.293, rely=0.216, relheight=0.655
                        , relwidth=0.452)
                self.Frame1.configure(relief='groove')
                self.Frame1.configure(borderwidth="2")
                self.Frame1.configure(relief='groove')
                self.Frame1.configure(background="#68ff3b")
                self.Frame1.configure(width=695)

                self.txtward = tk.Entry(self.Frame1)
                self.txtward.place(relx=0.504, rely=0.243,height=30, relwidth=0.437)
                self.txtward.configure(background="white")
                self.txtward.configure(disabledforeground="#a3a3a3")
                self.txtward.configure(font="TkFixedFont")
                self.txtward.configure(foreground="#000000")
                self.txtward.configure(insertbackground="black")
                self.txtward.configure(width=304)

                self.txtname = tk.Entry(self.Frame1)
                self.txtname.place(relx=0.504, rely=0.157,height=30, relwidth=0.437)
                self.txtname.configure(background="white")
                self.txtname.configure(disabledforeground="#a3a3a3")
                self.txtname.configure(font="TkFixedFont")
                self.txtname.configure(foreground="#000000")
                self.txtname.configure(insertbackground="black")
                self.txtname.configure(width=304)

                self.txtmeter = tk.Entry(self.Frame1)
                self.txtmeter.place(relx=0.504, rely=0.33,height=30, relwidth=0.437)
                self.txtmeter.configure(background="white")
                self.txtmeter.configure(disabledforeground="#a3a3a3")
                self.txtmeter.configure(font="TkFixedFont")
                self.txtmeter.configure(foreground="#000000")
                self.txtmeter.configure(insertbackground="black")
                self.txtmeter.configure(width=304)

                self.Label5 = tk.Label(self.Frame1)
                self.Label5.place(relx=0.23, rely=0.07, height=37, width=107)
                self.Label5.configure(background="#68ff3b")
                self.Label5.configure(disabledforeground="#a3a3a3")
                self.Label5.configure(font=font12)
                self.Label5.configure(foreground="#4124b5")
                self.Label5.configure(text='''UID Number :''')
                self.Label5.configure(width=107)

                self.Label5_2 = tk.Label(self.Frame1)
                self.Label5_2.place(relx=0.216, rely=0.313, height=47, width=127)
                self.Label5_2.configure(activebackground="#f9f9f9")
                self.Label5_2.configure(activeforeground="black")
                self.Label5_2.configure(background="#68ff3b")
                self.Label5_2.configure(disabledforeground="#a3a3a3")
                self.Label5_2.configure(font=font12)
                self.Label5_2.configure(foreground="#4124b5")
                self.Label5_2.configure(highlightbackground="#d9d9d9")
                self.Label5_2.configure(highlightcolor="black")
                self.Label5_2.configure(text='''Meter Number :''')
                self.Label5_2.configure(width=127)

                self.Label5_3 = tk.Label(self.Frame1)
                self.Label5_3.place(relx=0.23, rely=0.243, height=37, width=117)
                self.Label5_3.configure(activebackground="#f9f9f9")
                self.Label5_3.configure(activeforeground="black")
                self.Label5_3.configure(background="#68ff3b")
                self.Label5_3.configure(disabledforeground="#a3a3a3")
                self.Label5_3.configure(font=font12)
                self.Label5_3.configure(foreground="#4124b5")
                self.Label5_3.configure(highlightbackground="#d9d9d9")
                self.Label5_3.configure(highlightcolor="black")
                self.Label5_3.configure(text='''Ward Number :''')
                self.Label5_3.configure(width=117)

                self.Label5_4 = tk.Label(self.Frame1)
                self.Label5_4.place(relx=0.245, rely=0.417, height=37, width=97)
                self.Label5_4.configure(activebackground="#f9f9f9")
                self.Label5_4.configure(activeforeground="black")
                self.Label5_4.configure(background="#68ff3b")
                self.Label5_4.configure(disabledforeground="#a3a3a3")
                self.Label5_4.configure(font=font12)
                self.Label5_4.configure(foreground="#4124b5")
                self.Label5_4.configure(highlightbackground="#d9d9d9")
                self.Label5_4.configure(highlightcolor="black")
                self.Label5_4.configure(text='''Housetax :''')
                self.Label5_4.configure(width=97)

                self.Label5_5 = tk.Label(self.Frame1)
                self.Label5_5.place(relx=0.23, rely=0.157, height=37, width=107)
                self.Label5_5.configure(activebackground="#f9f9f9")
                self.Label5_5.configure(activeforeground="black")
                self.Label5_5.configure(background="#68ff3b")
                self.Label5_5.configure(disabledforeground="#a3a3a3")
                self.Label5_5.configure(font=font12)
                self.Label5_5.configure(foreground="#4124b5")
                self.Label5_5.configure(highlightbackground="#d9d9d9")
                self.Label5_5.configure(highlightcolor="black")
                self.Label5_5.configure(text='''Name :''')
                self.Label5_5.configure(width=107)

                self.Label5_6 = tk.Label(self.Frame1)
                self.Label5_6.place(relx=0.23, rely=0.504, height=37, width=107)
                self.Label5_6.configure(activebackground="#f9f9f9")
                self.Label5_6.configure(activeforeground="black")
                self.Label5_6.configure(background="#68ff3b")
                self.Label5_6.configure(disabledforeground="#a3a3a3")
                self.Label5_6.configure(font=font12)
                self.Label5_6.configure(foreground="#4124b5")
                self.Label5_6.configure(highlightbackground="#d9d9d9")
                self.Label5_6.configure(highlightcolor="black")
                self.Label5_6.configure(text='''Healthtax :''')
                self.Label5_6.configure(width=107)

                self.Label5_7 = tk.Label(self.Frame1)
                self.Label5_7.place(relx=0.23, rely=0.574, height=47, width=107)
                self.Label5_7.configure(activebackground="#f9f9f9")
                self.Label5_7.configure(activeforeground="black")
                self.Label5_7.configure(background="#68ff3b")
                self.Label5_7.configure(disabledforeground="#a3a3a3")
                self.Label5_7.configure(font=font12)
                self.Label5_7.configure(foreground="#4124b5")
                self.Label5_7.configure(highlightbackground="#d9d9d9")
                self.Label5_7.configure(highlightcolor="black")
                self.Label5_7.configure(text='''Lighttax :''')
                self.Label5_7.configure(width=107)

                self.Label5_8 = tk.Label(self.Frame1)
                self.Label5_8.place(relx=0.23, rely=0.661, height=47, width=107)
                self.Label5_8.configure(activebackground="#f9f9f9")
                self.Label5_8.configure(activeforeground="black")
                self.Label5_8.configure(background="#68ff3b")
                self.Label5_8.configure(disabledforeground="#a3a3a3")
                self.Label5_8.configure(font=font12)
                self.Label5_8.configure(foreground="#4124b5")
                self.Label5_8.configure(highlightbackground="#d9d9d9")
                self.Label5_8.configure(highlightcolor="black")
                self.Label5_8.configure(text='''Watertax :''')
                self.Label5_8.configure(width=107)

                self.buttncontinue = tk.Button(self.Frame1)
                self.buttncontinue.place(relx=0.763, rely=0.887, height=42, width=115)
                self.buttncontinue.configure(activebackground="#ececec")
                self.buttncontinue.configure(activeforeground="#000000")
                self.buttncontinue.configure(background="#5a54ff")
                self.buttncontinue.configure(disabledforeground="#a3a3a3")
                self.buttncontinue.configure(font=font13)
                self.buttncontinue.configure(foreground="#fafffc")
                self.buttncontinue.configure(highlightbackground="#d9d9d9")
                self.buttncontinue.configure(highlightcolor="black")
                self.buttncontinue.configure(pady="0")
                self.buttncontinue.configure(command = self.continues, text='''CONTINUE''')
                self.buttncontinue.configure(width=115)

                self.style.map('TCheckbutton',background=
                    [('selected', _bgcolor), ('active', _ana2color)])
                self.checkbox = ttk.Checkbutton(self.Frame1)
                self.checkbox.place(relx=0.216, rely=0.835, relwidth=0.706, relheight=0.0
                        , height=21)
                self.checkbox.configure(variable=new_support.tch66)
                self.checkbox.configure(takefocus="")
                self.checkbox.configure(text='''All the data filled is correct I Wish To Continue''')
                self.checkbox.configure(width=491)

                self.Label4 = tk.Label(top)
                self.Label4.place(relx=0.377, rely=0.057, height=101, width=444)
                self.Label4.configure(background="#ffff24")
                self.Label4.configure(disabledforeground="#a3a3a3")
                self.Label4.configure(font=font9)
                self.Label4.configure(foreground="#35a025")
                self.Label4.configure(text='''New Entry''')
                self.Label4.configure(width=444)

                self.txtid = tk.Entry(top)
                self.txtid.place(relx=0.52, rely=0.262,height=30, relwidth=0.198)
                self.txtid.configure(background="white")
                self.txtid.configure(disabledforeground="#a3a3a3")
                self.txtid.configure(font="TkFixedFont")
                self.txtid.configure(foreground="#000000")
                self.txtid.configure(insertbackground="black")
                self.txtid.configure(width=304)

                self.txtwatertax = tk.Entry(top)
                self.txtwatertax.place(relx=0.52, rely=0.661,height=30, relwidth=0.198)
                self.txtwatertax.configure(background="white")
                self.txtwatertax.configure(disabledforeground="#a3a3a3")
                self.txtwatertax.configure(font="TkFixedFont")
                self.txtwatertax.configure(foreground="#000000")
                self.txtwatertax.configure(insertbackground="black")
                self.txtwatertax.configure(width=304)

                self.txthousetax = tk.Entry(top)
                self.txthousetax.place(relx=0.52, rely=0.49,height=30, relwidth=0.198)
                self.txthousetax.configure(background="white")
                self.txthousetax.configure(disabledforeground="#a3a3a3")
                self.txthousetax.configure(font="TkFixedFont")
                self.txthousetax.configure(foreground="#000000")
                self.txthousetax.configure(insertbackground="black")
                self.txthousetax.configure(width=304)

                self.txthealthtax = tk.Entry(top)
                self.txthealthtax.place(relx=0.52, rely=0.547, height=30, relwidth=0.198)

                self.txthealthtax.configure(background="white")
                self.txthealthtax.configure(disabledforeground="#a3a3a3")
                self.txthealthtax.configure(font="TkFixedFont")
                self.txthealthtax.configure(foreground="#000000")
                self.txthealthtax.configure(insertbackground="black")
                self.txthealthtax.configure(width=304)

                self.txtlighttax = tk.Entry(top)
                self.txtlighttax.place(relx=0.52, rely=0.604,height=30, relwidth=0.198)
                self.txtlighttax.configure(background="white")
                self.txtlighttax.configure(disabledforeground="#a3a3a3")
                self.txtlighttax.configure(font="TkFixedFont")
                self.txtlighttax.configure(foreground="#000000")
                self.txtlighttax.configure(insertbackground="black")
                self.txtlighttax.configure(width=304)

                self.Label1_1 = tk.Label(top)
                self.Label1_1.place(relx=0.117, rely=0.023, height=81, width=156)
                self.Label1_1.configure(activebackground="#f9f9f9")
                self.Label1_1.configure(activeforeground="black")
                self.Label1_1.configure(background="#ffff24")
                self.Label1_1.configure(disabledforeground="#a3a3a3")
                self.Label1_1.configure(font=font9)
                self.Label1_1.configure(foreground="#2212ff")
                self.Label1_1.configure(highlightbackground="#d9d9d9")
                self.Label1_1.configure(highlightcolor="black")
                self.Label1_1.configure(text='''2019''')

                self.Label1_11 = tk.Label(top)
                self.Label1_11.place(relx=0.826, rely=0.182, height=81, width=246)
                self.Label1_11.configure(activebackground="#f9f9f9")
                self.Label1_11.configure(activeforeground="black")
                self.Label1_11.configure(background="#ffff24")
                self.Label1_11.configure(disabledforeground="#a3a3a3")
                self.Label1_11.configure(font=font14)
                self.Label1_11.configure(foreground="#268930")
                self.Label1_11.configure(highlightbackground="#d9d9d9")
                self.Label1_11.configure(highlightcolor="black")
                self.Label1_11.configure(text='''Village : Kalamwadi''')
                self.Label1_11.configure(width=246)

                self.Label1_12 = tk.Label(top)
                self.Label1_12.place(relx=0.826, rely=0.251, height=41, width=246)
                self.Label1_12.configure(activebackground="#f9f9f9")
                self.Label1_12.configure(activeforeground="black")
                self.Label1_12.configure(background="#ffff24")
                self.Label1_12.configure(disabledforeground="#a3a3a3")
                self.Label1_12.configure(font=font14)
                self.Label1_12.configure(foreground="#268930")
                self.Label1_12.configure(highlightbackground="#d9d9d9")
                self.Label1_12.configure(highlightcolor="black")
                self.Label1_12.configure(text='''Dist. : Sangli''')
                self.Label1_12.configure(width=246)

                self.Label6 = tk.Label(top)
                self.Label6.place(relx=0.904, rely=0.125, height=27, width=78)
                self.Label6.configure(background="#ffff24")
                self.Label6.configure(disabledforeground="#a3a3a3")
                self.Label6.configure(font=font15)
                self.Label6.configure(foreground="#000000")
                self.Label6.configure(text='''Hi User !!!''')

                self.Label7 = tk.Label(top)
                self.Label7.place(relx=0.904, rely=0.034, height=74, width=74)
                self.Label7.configure(background="#d9d9d9")
                self.Label7.configure(disabledforeground="#a3a3a3")
                self.Label7.configure(foreground="#000000")
                self._img1 = tk.PhotoImage(file="./login3.png")
                self.Label7.configure(image=self._img1)
                self.Label7.configure(text='''Label''')
                self.Label7.configure(width=74)

                self.TSeparator1 = ttk.Separator(top)
                self.TSeparator1.place(relx=0.826, rely=0.171, relwidth=0.176)

                self.TSeparator1_13 = ttk.Separator(top)
                self.TSeparator1_13.place(relx=0.829, rely=0.33, relwidth=0.202)

                self.Label8 = tk.Label(top)
                self.Label8.place(relx=0.026, rely=0.513, height=222, width=318)
                self.Label8.configure(background="#d9d9d9")
                self.Label8.configure(disabledforeground="#a3a3a3")
                self.Label8.configure(foreground="#000000")
                self._img2 = tk.PhotoImage(file="./login1.png")
                self.Label8.configure(image=self._img2)
                self.Label8.configure(text='''Label''')

                self.Label9 = tk.Label(top)
                self.Label9.place(relx=0.891, rely=0.945, height=21, width=117)
                self.Label9.configure(background="#ffff24")
                self.Label9.configure(disabledforeground="#a3a3a3")
                self.Label9.configure(foreground="#000000")
                self.Label9.configure(text='''Server Status : Online''')

                self.Label9_15 = tk.Label(top)
                self.Label9_15.place(relx=0.884, rely=0.968, height=21, width=147)
                self.Label9_15.configure(activebackground="#f9f9f9")
                self.Label9_15.configure(activeforeground="black")
                self.Label9_15.configure(background="#ffff24")
                self.Label9_15.configure(disabledforeground="#a3a3a3")
                self.Label9_15.configure(foreground="#000000")
                self.Label9_15.configure(highlightbackground="#d9d9d9")
                self.Label9_15.configure(highlightcolor="black")
                self.Label9_15.configure(text='''Connected To Local Host''')
                self.Label9_15.configure(width=147)

        if __name__ == '__main__':
            vp_start_gui()




















    


























































    def openworkspace1():

        try:
            import Tkinter as tk
        except ImportError:
            import tkinter as tk

        try:
            import ttk
            py3 = False
        except ImportError:
            import tkinter.ttk as ttk
            py3 = True

        import workspace_support
        import os.path

        def vp_start_gui():
            '''Starting point when module is the main routine.'''
            global val, w, root
            global prog_location
            prog_call = sys.argv[0]
            print ('prog_call = {}'.format(prog_call))
            prog_location = os.path.split(prog_call)[0]
            print ('prog_location = {}'.format(prog_location))
            sys.stdout.flush()
            root = tk.Tk()
            top = Toplevel1 (root)
            workspace_support.init(root, top)
            root.mainloop()

        w = None
        def create_Toplevel1(root, *args, **kwargs):
            '''Starting point when module is imported by another program.'''
            global w, w_win, rt
            global prog_location
            prog_call = sys.argv[0]
            print ('prog_call = {}'.format(prog_call))
            prog_location = os.path.split(prog_call)[0]
            print ('prog_location = {}'.format(prog_location))
            rt = root
            w = tk.Toplevel (root)
            top = Toplevel1 (w)
            workspace_support.init(w, top, *args, **kwargs)
            return (w, top)

        def destroy_Toplevel1():
            global w
            w.destroy()
            w = None

        class Toplevel1:
            def exits(self):
                msg=tkinter.messagebox.askyesno('etax-2019','Do You Want To Exit ?');
                if(msg):
                    os._exit(1)

            def viewnames(self):
                try :
                    mydb=mysql.connector.connect(host='remotemysql.com',user='NhpDbfVMT3',password='NXWSVJNu55',database='NhpDbfVMT3',port='3306')
                except :
                    tkinter.messagebox.showerror('etax-2019','Failed to connect server, Please contact your administrator')
                
                mycursor=mydb.cursor()
                query = ("SELECT idnumber, meternumber, wardnumber, name, housetax, healthtax, lighttax, watertax, total, rest FROM kalamwadiinfo")
                mycursor.execute(query)
                for(idnumber, meternumber, wardnumber, name, housetax, healthtax, lighttax, watertax, total, rest) in mycursor:
                    s="{}   {}".format(idnumber, name)
                    self.namebox.insert(0,s)


            def viewalls(self):
                try :
                    mydb=mysql.connector.connect(host='remotemysql.com',user='NhpDbfVMT3',password='NXWSVJNu55',database='NhpDbfVMT3',port='3306')
                except :
                    tkinter.messagebox.showerror('etax-2019','Failed to connect server, Please contact your administrator')
                
                mycursor=mydb.cursor()
                query = ("SELECT idnumber, meternumber, wardnumber, name, housetax, healthtax, lighttax, watertax, total ,rest FROM kalamwadiinfo")
                mycursor.execute(query)
                for(idnumber, meternumber, wardnumber, name, housetax, healthtax, lighttax, watertax, total, rest) in mycursor:
                    s="{}     {}      {}      {}      {}      {}       {}     {}    {}".format(idnumber, meternumber, wardnumber, housetax, healthtax, lighttax, watertax, total, rest)
                    self.taxbox.insert(0,s)


            def clear(self):
                self.namebox.delete(0,tk.END)
                self.taxbox.delete(0,tk.END)





            def submits(self):
                a=str(self.txt_idnumber.get());
                b=str(self.txt_name.get());
                c=str(self.txt_meternumber.get());
                d=str(self.txt_wardnumber.get());
                e=str(self.txt_house.get());
                f=str(self.txt_health.get());
                g=str(self.txt_light.get());
                h=str(self.txt_water.get());
                a1=float(e)
                b1=float(f)
                c1=float(g)
                d1=float(h)
                i=a1+b1+c1+d1
                t=int(i)
                i=str(i)
                paidrest=str(self.txt_reciept.get());
                k=str(self.txt_housepaid.get());
                l=str(self.txt_healthpaid.get());
                m=str(self.txt_lightpaid.get());
                n=str(self.txt_waterpaid.get());
                n1=float(n)
                k1=float(k)
                l1=float(l)
                m1=float(m)
                paidrest1=float(paidrest)
                o=n1+k1+l1+m1
                u=int(o)
                o=str(o)
                email=str(self.txt_findid_10.get());
                if (email==""):
                    email="kulkarnipranesh1767@gmail.com"

                localtime=str(time.asctime(time.localtime(time.time())))


                totalx=float(u)
                p=str(t-u)
                mydb=mysql.connector.connect(host='remotemysql.com',user='NhpDbfVMT3',password='NXWSVJNu55',database='NhpDbfVMT3',port='3306')
                mycursor=mydb.cursor()
                query = ("SELECT rest FROM kalamwadiinfo where idnumber ="+a)
                mycursor.execute(query)
                for(rest) in mycursor:
                    privrest="{}".format(rest)
                    privrest=privrest[2:len(privrest)-3]
                    privrest=float(privrest)
                    privrest = privrest-paidrest1-totalx
                    privrest=str(privrest)


                mydb=mysql.connector.connect(host='remotemysql.com',user='NhpDbfVMT3',password='NXWSVJNu55',database='NhpDbfVMT3',port='3306')
                mycursor=mydb.cursor()
                query = ("SELECT rec FROM kalamwadirec")
                mycursor.execute(query)
                for(rec) in mycursor:
                    j=str(rec)
                    j=j[1:len(j)-2]
                    rec=int(j)

                st=rec+1


                mycursor=mydb.cursor()
                query=("update kalamwadirec set rec = %s" %(st))
                mycursor.execute(query)
                mydb.commit()


                if (a=="") or (b=="") or (c=="") or (d=="") or (e=="") or (f=="") or (g=="") or (h=="") or (i=="") or (j=="") or (k=="") or (l=="") or (m=="") or (n=="") or (o=="") :
                    tkinter.messagebox.showerror("eTAX-2019","Kindly Fill All Details")
                else :
                    database = DBConnect(host='remotemysql.com',user='NhpDbfVMT3',password='NXWSVJNu55',database='NhpDbfVMT3',port='3306')
                    new_user = {'idnumber': a,'meternumber': c,'wardnumber': d,'name': b,'housetax': e,'healthtax': f,'lighttax': g,'watertax': h,'total': i,'reciptnumber':j,'housetaxpaid':k,'healthtaxpaid':l,'lighttaxpaid':m,'watertaxpaid':n,'totalpaid':o,'rest':p,'datei': localtime}
                    database.insert(new_user,'kalamwadi')
                    mydb=mysql.connector.connect(host='remotemysql.com',user='NhpDbfVMT3',password='NXWSVJNu55',database='NhpDbfVMT3',port='3306')
                    mycursor=mydb.cursor()
                    query=("update kalamwadiinfo set rest = %s where idnumber = %s" %(privrest,a))
                    mycursor.execute(query)
                    mydb.commit()
                    tkinter.messagebox.showinfo('etax-2019','Data entered to the database successfully')

                    content1="-------------Grampanchayat Kalamwadi----------------\n\t\tTal. : Walava\n\t\tDist. : Sangli\n\tDate :"+localtime+"\n-----------Tax Collection 2019---------\nIncharge Officer : XYZ\nDetails :\n"
                    content2="\nReciept Number :"+j+"\nName :"+b+"\nID Number :"+a+"\nMeter Number :"+c+"\nWard Number :"+d+"\nHouse Tax :"+e+"\nHealth Tax :"+f+"\nLight Tax :"+g+"\nWater Tax :"+h+"\nTotal Ammount of tax to be paid :"+i+"\nPaid House Tax :"+k+"\nPaid Health Tax :"+l+"\nPaid Light Tax :"+m+"\nPiad Water Tax :"+n+"\nTotal Tax paid  :"+o+"\nPrevious Tax Ammount Paid  :"+paidrest+"\n\n\nThank You..........."
                    totcontaint=content1+content2 


                    mail= smtplib.SMTP('smtp.gmail.com',587)
                    mail.ehlo()
                    mail.starttls()
                    mail.login('etaxsupp2019@gmail.com','Pass@123')
                    mail.sendmail('etaxsupp2019@gmail.com',email,totcontaint)
                    mail.close()
                    messagebox.showinfo("etax-2019","Reciept Delivered successfully...............")

                    root.destroy()
                    openworkspace1()





            def finds(self):
                id=str(self.txt_findid.get());
                if id=="":
                    tkinter.messagebox.showerror('eTAX-2019','Please Enter id number')
                mydb=mysql.connector.connect(host='remotemysql.com',user='NhpDbfVMT3',password='NXWSVJNu55',database='NhpDbfVMT3',port='3306')
                mycursor=mydb.cursor()
                query = ("SELECT idnumber, meternumber, wardnumber, name, housetax, healthtax, lighttax, watertax, total, rest FROM kalamwadiinfo WHERE idnumber = "+id)
                mycursor.execute(query)
                for(idnumber, meternumber, wardnumber, name, housetax, healthtax, lighttax, watertax, total, rest) in mycursor:
                    s="{}     {}      {}      {}      {}      {}       {}     {}    {}".format(idnumber, meternumber, wardnumber, housetax, healthtax, lighttax, watertax, total, rest)
                    self.taxbox.insert(0,s)


                mydb=mysql.connector.connect(host='remotemysql.com',user='NhpDbfVMT3',password='NXWSVJNu55',database='NhpDbfVMT3',port='3306')
                mycursor=mydb.cursor()
                query = ("SELECT idnumber, meternumber, wardnumber, name, housetax, healthtax, lighttax, watertax, total FROM kalamwadiinfo WHERE idnumber = "+id)
                mycursor.execute(query)
                for(idnumber, meternumber, wardnumber, name, housetax, healthtax, lighttax, watertax, total) in mycursor:
                    s="{}   {}".format(idnumber, name)
                    self.namebox.insert(0,s)



            def backs(self):
                root.destroy()
                main()

            def __init__(self, top=None):
                '''This class configures and populates the toplevel window.
                   top is the toplevel containing window.'''
                _bgcolor = '#d9d9d9'  # X11 color: 'gray85'
                _fgcolor = '#000000'  # X11 color: 'black'
                _compcolor = '#d9d9d9' # X11 color: 'gray85'
                _ana1color = '#d9d9d9' # X11 color: 'gray85'
                _ana2color = '#ececec' # Closest X11 color: 'gray92'
                self.style = ttk.Style()
                if sys.platform == "win32":
                    self.style.theme_use('winnative')
                self.style.configure('.',background=_bgcolor)
                self.style.configure('.',foreground=_fgcolor)
                self.style.configure('.',font="TkDefaultFont")
                self.style.map('.',background=
                    [('selected', _compcolor), ('active',_ana2color)])

                top.geometry("1538x862+-22+0")
                top.title("New Toplevel")
                top.configure(background="#ffff24")
                top.configure(highlightbackground="#d9d9d9")
                top.configure(highlightcolor="black")

                self.Label1 = tk.Label(top)
                self.Label1.place(relx=0.013, rely=0.023, height=81, width=156)
                self.Label1.configure(activebackground="#f9f9f9")
                self.Label1.configure(activeforeground="black")
                self.Label1.configure(background="#ffff24")
                self.Label1.configure(disabledforeground="#a3a3a3")
                self.Label1.configure(font="-family {Britannic Bold} -size 48 -weight bold")
                self.Label1.configure(foreground="#ff250d")
                self.Label1.configure(highlightbackground="#d9d9d9")
                self.Label1.configure(highlightcolor="black")
                self.Label1.configure(text='''eTAX''')

                self.Label1_1 = tk.Label(top)
                self.Label1_1.place(relx=0.117, rely=0.023, height=81, width=156)
                self.Label1_1.configure(activebackground="#f9f9f9")
                self.Label1_1.configure(activeforeground="black")
                self.Label1_1.configure(background="#ffff24")
                self.Label1_1.configure(disabledforeground="#a3a3a3")
                self.Label1_1.configure(font="-family {Britannic Bold} -size 48 -weight bold")
                self.Label1_1.configure(foreground="#2212ff")
                self.Label1_1.configure(highlightbackground="#d9d9d9")
                self.Label1_1.configure(highlightcolor="black")
                self.Label1_1.configure(text='''2019''')

                self.Label2 = tk.Label(top)
                self.Label2.place(relx=0.072, rely=0.104, height=31, width=141)
                self.Label2.configure(activebackground="#f9f9f9")
                self.Label2.configure(activeforeground="black")
                self.Label2.configure(background="#ffff24")
                self.Label2.configure(disabledforeground="#a3a3a3")
                self.Label2.configure(font="-family {Segoe Script} -size 12 -slant italic")
                self.Label2.configure(foreground="#13c12a")
                self.Label2.configure(highlightbackground="#d9d9d9")
                self.Label2.configure(highlightcolor="black")
                self.Label2.configure(text='''working for you''')

                self.backbutton = tk.Button(top)
                self.backbutton.place(relx=0.013, rely=0.162, height=44, width=97)
                self.backbutton.configure(activebackground="#ececec")
                self.backbutton.configure(activeforeground="#000000")
                self.backbutton.configure(background="#120bd8")
                self.backbutton.configure(disabledforeground="#a3a3a3")
                self.backbutton.configure(font="-family {Rockwell Extra Bold} -size 12 -weight bold")
                self.backbutton.configure(foreground="#fcffff")
                self.backbutton.configure(highlightbackground="#d9d9d9")
                self.backbutton.configure(highlightcolor="black")
                self.backbutton.configure(pady="0")
                self.backbutton.configure(text='''Back''',command=self.backs)

                self.exit = tk.Button(top)
                self.exit.place(relx=0.104, rely=0.162, height=44, width=97)
                self.exit.configure(activebackground="#ececec")
                self.exit.configure(activeforeground="#000000")
                self.exit.configure(background="#120bd8")
                self.exit.configure(disabledforeground="#a3a3a3")
                self.exit.configure(font="-family {Rockwell Extra Bold} -size 12 -weight bold")
                self.exit.configure(foreground="#fcffff")
                self.exit.configure(highlightbackground="#d9d9d9")
                self.exit.configure(highlightcolor="black")
                self.exit.configure(pady="0")
                self.exit.configure(text='''Exit''',command= self.exits)

                self.Label3 = tk.Label(top)
                self.Label3.place(relx=0.013, rely=0.916, height=21, width=56)
                self.Label3.configure(activebackground="#f9f9f9")
                self.Label3.configure(activeforeground="black")
                self.Label3.configure(background="#ffff24")
                self.Label3.configure(disabledforeground="#a3a3a3")
                self.Label3.configure(foreground="#000000")
                self.Label3.configure(highlightbackground="#d9d9d9")
                self.Label3.configure(highlightcolor="black")
                self.Label3.configure(text='''etax-2019''')

                self.Label3_3 = tk.Label(top)
                self.Label3_3.place(relx=0.013, rely=0.94, height=21, width=34)
                self.Label3_3.configure(activebackground="#f9f9f9")
                self.Label3_3.configure(activeforeground="black")
                self.Label3_3.configure(background="#ffff24")
                self.Label3_3.configure(disabledforeground="#a3a3a3")
                self.Label3_3.configure(foreground="#000000")
                self.Label3_3.configure(highlightbackground="#d9d9d9")
                self.Label3_3.configure(highlightcolor="black")
                self.Label3_3.configure(text='''v 1.0.2''')

                self.Label3_4 = tk.Label(top)
                self.Label3_4.place(relx=0.007, rely=0.986, height=21, width=134)
                self.Label3_4.configure(activebackground="#f9f9f9")
                self.Label3_4.configure(activeforeground="black")
                self.Label3_4.configure(background="#ffff24")
                self.Label3_4.configure(disabledforeground="#a3a3a3")
                self.Label3_4.configure(foreground="#000000")
                self.Label3_4.configure(highlightbackground="#d9d9d9")
                self.Label3_4.configure(highlightcolor="black")
                self.Label3_4.configure(text='''Working On Windows''')

                self.Label3_1 = tk.Label(top)
                self.Label3_1.place(relx=0.013, rely=0.963, height=21, width=164)
                self.Label3_1.configure(activebackground="#f9f9f9")
                self.Label3_1.configure(activeforeground="black")
                self.Label3_1.configure(background="#ffff24")
                self.Label3_1.configure(disabledforeground="#a3a3a3")
                self.Label3_1.configure(foreground="#000000")
                self.Label3_1.configure(highlightbackground="#d9d9d9")
                self.Label3_1.configure(highlightcolor="black")
                self.Label3_1.configure(text='''Connected to MySQL server 8.0''')

                self.Label4 = tk.Label(top)
                self.Label4.place(relx=0.39, rely=0.023, height=68, width=361)
                self.Label4.configure(activebackground="#f9f9f9")
                self.Label4.configure(activeforeground="black")
                self.Label4.configure(background="#ffff24")
                self.Label4.configure(disabledforeground="#36911a")
                self.Label4.configure(font="-family {Rockwell Extra Bold} -size 40 -weight bold")
                self.Label4.configure(foreground="#36911a")
                self.Label4.configure(highlightbackground="#d9d9d9")
                self.Label4.configure(highlightcolor="black")
                self.Label4.configure(text='''Workspace''')

                self.Label5 = tk.Label(top)
                self.Label5.place(relx=0.793, rely=0.035, height=28, width=192)
                self.Label5.configure(activebackground="#f9f9f9")
                self.Label5.configure(activeforeground="black")
                self.Label5.configure(background="#ffff24")
                self.Label5.configure(disabledforeground="#a3a3a3")
                self.Label5.configure(font="-family {Rockwell} -size 15")
                self.Label5.configure(foreground="#000000")
                self.Label5.configure(highlightbackground="#d9d9d9")
                self.Label5.configure(highlightcolor="black")
                self.Label5.configure(text='''Village : Kalamwadi''')

                self.Label5_2 = tk.Label(top)
                self.Label5_2.place(relx=0.813, rely=0.07, height=28, width=172)
                self.Label5_2.configure(activebackground="#f9f9f9")
                self.Label5_2.configure(activeforeground="black")
                self.Label5_2.configure(background="#ffff24")
                self.Label5_2.configure(disabledforeground="#a3a3a3")
                self.Label5_2.configure(font="-family {Rockwell} -size 15")
                self.Label5_2.configure(foreground="#000000")
                self.Label5_2.configure(highlightbackground="#d9d9d9")
                self.Label5_2.configure(highlightcolor="black")
                self.Label5_2.configure(text='''District : Sangli''')

                self.Label5_3 = tk.Label(top)
                self.Label5_3.place(relx=0.897, rely=0.94, height=28, width=172)
                self.Label5_3.configure(activebackground="#f9f9f9")
                self.Label5_3.configure(activeforeground="black")
                self.Label5_3.configure(background="#ffff24")
                self.Label5_3.configure(disabledforeground="#a3a3a3")
                self.Label5_3.configure(font="-family {Rockwell} -size 9")
                self.Label5_3.configure(foreground="#000000")
                self.Label5_3.configure(highlightbackground="#d9d9d9")
                self.Label5_3.configure(highlightcolor="black")
                self.Label5_3.configure(text='''Server  Status : Online''')

                self.Label5_4 = tk.Label(top)
                self.Label5_4.place(relx=0.904, rely=0.963, height=28, width=172)
                self.Label5_4.configure(activebackground="#f9f9f9")
                self.Label5_4.configure(activeforeground="black")
                self.Label5_4.configure(background="#ffff24")
                self.Label5_4.configure(disabledforeground="#a3a3a3")
                self.Label5_4.configure(font="-family {Rockwell} -size 9")
                self.Label5_4.configure(foreground="#000000")
                self.Label5_4.configure(highlightbackground="#d9d9d9")
                self.Label5_4.configure(highlightcolor="black")
                self.Label5_4.configure(text='''Host : localhost''')

                self.Label5_5 = tk.Label(top)
                self.Label5_5.place(relx=0.904, rely=0.986, height=28, width=172)
                self.Label5_5.configure(activebackground="#f9f9f9")
                self.Label5_5.configure(activeforeground="black")
                self.Label5_5.configure(background="#ffff24")
                self.Label5_5.configure(disabledforeground="#a3a3a3")
                self.Label5_5.configure(font="-family {Rockwell} -size 9")
                self.Label5_5.configure(foreground="#000000")
                self.Label5_5.configure(highlightbackground="#d9d9d9")
                self.Label5_5.configure(highlightcolor="black")
                self.Label5_5.configure(text='''Port : 3306''')

                self.Label5_1 = tk.Label(top)
                self.Label5_1.place(relx=0.91, rely=0.093, height=28, width=172)
                self.Label5_1.configure(activebackground="#f9f9f9")
                self.Label5_1.configure(activeforeground="black")
                self.Label5_1.configure(background="#ffff24")
                self.Label5_1.configure(disabledforeground="#a3a3a3")
                self.Label5_1.configure(font="-family {Rockwell} -size 12")
                self.Label5_1.configure(foreground="#000000")
                self.Label5_1.configure(highlightbackground="#d9d9d9")
                self.Label5_1.configure(highlightcolor="black")
                self.Label5_1.configure(text='''User : user''')

                self.namebox = ScrolledListBox(top)
                self.namebox.place(relx=0.403, rely=0.278, relheight=0.621
                        , relwidth=0.248)
                self.namebox.configure(background="white")
                self.namebox.configure(disabledforeground="#a3a3a3")
                self.namebox.configure(font="TkFixedFont")
                self.namebox.configure(foreground="black")
                self.namebox.configure(highlightbackground="#d9d9d9")
                self.namebox.configure(highlightcolor="#d9d9d9")
                self.namebox.configure(selectbackground="#c4c4c4")
                self.namebox.configure(selectforeground="black")
                self.namebox.configure(width=10)

                self.taxbox = ScrolledListBox(top)
                self.taxbox.place(relx=0.65, rely=0.278, relheight=0.621, relwidth=0.339)

                self.taxbox.configure(background="white")
                self.taxbox.configure(disabledforeground="#a3a3a3")
                self.taxbox.configure(font="TkFixedFont")
                self.taxbox.configure(foreground="black")
                self.taxbox.configure(highlightbackground="#d9d9d9")
                self.taxbox.configure(highlightcolor="#d9d9d9")
                self.taxbox.configure(selectbackground="#c4c4c4")
                self.taxbox.configure(selectforeground="black")
                self.taxbox.configure(width=10)

                self.TSeparator1 = ttk.Separator(top)
                self.TSeparator1.place(relx=0.923, rely=0.012, relheight=0.116)
                self.TSeparator1.configure(orient="vertical")

                self.TSeparator2 = ttk.Separator(top)
                self.TSeparator2.place(relx=0.013, rely=0.139, relwidth=0.202)

                self.TSeparator3 = ttk.Separator(top)
                self.TSeparator3.place(relx=0.013, rely=0.232, relwidth=0.975)

                self.TSeparator3_6 = ttk.Separator(top)
                self.TSeparator3_6.place(relx=0.013, rely=0.911, relwidth=0.975)

                self.viewbutton = tk.Button(top)
                self.viewbutton.place(relx=0.494, rely=0.94, height=33, width=148)
                self.viewbutton.configure(activebackground="#ececec")
                self.viewbutton.configure(activeforeground="#000000")
                self.viewbutton.configure(background="#2020d8")
                self.viewbutton.configure(disabledforeground="#a3a3a3")
                self.viewbutton.configure(font="-family {Rockwell} -size 13 -weight bold")
                self.viewbutton.configure(foreground="#ffffff")
                self.viewbutton.configure(highlightbackground="#d9d9d9")
                self.viewbutton.configure(highlightcolor="black")
                self.viewbutton.configure(pady="0")
                self.viewbutton.configure(takefocus="0")
                self.viewbutton.configure(text='''View all Names''',command=self.viewnames)

                self.viewallbutton = tk.Button(top)
                self.viewallbutton.place(relx=0.683, rely=0.94, height=33, width=148)
                self.viewallbutton.configure(activebackground="#ececec")
                self.viewallbutton.configure(activeforeground="#000000")
                self.viewallbutton.configure(background="#2020d8")
                self.viewallbutton.configure(disabledforeground="#a3a3a3")
                self.viewallbutton.configure(font="-family {Rockwell} -size 13 -weight bold")
                self.viewallbutton.configure(foreground="#ffffff")
                self.viewallbutton.configure(highlightbackground="#d9d9d9")
                self.viewallbutton.configure(highlightcolor="black")
                self.viewallbutton.configure(pady="0")
                self.viewallbutton.configure(takefocus="0")
                self.viewallbutton.configure(text='''View all Data''',command=self.viewalls)

                self.clearall_button = tk.Button(top)
                self.clearall_button.place(relx=0.8, rely=0.94, height=33, width=108)
                self.clearall_button.configure(activebackground="#ececec")
                self.clearall_button.configure(activeforeground="#000000")
                self.clearall_button.configure(background="#2020d8")
                self.clearall_button.configure(disabledforeground="#a3a3a3")
                self.clearall_button.configure(font="-family {Rockwell} -size 13 -weight bold")
                self.clearall_button.configure(foreground="#ffffff")
                self.clearall_button.configure(highlightbackground="#d9d9d9")
                self.clearall_button.configure(highlightcolor="black")
                self.clearall_button.configure(pady="0")
                self.clearall_button.configure(takefocus="0")
                self.clearall_button.configure(text='''Clear all''',command=self.clear)

                self.Label6 = tk.Label(top)
                self.Label6.place(relx=0.949, rely=0.035, height=44, width=44)
                self.Label6.configure(activebackground="#f9f9f9")
                self.Label6.configure(activeforeground="black")
                self.Label6.configure(background="#d9d9d9")
                self.Label6.configure(disabledforeground="#a3a3a3")
                self.Label6.configure(foreground="#000000")
                self.Label6.configure(highlightbackground="#d9d9d9")
                self.Label6.configure(highlightcolor="black")
                photo_location = os.path.join(prog_location,"../view database/original/login3.png")
                self._img0 = tk.PhotoImage(file=photo_location)
                self.Label6.configure(image=self._img0)
                self.Label6.configure(text='''Label''')

                self.Frame1 = tk.Frame(top)
                self.Frame1.place(relx=0.0, rely=0.300, relheight=0.493, relwidth=0.198)
                self.Frame1.configure(relief='ridge')
                self.Frame1.configure(borderwidth="10")
                self.Frame1.configure(relief='ridge')
                self.Frame1.configure(background="#50e82a")
                self.Frame1.configure(highlightbackground="#d9d9d9")
                self.Frame1.configure(highlightcolor="black")

                self.Label7 = tk.Label(self.Frame1)
                self.Label7.place(relx=0.066, rely=0.071, height=39, width=106)
                self.Label7.configure(activebackground="#f9f9f9")
                self.Label7.configure(activeforeground="black")
                self.Label7.configure(background="#50e82a")
                self.Label7.configure(disabledforeground="#a3a3a3")
                self.Label7.configure(font="-family {Plantagenet Cherokee} -size 13")
                self.Label7.configure(foreground="#000000")
                self.Label7.configure(highlightbackground="#d9d9d9")
                self.Label7.configure(highlightcolor="black")
                self.Label7.configure(text='''ID Number :''')

                self.txt_idnumber = tk.Entry(self.Frame1)
                self.txt_idnumber.place(relx=0.426, rely=0.094, height=20
                        , relwidth=0.538)
                self.txt_idnumber.configure(background="white")
                self.txt_idnumber.configure(disabledforeground="#a3a3a3")
                self.txt_idnumber.configure(font="TkFixedFont")
                self.txt_idnumber.configure(foreground="#000000")
                self.txt_idnumber.configure(highlightbackground="#d9d9d9")
                self.txt_idnumber.configure(highlightcolor="black")
                self.txt_idnumber.configure(insertbackground="black")
                self.txt_idnumber.configure(selectbackground="#c4c4c4")
                self.txt_idnumber.configure(selectforeground="black")

                self.Label7_1 = tk.Label(self.Frame1)
                self.Label7_1.place(relx=0.033, rely=0.165, height=39, width=106)
                self.Label7_1.configure(activebackground="#f9f9f9")
                self.Label7_1.configure(activeforeground="black")
                self.Label7_1.configure(background="#50e82a")
                self.Label7_1.configure(disabledforeground="#a3a3a3")
                self.Label7_1.configure(font="-family {Plantagenet Cherokee} -size 13")
                self.Label7_1.configure(foreground="#000000")
                self.Label7_1.configure(highlightbackground="#d9d9d9")
                self.Label7_1.configure(highlightcolor="black")
                self.Label7_1.configure(text='''Name :''')

                self.Label7_2 = tk.Label(self.Frame1)
                self.Label7_2.place(relx=0.033, rely=0.282, height=29, width=116)
                self.Label7_2.configure(activebackground="#f9f9f9")
                self.Label7_2.configure(activeforeground="black")
                self.Label7_2.configure(background="#50e82a")
                self.Label7_2.configure(disabledforeground="#a3a3a3")
                self.Label7_2.configure(font="-family {Plantagenet Cherokee} -size 13")
                self.Label7_2.configure(foreground="#000000")
                self.Label7_2.configure(highlightbackground="#d9d9d9")
                self.Label7_2.configure(highlightcolor="black")
                self.Label7_2.configure(text='''Meter Number :''')

                self.Label7_3 = tk.Label(self.Frame1)
                self.Label7_3.place(relx=0.033, rely=0.353, height=39, width=116)
                self.Label7_3.configure(activebackground="#f9f9f9")
                self.Label7_3.configure(activeforeground="black")
                self.Label7_3.configure(background="#50e82a")
                self.Label7_3.configure(disabledforeground="#a3a3a3")
                self.Label7_3.configure(font="-family {Plantagenet Cherokee} -size 13")
                self.Label7_3.configure(foreground="#000000")
                self.Label7_3.configure(highlightbackground="#d9d9d9")
                self.Label7_3.configure(highlightcolor="black")
                self.Label7_3.configure(text='''Ward Number :''')

                self.Label7_4 = tk.Label(self.Frame1)
                self.Label7_4.place(relx=0.033, rely=0.447, height=39, width=106)
                self.Label7_4.configure(activebackground="#f9f9f9")
                self.Label7_4.configure(activeforeground="black")
                self.Label7_4.configure(background="#50e82a")
                self.Label7_4.configure(disabledforeground="#a3a3a3")
                self.Label7_4.configure(font="-family {Plantagenet Cherokee} -size 13")
                self.Label7_4.configure(foreground="#000000")
                self.Label7_4.configure(highlightbackground="#d9d9d9")
                self.Label7_4.configure(highlightcolor="black")
                self.Label7_4.configure(text='''House Tax :''')

                self.Label7_5 = tk.Label(self.Frame1)
                self.Label7_5.place(relx=0.033, rely=0.541, height=39, width=106)
                self.Label7_5.configure(activebackground="#f9f9f9")
                self.Label7_5.configure(activeforeground="black")
                self.Label7_5.configure(background="#50e82a")
                self.Label7_5.configure(disabledforeground="#a3a3a3")
                self.Label7_5.configure(font="-family {Plantagenet Cherokee} -size 13")
                self.Label7_5.configure(foreground="#000000")
                self.Label7_5.configure(highlightbackground="#d9d9d9")
                self.Label7_5.configure(highlightcolor="black")
                self.Label7_5.configure(text='''Health Tax :''')

                self.Label7_6 = tk.Label(self.Frame1)
                self.Label7_6.place(relx=0.033, rely=0.635, height=39, width=106)
                self.Label7_6.configure(activebackground="#f9f9f9")
                self.Label7_6.configure(activeforeground="black")
                self.Label7_6.configure(background="#50e82a")
                self.Label7_6.configure(disabledforeground="#a3a3a3")
                self.Label7_6.configure(font="-family {Plantagenet Cherokee} -size 13")
                self.Label7_6.configure(foreground="#000000")
                self.Label7_6.configure(highlightbackground="#d9d9d9")
                self.Label7_6.configure(highlightcolor="black")
                self.Label7_6.configure(text='''Light Tax :''')

                self.Label7_7 = tk.Label(self.Frame1)
                self.Label7_7.place(relx=0.033, rely=0.729, height=39, width=106)
                self.Label7_7.configure(activebackground="#f9f9f9")
                self.Label7_7.configure(activeforeground="black")
                self.Label7_7.configure(background="#50e82a")
                self.Label7_7.configure(disabledforeground="#a3a3a3")
                self.Label7_7.configure(font="-family {Plantagenet Cherokee} -size 13")
                self.Label7_7.configure(foreground="#000000")
                self.Label7_7.configure(highlightbackground="#d9d9d9")
                self.Label7_7.configure(highlightcolor="black")
                self.Label7_7.configure(text='''Water Tax :''')

                self.txt_name = tk.Entry(self.Frame1)
                self.txt_name.place(relx=0.426, rely=0.188,height=20, relwidth=0.538)
                self.txt_name.configure(background="white")
                self.txt_name.configure(disabledforeground="#a3a3a3")
                self.txt_name.configure(font="TkFixedFont")
                self.txt_name.configure(foreground="#000000")
                self.txt_name.configure(highlightbackground="#d9d9d9")
                self.txt_name.configure(highlightcolor="black")
                self.txt_name.configure(insertbackground="black")
                self.txt_name.configure(selectbackground="#c4c4c4")
                self.txt_name.configure(selectforeground="black")

                self.txt_meternumber = tk.Entry(self.Frame1)
                self.txt_meternumber.place(relx=0.426, rely=0.282, height=20
                        , relwidth=0.505)
                self.txt_meternumber.configure(background="white")
                self.txt_meternumber.configure(disabledforeground="#a3a3a3")
                self.txt_meternumber.configure(font="TkFixedFont")
                self.txt_meternumber.configure(foreground="#000000")
                self.txt_meternumber.configure(highlightbackground="#d9d9d9")
                self.txt_meternumber.configure(highlightcolor="black")
                self.txt_meternumber.configure(insertbackground="black")
                self.txt_meternumber.configure(selectbackground="#c4c4c4")
                self.txt_meternumber.configure(selectforeground="black")

                self.txt_wardnumber = tk.Entry(self.Frame1)
                self.txt_wardnumber.place(relx=0.426, rely=0.353, height=20
                        , relwidth=0.505)
                self.txt_wardnumber.configure(background="white")
                self.txt_wardnumber.configure(disabledforeground="#a3a3a3")
                self.txt_wardnumber.configure(font="TkFixedFont")
                self.txt_wardnumber.configure(foreground="#000000")
                self.txt_wardnumber.configure(highlightbackground="#d9d9d9")
                self.txt_wardnumber.configure(highlightcolor="black")
                self.txt_wardnumber.configure(insertbackground="black")
                self.txt_wardnumber.configure(selectbackground="#c4c4c4")
                self.txt_wardnumber.configure(selectforeground="black")

                self.txt_house = tk.Entry(self.Frame1)
                self.txt_house.place(relx=0.393, rely=0.471,height=20, relwidth=0.538)
                self.txt_house.configure(background="white")
                self.txt_house.configure(disabledforeground="#a3a3a3")
                self.txt_house.configure(font="TkFixedFont")
                self.txt_house.configure(foreground="#000000")
                self.txt_house.configure(highlightbackground="#d9d9d9")
                self.txt_house.configure(highlightcolor="black")
                self.txt_house.configure(insertbackground="black")
                self.txt_house.configure(selectbackground="#c4c4c4")
                self.txt_house.configure(selectforeground="black")

                self.txt_light = tk.Entry(self.Frame1)
                self.txt_light.place(relx=0.393, rely=0.659,height=20, relwidth=0.538)
                self.txt_light.configure(background="white")
                self.txt_light.configure(disabledforeground="#a3a3a3")
                self.txt_light.configure(font="TkFixedFont")
                self.txt_light.configure(foreground="#000000")
                self.txt_light.configure(highlightbackground="#d9d9d9")
                self.txt_light.configure(highlightcolor="black")
                self.txt_light.configure(insertbackground="black")
                self.txt_light.configure(selectbackground="#c4c4c4")
                self.txt_light.configure(selectforeground="black")

                self.txt_health = tk.Entry(self.Frame1)
                self.txt_health.place(relx=0.393, rely=0.565,height=20, relwidth=0.538)
                self.txt_health.configure(background="white")
                self.txt_health.configure(disabledforeground="#a3a3a3")
                self.txt_health.configure(font="TkFixedFont")
                self.txt_health.configure(foreground="#000000")
                self.txt_health.configure(highlightbackground="#d9d9d9")
                self.txt_health.configure(highlightcolor="black")
                self.txt_health.configure(insertbackground="black")
                self.txt_health.configure(selectbackground="#c4c4c4")
                self.txt_health.configure(selectforeground="black")

                self.txt_water = tk.Entry(self.Frame1)
                self.txt_water.place(relx=0.393, rely=0.753,height=20, relwidth=0.538)
                self.txt_water.configure(background="white")
                self.txt_water.configure(disabledforeground="#a3a3a3")
                self.txt_water.configure(font="TkFixedFont")
                self.txt_water.configure(foreground="#000000")
                self.txt_water.configure(highlightbackground="#d9d9d9")
                self.txt_water.configure(highlightcolor="black")
                self.txt_water.configure(insertbackground="black")
                self.txt_water.configure(selectbackground="#c4c4c4")
                self.txt_water.configure(selectforeground="black")

                self.Frame1_20 = tk.Frame(top)
                self.Frame1_20.place(relx=0.202, rely=0.267, relheight=0.4
                        , relwidth=0.198)
                self.Frame1_20.configure(relief='ridge')
                self.Frame1_20.configure(borderwidth="10")
                self.Frame1_20.configure(relief='ridge')
                self.Frame1_20.configure(background="#50e82a")
                self.Frame1_20.configure(highlightbackground="#d9d9d9")
                self.Frame1_20.configure(highlightcolor="black")

                self.Label7_21 = tk.Label(self.Frame1_20)
                self.Label7_21.place(relx=0.033, rely=0.087, height=39, width=126)
                self.Label7_21.configure(activebackground="#f9f9f9")
                self.Label7_21.configure(activeforeground="black")
                self.Label7_21.configure(background="#50e82a")
                self.Label7_21.configure(disabledforeground="#a3a3a3")
                self.Label7_21.configure(font="-family {Plantagenet Cherokee} -size 13")
                self.Label7_21.configure(foreground="#000000")
                self.Label7_21.configure(highlightbackground="#d9d9d9")
                self.Label7_21.configure(highlightcolor="black")
                self.Label7_21.configure(text='''Previous Remaining Ammount :''')

                self.txt_reciept = tk.Entry(self.Frame1_20)
                self.txt_reciept.place(relx=0.492, rely=0.116, height=20, relwidth=0.439)

                self.txt_reciept.configure(background="white")
                self.txt_reciept.configure(disabledforeground="#a3a3a3")
                self.txt_reciept.configure(font="TkFixedFont")
                self.txt_reciept.configure(foreground="#000000")
                self.txt_reciept.configure(highlightbackground="#d9d9d9")
                self.txt_reciept.configure(highlightcolor="black")
                self.txt_reciept.configure(insertbackground="black")
                self.txt_reciept.configure(selectbackground="#c4c4c4")
                self.txt_reciept.configure(selectforeground="black")

                self.Label7_1 = tk.Label(self.Frame1_20)
                self.Label7_1.place(relx=0.033, rely=0.203, height=39, width=126)
                self.Label7_1.configure(activebackground="#f9f9f9")
                self.Label7_1.configure(activeforeground="black")
                self.Label7_1.configure(background="#50e82a")
                self.Label7_1.configure(disabledforeground="#a3a3a3")
                self.Label7_1.configure(font="-family {Plantagenet Cherokee} -size 13")
                self.Label7_1.configure(foreground="#000000")
                self.Label7_1.configure(highlightbackground="#d9d9d9")
                self.Label7_1.configure(highlightcolor="black")
                self.Label7_1.configure(text='''Housetax(Paid) :''')

                self.Label7_2 = tk.Label(self.Frame1_20)
                self.Label7_2.place(relx=0.033, rely=0.319, height=39, width=126)
                self.Label7_2.configure(activebackground="#f9f9f9")
                self.Label7_2.configure(activeforeground="black")
                self.Label7_2.configure(background="#50e82a")
                self.Label7_2.configure(disabledforeground="#a3a3a3")
                self.Label7_2.configure(font="-family {Plantagenet Cherokee} -size 13")
                self.Label7_2.configure(foreground="#000000")
                self.Label7_2.configure(highlightbackground="#d9d9d9")
                self.Label7_2.configure(highlightcolor="black")
                self.Label7_2.configure(text='''Healthtax (Paid) :''')

                self.Label7_3 = tk.Label(self.Frame1_20)
                self.Label7_3.place(relx=0.066, rely=0.435, height=39, width=116)
                self.Label7_3.configure(activebackground="#f9f9f9")
                self.Label7_3.configure(activeforeground="black")
                self.Label7_3.configure(background="#50e82a")
                self.Label7_3.configure(disabledforeground="#a3a3a3")
                self.Label7_3.configure(font="-family {Plantagenet Cherokee} -size 13")
                self.Label7_3.configure(foreground="#000000")
                self.Label7_3.configure(highlightbackground="#d9d9d9")
                self.Label7_3.configure(highlightcolor="black")
                self.Label7_3.configure(text='''Lighttax(Paid) :''')

                self.Label7_4 = tk.Label(self.Frame1_20)
                self.Label7_4.place(relx=0.033, rely=0.551, height=39, width=116)
                self.Label7_4.configure(activebackground="#f9f9f9")
                self.Label7_4.configure(activeforeground="black")
                self.Label7_4.configure(background="#50e82a")
                self.Label7_4.configure(disabledforeground="#a3a3a3")
                self.Label7_4.configure(font="-family {Plantagenet Cherokee} -size 13")
                self.Label7_4.configure(foreground="#000000")
                self.Label7_4.configure(highlightbackground="#d9d9d9")
                self.Label7_4.configure(highlightcolor="black")
                self.Label7_4.configure(text='''Watertax(Paid) :''')

                self.txt_housepaid = tk.Entry(self.Frame1_20)
                self.txt_housepaid.place(relx=0.492, rely=0.232, height=20
                        , relwidth=0.439)
                self.txt_housepaid.configure(background="white")
                self.txt_housepaid.configure(disabledforeground="#a3a3a3")
                self.txt_housepaid.configure(font="TkFixedFont")
                self.txt_housepaid.configure(foreground="#000000")
                self.txt_housepaid.configure(highlightbackground="#d9d9d9")
                self.txt_housepaid.configure(highlightcolor="black")
                self.txt_housepaid.configure(insertbackground="black")
                self.txt_housepaid.configure(selectbackground="#c4c4c4")
                self.txt_housepaid.configure(selectforeground="black")

                self.txt_healthpaid = tk.Entry(self.Frame1_20)
                self.txt_healthpaid.place(relx=0.492, rely=0.348, height=20
                        , relwidth=0.439)
                self.txt_healthpaid.configure(background="white")
                self.txt_healthpaid.configure(disabledforeground="#a3a3a3")
                self.txt_healthpaid.configure(font="TkFixedFont")
                self.txt_healthpaid.configure(foreground="#000000")
                self.txt_healthpaid.configure(highlightbackground="#d9d9d9")
                self.txt_healthpaid.configure(highlightcolor="black")
                self.txt_healthpaid.configure(insertbackground="black")
                self.txt_healthpaid.configure(selectbackground="#c4c4c4")
                self.txt_healthpaid.configure(selectforeground="black")

                self.txt_lightpaid = tk.Entry(self.Frame1_20)
                self.txt_lightpaid.place(relx=0.492, rely=0.464, height=20
                        , relwidth=0.439)
                self.txt_lightpaid.configure(background="white")
                self.txt_lightpaid.configure(disabledforeground="#a3a3a3")
                self.txt_lightpaid.configure(font="TkFixedFont")
                self.txt_lightpaid.configure(foreground="#000000")
                self.txt_lightpaid.configure(highlightbackground="#d9d9d9")
                self.txt_lightpaid.configure(highlightcolor="black")
                self.txt_lightpaid.configure(insertbackground="black")
                self.txt_lightpaid.configure(selectbackground="#c4c4c4")
                self.txt_lightpaid.configure(selectforeground="black")

                self.txt_waterpaid = tk.Entry(self.Frame1_20)
                self.txt_waterpaid.place(relx=0.492, rely=0.58, height=20
                        , relwidth=0.439)
                self.txt_waterpaid.configure(background="white")
                self.txt_waterpaid.configure(disabledforeground="#a3a3a3")
                self.txt_waterpaid.configure(font="TkFixedFont")
                self.txt_waterpaid.configure(foreground="#000000")
                self.txt_waterpaid.configure(highlightbackground="#d9d9d9")
                self.txt_waterpaid.configure(highlightcolor="black")
                self.txt_waterpaid.configure(insertbackground="black")
                self.txt_waterpaid.configure(selectbackground="#c4c4c4")
                self.txt_waterpaid.configure(selectforeground="black")

                self.btn_submit = tk.Button(self.Frame1_20)
                self.btn_submit.place(relx=0.328, rely=0.812, height=44, width=97)
                self.btn_submit.configure(activebackground="#ececec")
                self.btn_submit.configure(activeforeground="#000000")
                self.btn_submit.configure(background="#ff350d")
                self.btn_submit.configure(disabledforeground="#a3a3a3")
                self.btn_submit.configure(font="-family {Segoe UI} -size 13")
                self.btn_submit.configure(foreground="#ffffff")
                self.btn_submit.configure(highlightbackground="#d9d9d9")
                self.btn_submit.configure(highlightcolor="black")
                self.btn_submit.configure(pady="0")
                self.btn_submit.configure(text='''SUBMIT''',command= self.submits)

                self.Frame1_18 = tk.Frame(top)
                self.Frame1_18.place(relx=0.202, rely=0.673, relheight=0.226
                        , relwidth=0.198)
                self.Frame1_18.configure(relief='ridge')
                self.Frame1_18.configure(borderwidth="10")
                self.Frame1_18.configure(relief='ridge')
                self.Frame1_18.configure(background="#50e82a")
                self.Frame1_18.configure(highlightbackground="#d9d9d9")
                self.Frame1_18.configure(highlightcolor="black")

                self.Label7_19 = tk.Label(self.Frame1_18)
                self.Label7_19.place(relx=0.098, rely=0.359, height=29, width=256)
                self.Label7_19.configure(activebackground="#f9f9f9")
                self.Label7_19.configure(activeforeground="black")
                self.Label7_19.configure(background="#50e82a")
                self.Label7_19.configure(disabledforeground="#a3a3a3")
                self.Label7_19.configure(font="-family {Plantagenet Cherokee} -size 13")
                self.Label7_19.configure(foreground="#000000")
                self.Label7_19.configure(highlightbackground="#d9d9d9")
                self.Label7_19.configure(highlightcolor="black")
                self.Label7_19.configure(text='''Enter The ID Number Here :''')

                self.txt_findid = tk.Entry(self.Frame1_18)
                self.txt_findid.place(relx=0.098, rely=0.513,height=20, relwidth=0.833)
                self.txt_findid.configure(background="white")
                self.txt_findid.configure(disabledforeground="#a3a3a3")
                self.txt_findid.configure(font="TkFixedFont")
                self.txt_findid.configure(foreground="#000000")
                self.txt_findid.configure(highlightbackground="#d9d9d9")
                self.txt_findid.configure(highlightcolor="black")
                self.txt_findid.configure(insertbackground="black")
                self.txt_findid.configure(selectbackground="#c4c4c4")
                self.txt_findid.configure(selectforeground="black")

                self.Label8 = tk.Label(self.Frame1_18)
                self.Label8.place(relx=0.23, rely=0.103, height=27, width=168)
                self.Label8.configure(activebackground="#f9f9f9")
                self.Label8.configure(activeforeground="black")
                self.Label8.configure(background="#50e82a")
                self.Label8.configure(disabledforeground="#a3a3a3")
                self.Label8.configure(font="-family {Rockwell} -size 14")
                self.Label8.configure(foreground="#000000")
                self.Label8.configure(highlightbackground="#d9d9d9")
                self.Label8.configure(highlightcolor="black")
                self.Label8.configure(text='''-----Find Entry-----''')

                self.Frame2 = tk.Frame(top)
                self.Frame2.place(relx=0.0, rely=0.752, relheight=0.131, relwidth=0.198)
                self.Frame2.configure(relief='ridge')
                self.Frame2.configure(borderwidth="10")
                self.Frame2.configure(relief='ridge')
                self.Frame2.configure(background="#50e82a")
                self.Frame2.configure(width=305)

                self.Label8_8 = tk.Label(self.Frame2)
                self.Label8_8.place(relx=0.23, rely=0.087, height=27, width=168)
                self.Label8_8.configure(activebackground="#f9f9f9")
                self.Label8_8.configure(activeforeground="black")
                self.Label8_8.configure(background="#50e82a")
                self.Label8_8.configure(disabledforeground="#a3a3a3")
                self.Label8_8.configure(font="-family {Rockwell} -size 14")
                self.Label8_8.configure(foreground="#000000")
                self.Label8_8.configure(highlightbackground="#d9d9d9")
                self.Label8_8.configure(highlightcolor="black")
                self.Label8_8.configure(text='''-----Email-----''')

                self.Label7_9 = tk.Label(self.Frame2)
                self.Label7_9.place(relx=0.033, rely=0.261, height=29, width=286)
                self.Label7_9.configure(activebackground="#f9f9f9")
                self.Label7_9.configure(activeforeground="black")
                self.Label7_9.configure(background="#50e82a")
                self.Label7_9.configure(disabledforeground="#a3a3a3")
                self.Label7_9.configure(font="-family {Plantagenet Cherokee} -size 13")
                self.Label7_9.configure(foreground="#000000")
                self.Label7_9.configure(highlightbackground="#d9d9d9")
                self.Label7_9.configure(highlightcolor="black")
                self.Label7_9.configure(text='''Enter The Email  Here :''')
                self.Label7_9.configure(width=286)

                self.txt_findid_10 = tk.Entry(self.Frame2)
                self.txt_findid_10.place(relx=0.066, rely=0.609, height=20
                        , relwidth=0.866)
                self.txt_findid_10.configure(background="white")
                self.txt_findid_10.configure(disabledforeground="#a3a3a3")
                self.txt_findid_10.configure(font="TkFixedFont")
                self.txt_findid_10.configure(foreground="#000000")
                self.txt_findid_10.configure(highlightbackground="#d9d9d9")
                self.txt_findid_10.configure(highlightcolor="black")
                self.txt_findid_10.configure(insertbackground="black")
                self.txt_findid_10.configure(selectbackground="#c4c4c4")
                self.txt_findid_10.configure(selectforeground="black")
                self.txt_findid_10.configure(width=264)

                self.Label9_11 = tk.Label(top)
                self.Label9_11.place(relx=0.475, rely=0.239, height=21, width=64)
                self.Label9_11.configure(activebackground="#f9f9f9")
                self.Label9_11.configure(activeforeground="black")
                self.Label9_11.configure(background="#ffff24")
                self.Label9_11.configure(disabledforeground="#a3a3a3")
                self.Label9_11.configure(foreground="#000000")
                self.Label9_11.configure(highlightbackground="#d9d9d9")
                self.Label9_11.configure(highlightcolor="black")
                self.Label9_11.configure(text='''Name''')

                self.Label9_12 = tk.Label(top)
                self.Label9_12.place(relx=0.65, rely=0.251, height=11, width=54)
                self.Label9_12.configure(activebackground="#f9f9f9")
                self.Label9_12.configure(activeforeground="black")
                self.Label9_12.configure(background="#ffff24")
                self.Label9_12.configure(disabledforeground="#a3a3a3")
                self.Label9_12.configure(foreground="#000000")
                self.Label9_12.configure(highlightbackground="#d9d9d9")
                self.Label9_12.configure(highlightcolor="black")
                self.Label9_12.configure(text='''ID No.''')
                self.Label9_12.configure(width=54)

                self.Label9_13 = tk.Label(top)
                self.Label9_13.place(relx=0.696, rely=0.251, height=11, width=54)
                self.Label9_13.configure(activebackground="#f9f9f9")
                self.Label9_13.configure(activeforeground="black")
                self.Label9_13.configure(background="#ffff24")
                self.Label9_13.configure(disabledforeground="#a3a3a3")
                self.Label9_13.configure(foreground="#000000")
                self.Label9_13.configure(highlightbackground="#d9d9d9")
                self.Label9_13.configure(highlightcolor="black")
                self.Label9_13.configure(text='''Meter No.''')

                self.Label9_13 = tk.Label(top)
                self.Label9_13.place(relx=0.741, rely=0.251, height=11, width=54)
                self.Label9_13.configure(activebackground="#f9f9f9")
                self.Label9_13.configure(activeforeground="black")
                self.Label9_13.configure(background="#ffff24")
                self.Label9_13.configure(disabledforeground="#a3a3a3")
                self.Label9_13.configure(foreground="#000000")
                self.Label9_13.configure(highlightbackground="#d9d9d9")
                self.Label9_13.configure(highlightcolor="black")
                self.Label9_13.configure(text='''Ward No.''')
                self.Label9_13.configure(width=54)

                self.Label9_13 = tk.Label(top)
                self.Label9_13.place(relx=0.78, rely=0.251, height=11, width=64)
                self.Label9_13.configure(activebackground="#f9f9f9")
                self.Label9_13.configure(activeforeground="black")
                self.Label9_13.configure(background="#ffff24")
                self.Label9_13.configure(disabledforeground="#a3a3a3")
                self.Label9_13.configure(foreground="#000000")
                self.Label9_13.configure(highlightbackground="#d9d9d9")
                self.Label9_13.configure(highlightcolor="black")
                self.Label9_13.configure(text='''Houseatax''')
                self.Label9_13.configure(width=64)

                self.Label9_14 = tk.Label(top)
                self.Label9_14.place(relx=0.826, rely=0.251, height=11, width=64)
                self.Label9_14.configure(activebackground="#f9f9f9")
                self.Label9_14.configure(activeforeground="black")
                self.Label9_14.configure(background="#ffff24")
                self.Label9_14.configure(disabledforeground="#a3a3a3")
                self.Label9_14.configure(foreground="#000000")
                self.Label9_14.configure(highlightbackground="#d9d9d9")
                self.Label9_14.configure(highlightcolor="black")
                self.Label9_14.configure(text='''Healthtaxtax''')

                self.Label9_14 = tk.Label(top)
                self.Label9_14.place(relx=0.871, rely=0.251, height=11, width=54)
                self.Label9_14.configure(activebackground="#f9f9f9")
                self.Label9_14.configure(activeforeground="black")
                self.Label9_14.configure(background="#ffff24")
                self.Label9_14.configure(disabledforeground="#a3a3a3")
                self.Label9_14.configure(foreground="#000000")
                self.Label9_14.configure(highlightbackground="#d9d9d9")
                self.Label9_14.configure(highlightcolor="black")
                self.Label9_14.configure(text='''Lighttax''')
                self.Label9_14.configure(width=54)

                self.Label9_14 = tk.Label(top)
                self.Label9_14.place(relx=0.91, rely=0.251, height=11, width=64)
                self.Label9_14.configure(activebackground="#f9f9f9")
                self.Label9_14.configure(activeforeground="black")
                self.Label9_14.configure(background="#ffff24")
                self.Label9_14.configure(disabledforeground="#a3a3a3")
                self.Label9_14.configure(foreground="#000000")
                self.Label9_14.configure(highlightbackground="#d9d9d9")
                self.Label9_14.configure(highlightcolor="black")
                self.Label9_14.configure(text='''Watertax''')

                self.Label9_14 = tk.Label(top)
                self.Label9_14.place(relx=0.949, rely=0.251, height=11, width=64)
                self.Label9_14.configure(activebackground="#f9f9f9")
                self.Label9_14.configure(activeforeground="black")
                self.Label9_14.configure(background="#ffff24")
                self.Label9_14.configure(disabledforeground="#a3a3a3")
                self.Label9_14.configure(foreground="#000000")
                self.Label9_14.configure(highlightbackground="#d9d9d9")
                self.Label9_14.configure(highlightcolor="black")
                self.Label9_14.configure(text='''Total''')

                self.btn_find = tk.Button(self.Frame1_18)
                self.btn_find.place(relx=0.328, rely=0.718, height=34, width=97)
                self.btn_find.configure(activebackground="#ececec")
                self.btn_find.configure(activeforeground="#000000")
                self.btn_find.configure(background="#252ce8")
                self.btn_find.configure(disabledforeground="#a3a3a3")
                self.btn_find.configure(foreground="#ffffff")
                self.btn_find.configure(highlightbackground="#d9d9d9")
                self.btn_find.configure(highlightcolor="black")
                self.btn_find.configure(pady="0")
                self.btn_find.configure(text='''FIND''',command=self.finds)

                self.menubar = tk.Menu(top,font="TkMenuFont",bg=_bgcolor,fg=_fgcolor)
                top.configure(menu = self.menubar)

        # The following code is added to facilitate the Scrolled widgets you specified.
        class AutoScroll(object):
            '''Configure the scrollbars for a widget.'''

            def __init__(self, master):
                #  Rozen. Added the try-except clauses so that this class
                #  could be used for scrolled entry widget for which vertical
                #  scrolling is not supported. 5/7/14.
                try:
                    vsb = ttk.Scrollbar(master, orient='vertical', command=self.yview)
                except:
                    pass
                hsb = ttk.Scrollbar(master, orient='horizontal', command=self.xview)

                #self.configure(yscrollcommand=_autoscroll(vsb),
                #    xscrollcommand=_autoscroll(hsb))
                try:
                    self.configure(yscrollcommand=self._autoscroll(vsb))
                except:
                    pass
                self.configure(xscrollcommand=self._autoscroll(hsb))

                self.grid(column=0, row=0, sticky='nsew')
                try:
                    vsb.grid(column=1, row=0, sticky='ns')
                except:
                    pass
                hsb.grid(column=0, row=1, sticky='ew')

                master.grid_columnconfigure(0, weight=1)
                master.grid_rowconfigure(0, weight=1)

                # Copy geometry methods of master  (taken from ScrolledText.py)
                if py3:
                    methods = tk.Pack.__dict__.keys() | tk.Grid.__dict__.keys() \
                          | tk.Place.__dict__.keys()
                else:
                    methods = tk.Pack.__dict__.keys() + tk.Grid.__dict__.keys() \
                          + tk.Place.__dict__.keys()

                for meth in methods:
                    if meth[0] != '_' and meth not in ('config', 'configure'):
                        setattr(self, meth, getattr(master, meth))

            @staticmethod
            def _autoscroll(sbar):
                '''Hide and show scrollbar as needed.'''
                def wrapped(first, last):
                    first, last = float(first), float(last)
                    if first <= 0 and last >= 1:
                        sbar.grid_remove()
                    else:
                        sbar.grid()
                    sbar.set(first, last)
                return wrapped

            def __str__(self):
                return str(self.master)

        def _create_container(func):
            '''Creates a ttk Frame with a given master, and use this new frame to
            place the scrollbars and the widget.'''
            def wrapped(cls, master, **kw):
                container = ttk.Frame(master)
                container.bind('<Enter>', lambda e: _bound_to_mousewheel(e, container))
                container.bind('<Leave>', lambda e: _unbound_to_mousewheel(e, container))
                return func(cls, container, **kw)
            return wrapped

        class ScrolledListBox(AutoScroll, tk.Listbox):
            '''A standard Tkinter Text widget with scrollbars that will
            automatically show/hide as needed.'''
            @_create_container
            def __init__(self, master, **kw):
                tk.Listbox.__init__(self, master, **kw)
                AutoScroll.__init__(self, master)

        import platform
        def _bound_to_mousewheel(event, widget):
            child = widget.winfo_children()[0]
            if platform.system() == 'Windows' or platform.system() == 'Darwin':
                child.bind_all('<MouseWheel>', lambda e: _on_mousewheel(e, child))
                child.bind_all('<Shift-MouseWheel>', lambda e: _on_shiftmouse(e, child))
            else:
                child.bind_all('<Button-4>', lambda e: _on_mousewheel(e, child))
                child.bind_all('<Button-5>', lambda e: _on_mousewheel(e, child))
                child.bind_all('<Shift-Button-4>', lambda e: _on_shiftmouse(e, child))
                child.bind_all('<Shift-Button-5>', lambda e: _on_shiftmouse(e, child))

        def _unbound_to_mousewheel(event, widget):
            if platform.system() == 'Windows' or platform.system() == 'Darwin':
                widget.unbind_all('<MouseWheel>')
                widget.unbind_all('<Shift-MouseWheel>')
            else:
                widget.unbind_all('<Button-4>')
                widget.unbind_all('<Button-5>')
                widget.unbind_all('<Shift-Button-4>')
                widget.unbind_all('<Shift-Button-5>')

        def _on_mousewheel(event, widget):
            if platform.system() == 'Windows':
                widget.yview_scroll(-1*int(event.delta/120),'units')
            elif platform.system() == 'Darwin':
                widget.yview_scroll(-1*int(event.delta),'units')
            else:
                if event.num == 4:
                    widget.yview_scroll(-1, 'units')
                elif event.num == 5:
                    widget.yview_scroll(1, 'units')

        def _on_shiftmouse(event, widget):
            if platform.system() == 'Windows':
                widget.xview_scroll(-1*int(event.delta/120), 'units')
            elif platform.system() == 'Darwin':
                widget.xview_scroll(-1*int(event.delta), 'units')
            else:
                if event.num == 4:
                    widget.xview_scroll(-1, 'units')
                elif event.num == 5:
                    widget.xview_scroll(1, 'units')

        if __name__ == '__main__':
            vp_start_gui()

















    
































































    def viewentry1():         # view all entries

        import sys
        import tkinter
        


        try:
            import Tkinter as tk
        except ImportError:
            import tkinter as tk

        try:
            import ttk
            py3 = False
        except ImportError:
            import tkinter.ttk as ttk
            py3 = True

        import unknown_support

        def vp_start_gui():
            '''Starting point when module is the main routine.'''
            global val, w, root
            root = tk.Tk()
            top = Toplevel1 (root)
            unknown_support.init(root, top)
            root.mainloop()

        w = None
        def create_Toplevel1(root, *args, **kwargs):
            '''Starting point when module is imported by another program.'''
            global w, w_win, rt
            rt = root
            w = tk.Toplevel (root)
            top = Toplevel1 (w)
            unknown_support.init(w, top, *args, **kwargs)
            return (w, top)

        def destroy_Toplevel1():
            global w
            w.destroy()
            w = None











        def advances1():
            import mysql.connector
            from openpyxl import Workbook
            from reportlab.pdfgen import canvas
            import tkinter
            from tkinter import messagebox

            try:
                import Tkinter as tk
            except ImportError:
                import tkinter as tk

            try:
                import ttk
                py3 = False
            except ImportError:
                import tkinter.ttk as ttk
                py3 = True

            import advamce_support
            import os.path

            def vp_start_gui():
                '''Starting point when module is the main routine.'''
                global val, w, root
                global prog_location
                prog_call = sys.argv[0]
                print ('prog_call = {}'.format(prog_call))
                prog_location = os.path.split(prog_call)[0]
                print ('prog_location = {}'.format(prog_location))
                sys.stdout.flush()
                root = tk.Tk()
                top = Toplevel1 (root)
                advamce_support.init(root, top)
                root.mainloop()

            w = None
            def create_Toplevel1(root, *args, **kwargs):
                '''Starting point when module is imported by another program.'''
                global w, w_win, rt
                global prog_location
                prog_call = sys.argv[0]
                print ('prog_call = {}'.format(prog_call))
                prog_location = os.path.split(prog_call)[0]
                print ('prog_location = {}'.format(prog_location))
                rt = root
                w = tk.Toplevel (root)
                top = Toplevel1 (w)
                advamce_support.init(w, top, *args, **kwargs)
                return (w, top)

            def destroy_Toplevel1():
                global w
                w.destroy()
                w = None

            class Toplevel1:
                def pdfs(self):
                    idd=str(self.Entry1.get());
                    mydb=mysql.connector.connect(host='remotemysql.com',user='NhpDbfVMT3',password='NXWSVJNu55',database='NhpDbfVMT3',port='3306')
                    mycursor=mydb.cursor()
                    query = ("SELECT idnumber, meternumber, wardnumber, name, housetax, healthtax, lighttax, watertax, total, reciptnumber, housetaxpaid, healthtaxpaid, lighttaxpaid, watertaxpaid, totalpaid, rest, datei FROM kalamwadi where idnumber = %s" %(idd))
                    mycursor.execute(query)
                    for(idnumber, meternumber, wardnumber, name, housetax, healthtax, lighttax, watertax, total, reciptnumber, housetaxpaid, healthtaxpaid, lighttaxpaid, watertaxpaid, totalpaid, rest, datei) in mycursor:
                        a="UID Number :                      "+idnumber
                        b="Meter Number :                     "+meternumber
                        z="Ward Number :                      "+wardnumber
                        d="Name :                                  "+name
                        e="Housetax :                             "+housetax
                        f="Healthtax :                            "+healthtax
                        g="Lighttax :                              "+lighttax
                        h="Watertax :                             "+watertax
                        j='Receiptnumber : '+reciptnumber
                        k="Paid Housetax :                     "+housetaxpaid
                        l="Paid Healthtax :                    "+healthtaxpaid
                        m="Paid Lighttax :                      "+lighttaxpaid
                        n="Paid Watertax :                     "+watertaxpaid
                        q="Total Paid Tax :                    "+totalpaid
                        o="Remaining Ammount :                0"+rest
                        p='Date : '+datei
                        x=reciptnumber
                    filename=x+'.pdf'
                    c= canvas.Canvas(filename)
                    c.setFont("Helvetica",25)
                    c.setFillColorRGB(100,0,0)
                    c.drawString(265,800,"eTAX 2019")
                    c.setFont("Helvetica",20)
                    c.setFillColorRGB(0,0,110)
                    c.drawString(165,775,'Village Kalamwadi, District : Sangali')
                    c.setFont("Times-Roman",12)
                    c.setFillColorRGB(0,0,0)
                    c.drawString(385,725,p)
                    c.drawString(385,710,j)
                    c.drawString(385,695,"Incharge Officer : Shridhar Kulkarni")
                    c.line(20,675,600,675)
                    c.drawString(40,655,d)
                    c.drawString(40,640,a)
                    c.drawString(40,625,b)
                    c.drawString(40,610,z)
                    c.drawString(40,595,e)
                    c.drawString(40,580,f)
                    c.drawString(40,565,g)
                    c.drawString(40,550,h)
                    c.line(20,535,600,535)
                    c.drawString(40,520,"---------------PAID AMMOUNT---------------")
                    c.drawString(40,505,k)
                    c.drawString(40,490,l)
                    c.drawString(40,475,m)
                    c.drawString(40,460,n)
                    c.drawString(40,445,q)
                    c.line(20,440,600,440)
                    c.drawString(40,425,o)
                    c.line(20,410,600,410)
                    c.drawString(500,360,"Sign")
                    c.setFont("Helvetica",15)
                    c.line(20,350,600,350)
                    c.drawString(265,320,"Thank You !!!")
                    c.drawImage("image.png",180,80, width=250,height=140,mask=None)
                    c.save()
                    tkinter.messagebox.showinfo('etax-2019','Receipt Created Successfully You Can Get It from directory')
                    os.startfile(filename)


                def getlists(self):
                    mydb=mysql.connector.connect(host='remotemysql.com',user='NhpDbfVMT3',password='NXWSVJNu55',database='NhpDbfVMT3',port='3306')
                    mycursor=mydb.cursor()
                    wb = Workbook()

                    SQL = 'SELECT idnumber,name,total,rest from kalamwadiinfo;'
                    mycursor.execute(SQL)
                    ws = wb.create_sheet(0)
                    ws.title = 'kalamwadi_info'
                    ws.append(mycursor.column_names)
                    for (idnumber,name,total,rest) in mycursor:
                        a=float(rest)
                        b=float(total)
                        if a > b/2:
                            row=(idnumber,name,total,rest)
                            ws.append(row)

                    workbook_name = "Kalamwadi_blacklist"
                    wb.save(workbook_name + ".xlsx")
                    tkinter.messagebox.showinfo('etax2019','Successfully Created Exel File. You Can Access the file from directory ')
                    os.startfile("Kalamwadi_blacklist.xlsx")


                def backs(self):
                    root.destroy()
                    viewentry1()

                def __init__(self, top=None):
                    '''This class configures and populates the toplevel window.
                       top is the toplevel containing window.'''
                    _bgcolor = '#d9d9d9'  # X11 color: 'gray85'
                    _fgcolor = '#000000'  # X11 color: 'black'
                    _compcolor = '#d9d9d9' # X11 color: 'gray85'
                    _ana1color = '#d9d9d9' # X11 color: 'gray85'
                    _ana2color = '#ececec' # Closest X11 color: 'gray92'
                    font10 = "-family {Captain Howdy} -size 13"
                    font12 = "-family {Bodoni MT Black} -size 11 -weight bold"
                    font15 = "-family {Rockwell Extra} -size 15 -weight bold"
                    font18 = "-family {Rockwell Condensed} -size 40 -weight bold"
                    font9 = "-family {SI Font} -size 36"

                    top.geometry("935x638+237+119")
                    top.title("New Toplevel")
                    top.configure(background="#7a7a7a")

                    self.Frame1 = tk.Frame(top)
                    self.Frame1.place(relx=0.043, rely=0.047, relheight=0.149
                            , relwidth=0.754)
                    self.Frame1.configure(relief='raised')
                    self.Frame1.configure(borderwidth="10")
                    self.Frame1.configure(relief='raised')
                    self.Frame1.configure(background="#5b5b5b")
                    self.Frame1.configure(width=705)

                    self.Label1 = tk.Label(self.Frame1)
                    self.Label1.place(relx=0.113, rely=0.105, height=58, width=91)
                    self.Label1.configure(background="#5b5b5b")
                    self.Label1.configure(disabledforeground="#a3a3a3")
                    self.Label1.configure(font=font9)
                    self.Label1.configure(foreground="#ff2908")
                    self.Label1.configure(text='''TAX''')
                    self.Label1.configure(width=91)

                    self.Label1_1 = tk.Label(self.Frame1)
                    self.Label1_1.place(relx=0.071, rely=0.105, height=58, width=41)
                    self.Label1_1.configure(activebackground="#f9f9f9")
                    self.Label1_1.configure(activeforeground="black")
                    self.Label1_1.configure(background="#5b5b5b")
                    self.Label1_1.configure(disabledforeground="#a3a3a3")
                    self.Label1_1.configure(font="-family {SI Font} -size 36")
                    self.Label1_1.configure(foreground="#2631d1")
                    self.Label1_1.configure(highlightbackground="#d9d9d9")
                    self.Label1_1.configure(highlightcolor="black")
                    self.Label1_1.configure(text='''e''')
                    self.Label1_1.configure(width=41)

                    self.Label1_2 = tk.Label(self.Frame1)
                    self.Label1_2.place(relx=0.241, rely=0.105, height=58, width=101)
                    self.Label1_2.configure(activebackground="#f9f9f9")
                    self.Label1_2.configure(activeforeground="black")
                    self.Label1_2.configure(background="#5b5b5b")
                    self.Label1_2.configure(disabledforeground="#a3a3a3")
                    self.Label1_2.configure(font="-family {SI Font} -size 36")
                    self.Label1_2.configure(foreground="#fff71c")
                    self.Label1_2.configure(highlightbackground="#d9d9d9")
                    self.Label1_2.configure(highlightcolor="black")
                    self.Label1_2.configure(text='''2019''')
                    self.Label1_2.configure(width=101)

                    self.Label1_3 = tk.Label(self.Frame1)
                    self.Label1_3.place(relx=0.411, rely=0.105, height=58, width=371)
                    self.Label1_3.configure(activebackground="#f9f9f9")
                    self.Label1_3.configure(activeforeground="black")
                    self.Label1_3.configure(background="#5b5b5b")
                    self.Label1_3.configure(disabledforeground="#a3a3a3")
                    self.Label1_3.configure(font="-family {SI Font} -size 36")
                    self.Label1_3.configure(foreground="#33a521")
                    self.Label1_3.configure(highlightbackground="#d9d9d9")
                    self.Label1_3.configure(highlightcolor="black")
                    self.Label1_3.configure(text='''VIEW DATABASE''')
                    self.Label1_3.configure(width=371)

                    self.Frame2 = tk.Frame(top)
                    self.Frame2.place(relx=0.032, rely=0.266, relheight=0.697
                            , relwidth=0.455)
                    self.Frame2.configure(relief='sunken')
                    self.Frame2.configure(borderwidth="10")
                    self.Frame2.configure(relief='sunken')
                    self.Frame2.configure(background="#5b5b5b")
                    self.Frame2.configure(width=425)

                    self.Label1_5 = tk.Label(self.Frame2)
                    self.Label1_5.place(relx=0.259, rely=0.045, height=48, width=201)
                    self.Label1_5.configure(activebackground="#f9f9f9")
                    self.Label1_5.configure(activeforeground="black")
                    self.Label1_5.configure(background="#5b5b5b")
                    self.Label1_5.configure(disabledforeground="#a3a3a3")
                    self.Label1_5.configure(font="-family {SI Font} -size 36")
                    self.Label1_5.configure(foreground="#432eff")
                    self.Label1_5.configure(highlightbackground="#d9d9d9")
                    self.Label1_5.configure(highlightcolor="black")
                    self.Label1_5.configure(text='''Receipt''')
                    self.Label1_5.configure(width=201)

                    self.Label2 = tk.Label(self.Frame2)
                    self.Label2.place(relx=0.212, rely=0.225, height=27, width=244)
                    self.Label2.configure(background="#5b5b5b")
                    self.Label2.configure(disabledforeground="#a3a3a3")
                    self.Label2.configure(font=font10)
                    self.Label2.configure(foreground="#f7ff17")
                    self.Label2.configure(text='''Get PDF copy of Reciept''')

                    self.Label2_8 = tk.Label(self.Frame2)
                    self.Label2_8.place(relx=0.094, rely=0.315, height=27, width=344)
                    self.Label2_8.configure(activebackground="#f9f9f9")
                    self.Label2_8.configure(activeforeground="black")
                    self.Label2_8.configure(background="#5b5b5b")
                    self.Label2_8.configure(disabledforeground="#a3a3a3")
                    self.Label2_8.configure(font="-family {Bodoni MT Black} -size 11 -weight bold")
                    self.Label2_8.configure(foreground="#f7ff17")
                    self.Label2_8.configure(highlightbackground="#d9d9d9")
                    self.Label2_8.configure(highlightcolor="black")
                    self.Label2_8.configure(text='''Enter UID Number Here :''')
                    self.Label2_8.configure(width=344)

                    self.Entry1 = tk.Entry(self.Frame2)
                    self.Entry1.place(relx=0.306, rely=0.404,height=20, relwidth=0.386)
                    self.Entry1.configure(background="white")
                    self.Entry1.configure(disabledforeground="#a3a3a3")
                    self.Entry1.configure(font="TkFixedFont")
                    self.Entry1.configure(foreground="#000000")
                    self.Entry1.configure(insertbackground="black")

                    self.btn_pdf = tk.Button(self.Frame2)
                    self.btn_pdf.place(relx=0.306, rely=0.517, height=56, width=165)
                    self.btn_pdf.configure(activebackground="#ececec")
                    self.btn_pdf.configure(activeforeground="#000000")
                    self.btn_pdf.configure(background="#3f8e36")
                    self.btn_pdf.configure(borderwidth="10")
                    self.btn_pdf.configure(disabledforeground="#a3a3a3")
                    self.btn_pdf.configure(font=font15)
                    self.btn_pdf.configure(foreground="#000000")
                    self.btn_pdf.configure(highlightbackground="#d9d9d9")
                    self.btn_pdf.configure(highlightcolor="black")
                    self.btn_pdf.configure(pady="0")
                    self.btn_pdf.configure(text='''GENERATE''',command=self.pdfs)

                    self.Label3 = tk.Label(self.Frame2)
                    self.Label3.place(relx=0.353, rely=0.674, height=114, width=104)
                    self.Label3.configure(background="#d9d9d9")
                    self.Label3.configure(disabledforeground="#a3a3a3")
                    self.Label3.configure(foreground="#000000")
                    photo_location = os.path.join(prog_location,"./adv1.png")
                    self._img0 = tk.PhotoImage(file=photo_location)
                    self.Label3.configure(image=self._img0)
                    self.Label3.configure(text='''Label''')
                    self.Label3.configure(width=104)

                    self.Frame2_4 = tk.Frame(top)
                    self.Frame2_4.place(relx=0.524, rely=0.266, relheight=0.697
                            , relwidth=0.455)
                    self.Frame2_4.configure(relief='sunken')
                    self.Frame2_4.configure(borderwidth="10")
                    self.Frame2_4.configure(relief='sunken')
                    self.Frame2_4.configure(background="#5b5b5b")
                    self.Frame2_4.configure(highlightbackground="#d9d9d9")
                    self.Frame2_4.configure(highlightcolor="black")
                    self.Frame2_4.configure(width=425)

                    self.Label1_6 = tk.Label(self.Frame2_4)
                    self.Label1_6.place(relx=0.235, rely=0.045, height=58, width=241)
                    self.Label1_6.configure(activebackground="#f9f9f9")
                    self.Label1_6.configure(activeforeground="black")
                    self.Label1_6.configure(background="#5b5b5b")
                    self.Label1_6.configure(disabledforeground="#a3a3a3")
                    self.Label1_6.configure(font="-family {SI Font} -size 36")
                    self.Label1_6.configure(foreground="#432eff")
                    self.Label1_6.configure(highlightbackground="#d9d9d9")
                    self.Label1_6.configure(highlightcolor="black")
                    self.Label1_6.configure(text='''Special List''')
                    self.Label1_6.configure(width=241)

                    self.Label2_7 = tk.Label(self.Frame2_4)
                    self.Label2_7.place(relx=0.047, rely=0.27, height=27, width=384)
                    self.Label2_7.configure(activebackground="#f9f9f9")
                    self.Label2_7.configure(activeforeground="black")
                    self.Label2_7.configure(background="#5b5b5b")
                    self.Label2_7.configure(disabledforeground="#a3a3a3")
                    self.Label2_7.configure(font=font12)
                    self.Label2_7.configure(foreground="#f7ff17")
                    self.Label2_7.configure(highlightbackground="#d9d9d9")
                    self.Label2_7.configure(highlightcolor="black")
                    self.Label2_7.configure(text='''of people who paid less than 50% Ammount of tax''')
                    self.Label2_7.configure(width=384)

                    self.Label2_8 = tk.Label(self.Frame2_4)
                    self.Label2_8.place(relx=0.165, rely=0.202, height=27, width=284)
                    self.Label2_8.configure(activebackground="#f9f9f9")
                    self.Label2_8.configure(activeforeground="black")
                    self.Label2_8.configure(background="#5b5b5b")
                    self.Label2_8.configure(disabledforeground="#a3a3a3")
                    self.Label2_8.configure(font="-family {Captain Howdy} -size 13")
                    self.Label2_8.configure(foreground="#f7ff17")
                    self.Label2_8.configure(highlightbackground="#d9d9d9")
                    self.Label2_8.configure(highlightcolor="black")
                    self.Label2_8.configure(text='''Get Excel sheet''')
                    self.Label2_8.configure(width=284)

                    self.btn_getlist = tk.Button(self.Frame2_4)
                    self.btn_getlist.place(relx=0.212, rely=0.382, height=86, width=265)
                    self.btn_getlist.configure(activebackground="#ececec")
                    self.btn_getlist.configure(activeforeground="#000000")
                    self.btn_getlist.configure(background="#3f8e36")
                    self.btn_getlist.configure(borderwidth="20")
                    self.btn_getlist.configure(disabledforeground="#a3a3a3")
                    self.btn_getlist.configure(font=font18)
                    self.btn_getlist.configure(foreground="#111c7a")
                    self.btn_getlist.configure(highlightbackground="#d9d9d9")
                    self.btn_getlist.configure(highlightcolor="black")
                    self.btn_getlist.configure(pady="0")
                    self.btn_getlist.configure(text='''GET LIST''')
                    self.btn_getlist.configure(width=265,command=self.getlists)

                    self.Label4 = tk.Label(self.Frame2_4)
                    self.Label4.place(relx=0.376, rely=0.629, height=130, width=130)
                    self.Label4.configure(background="#d9d9d9")
                    self.Label4.configure(disabledforeground="#a3a3a3")
                    self.Label4.configure(foreground="#000000")
                    photo_location = os.path.join(prog_location,"./adv2.png")
                    self._img1 = tk.PhotoImage(file=photo_location)
                    self.Label4.configure(image=self._img1)
                    self.Label4.configure(text='''Label''')

                    self.btn_back = tk.Button(top)
                    self.btn_back.place(relx=0.845, rely=0.094, height=46, width=95)
                    self.btn_back.configure(activebackground="#ececec")
                    self.btn_back.configure(activeforeground="#000000")
                    self.btn_back.configure(background="#ff2008")
                    self.btn_back.configure(borderwidth="10")
                    self.btn_back.configure(disabledforeground="#a3a3a3")
                    self.btn_back.configure(font="-family {Rockwell Extra} -size 15 -weight bold")
                    self.btn_back.configure(foreground="#000000")
                    self.btn_back.configure(highlightbackground="#d9d9d9")
                    self.btn_back.configure(highlightcolor="black")
                    self.btn_back.configure(pady="0")
                    self.btn_back.configure(text='''BACK''')
                    self.btn_back.configure(width=95,command=self.backs)

            if __name__ == '__main__':
                vp_start_gui()



















        class Toplevel1:
            def exits(self):
                msg=tkinter.messagebox.askyesno("etax-2019","Do You Want To Exit ?")
                if msg:
                    os._exit(1)
            def backs(self):
                msg=tkinter.messagebox.askyesno("etax-2019","Do You Want To Go Back To Main Window?")
                if msg:
                    root.destroy()
                    main()

            def exports(self):
                try :
                    mydb=mysql.connector.connect(host='remotemysql.com',user='NhpDbfVMT3',password='NXWSVJNu55',database='NhpDbfVMT3',port='3306')
                except :
                    tkinter.messagebox.showerror('etax-2019','Failed to connect server, Please contact your administrator')

                mycursor=mydb.cursor()
                wb = Workbook()

                SQL = 'SELECT * from kalamwadi;'
                mycursor.execute(SQL)
                results = mycursor.fetchall()
                ws = wb.create_sheet(0)
                ws.title = 'kalamwadi_data'
                ws.append(mycursor.column_names)
                for row in results:
                    ws.append(row)

                SQL = 'SELECT * from kalamwadiinfo;'
                mycursor.execute(SQL)
                results = mycursor.fetchall()
                ws = wb.create_sheet(0)
                ws.title = 'kalamwadi_info'
                ws.append(mycursor.column_names)
                for row in results:
                    ws.append(row)

                workbook_name = "Kalamwadi_data"
                wb.save(workbook_name + ".xlsx")
                tkinter.messagebox.showinfo('etax2019','Successfully Created Exel File. You Can Access the file from directory ')
                os.startfile("Kalamwadi_data.xlsx")



            def viewnames(self):
                try:
                    mydb=mysql.connector.connect(host='remotemysql.com',user='NhpDbfVMT3',password='NXWSVJNu55',database='NhpDbfVMT3',port='3306')
                    mycursor=mydb.cursor()
                    query = ("SELECT idnumber, meternumber, wardnumber, name, housetax, healthtax, lighttax, watertax, total, reciptnumber, housetaxpaid, healthtaxpaid, lighttaxpaid, watertaxpaid, totalpaid, rest FROM kalamwadi")
                    mycursor.execute(query)
                    for(idnumber, meternumber, wardnumber, name, housetax, healthtax, lighttax, watertax, total, reciptnumber, housetaxpaid, healthtaxpaid, lighttaxpaid, watertaxpaid, totalpaid, rest) in mycursor:
                        s="{}   {}".format(idnumber, name)
                        self.box1o1.insert(0,s)
                except:
                    msg= tkinter.messagebox.showerror("eTAX-2019",'Server Connection Failed Please Try To reconnect. If problem Continues, contact Administrator.')

            def viewalls(self):
                mydb=mysql.connector.connect(host='remotemysql.com',user='NhpDbfVMT3',password='NXWSVJNu55',database='NhpDbfVMT3',port='3306')
                mycursor=mydb.cursor()
                query = ("SELECT idnumber, meternumber, wardnumber, name, housetax, healthtax, lighttax, watertax, total, reciptnumber, housetaxpaid, healthtaxpaid, lighttaxpaid, watertaxpaid, totalpaid, rest, datei FROM kalamwadi")
                mycursor.execute(query)
                for(idnumber, meternumber, wardnumber, name, housetax, healthtax, lighttax, watertax, total, reciptnumber, housetaxpaid, healthtaxpaid, lighttaxpaid, watertaxpaid, totalpaid, rest, datei) in mycursor:
                    s="{}        {}       {}       {}       {}       {}       {}       {}       {}       {}       {}       {}       {}       {}       {}    {}".format(idnumber, meternumber, wardnumber, housetax, healthtax, lighttax, watertax, total, reciptnumber, housetaxpaid, healthtaxpaid, lighttaxpaid, watertaxpaid, totalpaid, rest, datei)
                    self.box2o1.insert(0,s)
            def clearalls(self):
                 self.box2o1.delete(0,tk.END)
                 self.box1o1.delete(0,tk.END)

            def viewtotals(self):
                try :
                    mydb=mysql.connector.connect(host='remotemysql.com',user='NhpDbfVMT3',password='NXWSVJNu55',database='NhpDbfVMT3',port='3306')
                except :
                    tkinter.messagebox.showerror('etax-2019','Failed to connect server, Please contact your administrator')
                mycursor=mydb.cursor()
                query=("SELECT totalpaid from kalamwadi")
                mycursor.execute(query)
                sumofall=0
                for totalpaid in mycursor:
                    p="{}".format(totalpaid)
                    p=p[2:len(p)-3]
                    p=float(p)
                    sumofall=sumofall+p
                sumofall=str(sumofall)
                self.txt_viewtotal.insert(1.0,sumofall)
            def advances(self):
                root.destroy()
                advances1()


            def __init__(self, top=None):
                '''This class configures and populates the toplevel window.
                   top is the toplevel containing window.'''
                _bgcolor = '#d9d9d9'  # X11 color: 'gray85'
                _fgcolor = '#000000'  # X11 color: 'black'
                _compcolor = '#d9d9d9' # X11 color: 'gray85'
                _ana1color = '#d9d9d9' # X11 color: 'gray85' 
                _ana2color = '#ececec' # Closest X11 color: 'gray92' 
                font10 = "-family {Rockwell Extra Bold} -size 12 -weight bold "  \
                    "-slant roman -underline 0 -overstrike 0"
                font11 = "-family {Rockwell Extra Bold} -size 40 -weight bold "  \
                    "-slant roman -underline 0 -overstrike 0"
                font12 = "-family Rockwell -size 15 -weight normal -slant "  \
                    "roman -underline 0 -overstrike 0"
                font13 = "-family Rockwell -size 9 -weight normal -slant roman"  \
                    " -underline 0 -overstrike 0"
                font14 = "-family Rockwell -size 12 -weight normal -slant "  \
                    "roman -underline 0 -overstrike 0"
                font15 = "-family Rockwell -size 13 -weight bold -slant roman "  \
                    "-underline 0 -overstrike 0"
                font20 = "-family {Britannic Bold} -size 48 -weight bold "  \
                    "-slant roman -underline 0 -overstrike 0"
                font16 = "-family {Palatino Linotype} -size 11 -weight normal "  \
                    "-slant roman -underline 0 -overstrike 0"
                font9 = "-family {Segoe Script} -size 12 -weight normal -slant"  \
                    " italic -underline 0 -overstrike 0"
                font41= "-family {Arial Black} -size 16 -weight bold"
                self.style = ttk.Style()
                if sys.platform == "win32":
                    self.style.theme_use('winnative')
                self.style.configure('.',background=_bgcolor)
                self.style.configure('.',foreground=_fgcolor)
                self.style.configure('.',font="TkDefaultFont")
                self.style.map('.',background=
                    [('selected', _compcolor), ('active',_ana2color)])

                top.geometry("1538x878+41+60")
                top.title("New Toplevel")
                top.configure(background="#ffff24")
                top.configure(highlightbackground="#d9d9d9")
                top.configure(highlightcolor="black")

                self.Label1 = tk.Label(top)
                self.Label1.place(relx=0.013, rely=0.023, height=81, width=156)
                self.Label1.configure(activebackground="#f9f9f9")
                self.Label1.configure(activeforeground="black")
                self.Label1.configure(background="#ffff24")
                self.Label1.configure(disabledforeground="#a3a3a3")
                self.Label1.configure(font=font20)
                self.Label1.configure(foreground="#ff250d")
                self.Label1.configure(highlightbackground="#d9d9d9")
                self.Label1.configure(highlightcolor="black")
                self.Label1.configure(text='''eTAX''')

                self.Label1_1 = tk.Label(top)
                self.Label1_1.place(relx=0.117, rely=0.023, height=81, width=156)
                self.Label1_1.configure(activebackground="#f9f9f9")
                self.Label1_1.configure(activeforeground="black")
                self.Label1_1.configure(background="#ffff24")
                self.Label1_1.configure(disabledforeground="#a3a3a3")
                self.Label1_1.configure(font=font20)
                self.Label1_1.configure(foreground="#2212ff")
                self.Label1_1.configure(highlightbackground="#d9d9d9")
                self.Label1_1.configure(highlightcolor="black")
                self.Label1_1.configure(text='''2019''')

                self.Label2 = tk.Label(top)
                self.Label2.place(relx=0.072, rely=0.103, height=31, width=141)
                self.Label2.configure(activebackground="#f9f9f9")
                self.Label2.configure(activeforeground="black")
                self.Label2.configure(background="#ffff24")
                self.Label2.configure(disabledforeground="#a3a3a3")
                self.Label2.configure(font=font9)
                self.Label2.configure(foreground="#13c12a")
                self.Label2.configure(highlightbackground="#d9d9d9")
                self.Label2.configure(highlightcolor="black")
                self.Label2.configure(text='''working for you''')

                self.backbutton = tk.Button(top)
                self.backbutton.place(relx=0.013, rely=0.159, height=44, width=97)
                self.backbutton.configure(activebackground="#ececec")
                self.backbutton.configure(activeforeground="#000000")
                self.backbutton.configure(background="#120bd8")
                self.backbutton.configure(disabledforeground="#a3a3a3")
                self.backbutton.configure(font=font10)
                self.backbutton.configure(foreground="#fcffff")
                self.backbutton.configure(highlightbackground="#d9d9d9")
                self.backbutton.configure(highlightcolor="black")
                self.backbutton.configure(pady="0")
                self.backbutton.configure(command = self.backs,text='''Back''')

                self.exit = tk.Button(top)
                self.exit.place(relx=0.104, rely=0.159, height=44, width=97)
                self.exit.configure(activebackground="#ececec")
                self.exit.configure(activeforeground="#000000")
                self.exit.configure(background="#120bd8")
                self.exit.configure(disabledforeground="#a3a3a3")
                self.exit.configure(font=font10)
                self.exit.configure(foreground="#fcffff")
                self.exit.configure(highlightbackground="#d9d9d9")
                self.exit.configure(highlightcolor="black")
                self.exit.configure(pady="0")
                self.exit.configure(command = self.exits, text='''Exit''')

                self.Label3 = tk.Label(top)
                self.Label3.place(relx=0.013, rely=0.9, height=21, width=56)
                self.Label3.configure(activebackground="#f9f9f9")
                self.Label3.configure(activeforeground="black")
                self.Label3.configure(background="#ffff24")
                self.Label3.configure(disabledforeground="#a3a3a3")
                self.Label3.configure(foreground="#000000")
                self.Label3.configure(highlightbackground="#d9d9d9")
                self.Label3.configure(highlightcolor="black")
                self.Label3.configure(text='''etax-2019''')

                self.Label3_3 = tk.Label(top)
                self.Label3_3.place(relx=0.013, rely=0.923, height=21, width=34)
                self.Label3_3.configure(activebackground="#f9f9f9")
                self.Label3_3.configure(activeforeground="black")
                self.Label3_3.configure(background="#ffff24")
                self.Label3_3.configure(disabledforeground="#a3a3a3")
                self.Label3_3.configure(foreground="#000000")
                self.Label3_3.configure(highlightbackground="#d9d9d9")
                self.Label3_3.configure(highlightcolor="black")
                self.Label3_3.configure(text='''v 1.0.2''')

                self.Label3_4 = tk.Label(top)
                self.Label3_4.place(relx=0.007, rely=0.968, height=21, width=134)
                self.Label3_4.configure(activebackground="#f9f9f9")
                self.Label3_4.configure(activeforeground="black")
                self.Label3_4.configure(background="#ffff24")
                self.Label3_4.configure(disabledforeground="#a3a3a3")
                self.Label3_4.configure(foreground="#000000")
                self.Label3_4.configure(highlightbackground="#d9d9d9")
                self.Label3_4.configure(highlightcolor="black")
                self.Label3_4.configure(text='''Working On Windows''')

                self.Label3_1 = tk.Label(top)
                self.Label3_1.place(relx=0.013, rely=0.945, height=21, width=164)
                self.Label3_1.configure(activebackground="#f9f9f9")
                self.Label3_1.configure(activeforeground="black")
                self.Label3_1.configure(background="#ffff24")
                self.Label3_1.configure(disabledforeground="#a3a3a3")
                self.Label3_1.configure(foreground="#000000")
                self.Label3_1.configure(highlightbackground="#d9d9d9")
                self.Label3_1.configure(highlightcolor="black")
                self.Label3_1.configure(text='''Connected to MySQL server 8.0''')

                self.Label4 = tk.Label(top)
                self.Label4.place(relx=0.416, rely=0.034, height=68, width=321)
                self.Label4.configure(activebackground="#f9f9f9")
                self.Label4.configure(activeforeground="black")
                self.Label4.configure(background="#ffff24")
                self.Label4.configure(disabledforeground="#36911a")
                self.Label4.configure(font=font11)
                self.Label4.configure(foreground="#36911a")
                self.Label4.configure(highlightbackground="#d9d9d9")
                self.Label4.configure(highlightcolor="black")
                self.Label4.configure(text='''Database''')

                self.Label5 = tk.Label(top)
                self.Label5.place(relx=0.793, rely=0.034, height=28, width=192)
                self.Label5.configure(activebackground="#f9f9f9")
                self.Label5.configure(activeforeground="black")
                self.Label5.configure(background="#ffff24")
                self.Label5.configure(disabledforeground="#a3a3a3")
                self.Label5.configure(font=font12)
                self.Label5.configure(foreground="#000000")
                self.Label5.configure(highlightbackground="#d9d9d9")
                self.Label5.configure(highlightcolor="black")
                self.Label5.configure(text='''Village : Kalamwadi''')

                self.Label5_2 = tk.Label(top)
                self.Label5_2.place(relx=0.813, rely=0.068, height=28, width=172)
                self.Label5_2.configure(activebackground="#f9f9f9")
                self.Label5_2.configure(activeforeground="black")
                self.Label5_2.configure(background="#ffff24")
                self.Label5_2.configure(disabledforeground="#a3a3a3")
                self.Label5_2.configure(font=font12)
                self.Label5_2.configure(foreground="#000000")
                self.Label5_2.configure(highlightbackground="#d9d9d9")
                self.Label5_2.configure(highlightcolor="black")
                self.Label5_2.configure(text='''District : Sangli''')

                self.Label5_3 = tk.Label(top)
                self.Label5_3.place(relx=0.897, rely=0.923, height=28, width=172)
                self.Label5_3.configure(activebackground="#f9f9f9")
                self.Label5_3.configure(activeforeground="black")
                self.Label5_3.configure(background="#ffff24")
                self.Label5_3.configure(disabledforeground="#a3a3a3")
                self.Label5_3.configure(font=font13)
                self.Label5_3.configure(foreground="#000000")
                self.Label5_3.configure(highlightbackground="#d9d9d9")
                self.Label5_3.configure(highlightcolor="black")
                self.Label5_3.configure(text='''Server  Status : Online''')

                self.Label5_4 = tk.Label(top)
                self.Label5_4.place(relx=0.904, rely=0.945, height=28, width=172)
                self.Label5_4.configure(activebackground="#f9f9f9")
                self.Label5_4.configure(activeforeground="black")
                self.Label5_4.configure(background="#ffff24")
                self.Label5_4.configure(disabledforeground="#a3a3a3")
                self.Label5_4.configure(font=font13)
                self.Label5_4.configure(foreground="#000000")
                self.Label5_4.configure(highlightbackground="#d9d9d9")
                self.Label5_4.configure(highlightcolor="black")
                self.Label5_4.configure(text='''Host : localhost''')

                self.Label5_5 = tk.Label(top)
                self.Label5_5.place(relx=0.904, rely=0.968, height=28, width=172)
                self.Label5_5.configure(activebackground="#f9f9f9")
                self.Label5_5.configure(activeforeground="black")
                self.Label5_5.configure(background="#ffff24")
                self.Label5_5.configure(disabledforeground="#a3a3a3")
                self.Label5_5.configure(font=font13)
                self.Label5_5.configure(foreground="#000000")
                self.Label5_5.configure(highlightbackground="#d9d9d9")
                self.Label5_5.configure(highlightcolor="black")
                self.Label5_5.configure(text='''Port : 3306''')

                self.Label5_1 = tk.Label(top)
                self.Label5_1.place(relx=0.91, rely=0.091, height=28, width=172)
                self.Label5_1.configure(activebackground="#f9f9f9")
                self.Label5_1.configure(activeforeground="black")
                self.Label5_1.configure(background="#ffff24")
                self.Label5_1.configure(disabledforeground="#a3a3a3")
                self.Label5_1.configure(font=font14)
                self.Label5_1.configure(foreground="#000000")
                self.Label5_1.configure(highlightbackground="#d9d9d9")
                self.Label5_1.configure(highlightcolor="black")
                self.Label5_1.configure(text='''User : user''')

                self.box1o1 = ScrolledListBox(top)
                self.box1o1.place(relx=0.013, rely=0.285, relheight=0.598
                        , relwidth=0.274)
                self.box1o1.configure(background="white")
                self.box1o1.configure(disabledforeground="#a3a3a3")
                self.box1o1.configure(font="TkFixedFont")
                self.box1o1.configure(foreground="black")
                self.box1o1.configure(highlightbackground="#d9d9d9")
                self.box1o1.configure(highlightcolor="#d9d9d9")
                self.box1o1.configure(selectbackground="#c4c4c4")
                self.box1o1.configure(selectforeground="black")
                self.box1o1.configure(width=10)

                self.box2o1 = ScrolledListBox(top)
                self.box2o1.place(relx=0.293, rely=0.285, relheight=0.598
                        , relwidth=0.703)
                self.box2o1.configure(background="white")
                self.box2o1.configure(disabledforeground="#a3a3a3")
                self.box2o1.configure(font="TkFixedFont")
                self.box2o1.configure(foreground="black")
                self.box2o1.configure(highlightbackground="#d9d9d9")
                self.box2o1.configure(highlightcolor="#d9d9d9")
                self.box2o1.configure(selectbackground="#c4c4c4")
                self.box2o1.configure(selectforeground="black")
                self.box2o1.configure(width=10)

                self.TSeparator1 = ttk.Separator(top)
                self.TSeparator1.place(relx=0.923, rely=0.011, relheight=0.114)
                self.TSeparator1.configure(orient="vertical")

                self.TSeparator2 = ttk.Separator(top)
                self.TSeparator2.place(relx=0.013, rely=0.137, relwidth=0.202)

                self.TSeparator3 = ttk.Separator(top)
                self.TSeparator3.place(relx=0.013, rely=0.228, relwidth=0.975)

                self.TSeparator3_6 = ttk.Separator(top)
                self.TSeparator3_6.place(relx=0.013, rely=0.894, relwidth=0.975)

                self.viewbutton = tk.Button(top)
                self.viewbutton.place(relx=0.364, rely=0.923, height=33, width=148)
                self.viewbutton.configure(activebackground="#ececec")
                self.viewbutton.configure(activeforeground="#000000")
                self.viewbutton.configure(background="#2020d8")
                self.viewbutton.configure(disabledforeground="#a3a3a3")
                self.viewbutton.configure(font=font15)
                self.viewbutton.configure(foreground="#ffffff")
                self.viewbutton.configure(highlightbackground="#d9d9d9")
                self.viewbutton.configure(highlightcolor="black")
                self.viewbutton.configure(pady="0")
                self.viewbutton.configure(takefocus="0")
                self.viewbutton.configure(command = self.viewnames , text='''View all Names''')

                self.viewbutton_8 = tk.Button(top)
                self.viewbutton_8.place(relx=0.501, rely=0.923, height=33, width=148)
                self.viewbutton_8.configure(activebackground="#ececec")
                self.viewbutton_8.configure(activeforeground="#000000")
                self.viewbutton_8.configure(background="#2020d8")
                self.viewbutton_8.configure(disabledforeground="#a3a3a3")
                self.viewbutton_8.configure(font=font15)
                self.viewbutton_8.configure(foreground="#ffffff")
                self.viewbutton_8.configure(highlightbackground="#d9d9d9")
                self.viewbutton_8.configure(highlightcolor="black")
                self.viewbutton_8.configure(pady="0")
                self.viewbutton_8.configure(takefocus="0")
                self.viewbutton_8.configure(command = self.viewalls , text='''View all Data''')

                self.viewbutton_9 = tk.Button(top)
                self.viewbutton_9.place(relx=0.637, rely=0.923, height=33, width=108)
                self.viewbutton_9.configure(activebackground="#ececec")
                self.viewbutton_9.configure(activeforeground="#000000")
                self.viewbutton_9.configure(background="#2020d8")
                self.viewbutton_9.configure(disabledforeground="#a3a3a3")
                self.viewbutton_9.configure(font=font15)
                self.viewbutton_9.configure(foreground="#ffffff")
                self.viewbutton_9.configure(highlightbackground="#d9d9d9")
                self.viewbutton_9.configure(highlightcolor="black")
                self.viewbutton_9.configure(pady="0")
                self.viewbutton_9.configure(takefocus="0")
                self.viewbutton_9.configure(command = self.clearalls , text='''Clear all''')


                self.Label7 = tk.Label(top)
                self.Label7.place(relx=0.013, rely=0.251, height=26, width=78)
                self.Label7.configure(activebackground="#f9f9f9")
                self.Label7.configure(activeforeground="black")
                self.Label7.configure(background="#ffff24")
                self.Label7.configure(disabledforeground="#a3a3a3")
                self.Label7.configure(font=font16)
                self.Label7.configure(foreground="#1220e0")
                self.Label7.configure(highlightbackground="#d9d9d9")
                self.Label7.configure(highlightcolor="black")
                self.Label7.configure(text='''Id number''')

                self.menubar = tk.Menu(top,font="TkMenuFont",bg=_bgcolor,fg=_fgcolor)
                top.configure(menu = self.menubar)

                self.Label7_1 = tk.Label(top)
                self.Label7_1.place(relx=0.104, rely=0.251, height=26, width=78)
                self.Label7_1.configure(activebackground="#f9f9f9")
                self.Label7_1.configure(activeforeground="black")
                self.Label7_1.configure(background="#ffff24")
                self.Label7_1.configure(disabledforeground="#a3a3a3")
                self.Label7_1.configure(font=font16)
                self.Label7_1.configure(foreground="#1220e0")
                self.Label7_1.configure(highlightbackground="#d9d9d9")
                self.Label7_1.configure(highlightcolor="black")
                self.Label7_1.configure(text='''Name''')

                self.Label7_2 = tk.Label(top)
                self.Label7_2.place(relx=0.293, rely=0.251, height=26, width=68)
                self.Label7_2.configure(activebackground="#f9f9f9")
                self.Label7_2.configure(activeforeground="black")
                self.Label7_2.configure(background="#ffff24")
                self.Label7_2.configure(disabledforeground="#a3a3a3")
                self.Label7_2.configure(font=font16)
                self.Label7_2.configure(foreground="#1220e0")
                self.Label7_2.configure(highlightbackground="#d9d9d9")
                self.Label7_2.configure(highlightcolor="black")
                self.Label7_2.configure(text='''Id number''')

                self.Label7_3 = tk.Label(top)
                self.Label7_3.place(relx=0.345, rely=0.251, height=26, width=68)
                self.Label7_3.configure(activebackground="#f9f9f9")
                self.Label7_3.configure(activeforeground="black")
                self.Label7_3.configure(background="#ffff24")
                self.Label7_3.configure(disabledforeground="#a3a3a3")
                self.Label7_3.configure(font=font16)
                self.Label7_3.configure(foreground="#1220e0")
                self.Label7_3.configure(highlightbackground="#d9d9d9")
                self.Label7_3.configure(highlightcolor="black")
                self.Label7_3.configure(text='''Meter No.''')

                self.Label7_4 = tk.Label(top)
                self.Label7_4.place(relx=0.397, rely=0.251, height=26, width=68)
                self.Label7_4.configure(activebackground="#f9f9f9")
                self.Label7_4.configure(activeforeground="black")
                self.Label7_4.configure(background="#ffff24")
                self.Label7_4.configure(disabledforeground="#a3a3a3")
                self.Label7_4.configure(font=font16)
                self.Label7_4.configure(foreground="#1220e0")
                self.Label7_4.configure(highlightbackground="#d9d9d9")
                self.Label7_4.configure(highlightcolor="black")
                self.Label7_4.configure(text='''Ward No.''')

                self.Label7_1 = tk.Label(top)
                self.Label7_1.place(relx=0.449, rely=0.251, height=26, width=58)
                self.Label7_1.configure(activebackground="#f9f9f9")
                self.Label7_1.configure(activeforeground="black")
                self.Label7_1.configure(background="#ffff24")
                self.Label7_1.configure(disabledforeground="#a3a3a3")
                self.Label7_1.configure(font=font16)
                self.Label7_1.configure(foreground="#1220e0")
                self.Label7_1.configure(highlightbackground="#d9d9d9")
                self.Label7_1.configure(highlightcolor="black")
                self.Label7_1.configure(text='''Housetax''')

                self.Label7_2 = tk.Label(top)
                self.Label7_2.place(relx=0.494, rely=0.251, height=26, width=68)
                self.Label7_2.configure(activebackground="#f9f9f9")
                self.Label7_2.configure(activeforeground="black")
                self.Label7_2.configure(background="#ffff24")
                self.Label7_2.configure(disabledforeground="#a3a3a3")
                self.Label7_2.configure(font=font16)
                self.Label7_2.configure(foreground="#1220e0")
                self.Label7_2.configure(highlightbackground="#d9d9d9")
                self.Label7_2.configure(highlightcolor="black")
                self.Label7_2.configure(text='''Healthtax''')

                self.Label7_3 = tk.Label(top)
                self.Label7_3.place(relx=0.54, rely=0.251, height=26, width=58)
                self.Label7_3.configure(activebackground="#f9f9f9")
                self.Label7_3.configure(activeforeground="black")
                self.Label7_3.configure(background="#ffff24")
                self.Label7_3.configure(disabledforeground="#a3a3a3")
                self.Label7_3.configure(font=font16)
                self.Label7_3.configure(foreground="#1220e0")
                self.Label7_3.configure(highlightbackground="#d9d9d9")
                self.Label7_3.configure(highlightcolor="black")
                self.Label7_3.configure(text='''Lighttax''')

                self.Label7_5 = tk.Label(top)
                self.Label7_5.place(relx=0.579, rely=0.251, height=26, width=68)
                self.Label7_5.configure(activebackground="#f9f9f9")
                self.Label7_5.configure(activeforeground="black")
                self.Label7_5.configure(background="#ffff24")
                self.Label7_5.configure(disabledforeground="#a3a3a3")
                self.Label7_5.configure(font=font16)
                self.Label7_5.configure(foreground="#1220e0")
                self.Label7_5.configure(highlightbackground="#d9d9d9")
                self.Label7_5.configure(highlightcolor="black")
                self.Label7_5.configure(text='''Watertax''')

                self.Label7_5 = tk.Label(top)
                self.Label7_5.place(relx=0.624, rely=0.251, height=26, width=48)
                self.Label7_5.configure(activebackground="#f9f9f9")
                self.Label7_5.configure(activeforeground="black")
                self.Label7_5.configure(background="#ffff24")
                self.Label7_5.configure(disabledforeground="#a3a3a3")
                self.Label7_5.configure(font=font16)
                self.Label7_5.configure(foreground="#1220e0")
                self.Label7_5.configure(highlightbackground="#d9d9d9")
                self.Label7_5.configure(highlightcolor="black")
                self.Label7_5.configure(text='''Total''')

                self.Label7_5 = tk.Label(top)
                self.Label7_5.place(relx=0.663, rely=0.251, height=26, width=78)
                self.Label7_5.configure(activebackground="#f9f9f9")
                self.Label7_5.configure(activeforeground="black")
                self.Label7_5.configure(background="#ffff24")
                self.Label7_5.configure(disabledforeground="#a3a3a3")
                self.Label7_5.configure(font=font16)
                self.Label7_5.configure(foreground="#1220e0")
                self.Label7_5.configure(highlightbackground="#d9d9d9")
                self.Label7_5.configure(highlightcolor="black")
                self.Label7_5.configure(text='''Reciept No.''')

                self.Label7_6 = tk.Label(top)
                self.Label7_6.place(relx=0.722, rely=0.251, height=26, width=68)
                self.Label7_6.configure(activebackground="#f9f9f9")
                self.Label7_6.configure(activeforeground="black")
                self.Label7_6.configure(background="#ffff24")
                self.Label7_6.configure(disabledforeground="#a3a3a3")
                self.Label7_6.configure(font=font16)
                self.Label7_6.configure(foreground="#1220e0")
                self.Label7_6.configure(highlightbackground="#d9d9d9")
                self.Label7_6.configure(highlightcolor="black")
                self.Label7_6.configure(text='''Housetax''')
                self.Label7_6.configure(width=68)

                self.Label7_6 = tk.Label(top)
                self.Label7_6.place(relx=0.767, rely=0.251, height=26, width=68)
                self.Label7_6.configure(activebackground="#f9f9f9")
                self.Label7_6.configure(activeforeground="black")
                self.Label7_6.configure(background="#ffff24")
                self.Label7_6.configure(disabledforeground="#a3a3a3")
                self.Label7_6.configure(font=font16)
                self.Label7_6.configure(foreground="#1220e0")
                self.Label7_6.configure(highlightbackground="#d9d9d9")
                self.Label7_6.configure(highlightcolor="black")
                self.Label7_6.configure(text='''Lighttax''')

                self.Label7_6 = tk.Label(top)
                self.Label7_6.place(relx=0.813, rely=0.251, height=26, width=68)
                self.Label7_6.configure(activebackground="#f9f9f9")
                self.Label7_6.configure(activeforeground="black")
                self.Label7_6.configure(background="#ffff24")
                self.Label7_6.configure(disabledforeground="#a3a3a3")
                self.Label7_6.configure(font=font16)
                self.Label7_6.configure(foreground="#1220e0")
                self.Label7_6.configure(highlightbackground="#d9d9d9")
                self.Label7_6.configure(highlightcolor="black")
                self.Label7_6.configure(text='''Watertax''')

                self.Label7_6 = tk.Label(top)
                self.Label7_6.place(relx=0.904, rely=0.251, height=26, width=48)
                self.Label7_6.configure(activebackground="#f9f9f9")
                self.Label7_6.configure(activeforeground="black")
                self.Label7_6.configure(background="#ffff24")
                self.Label7_6.configure(disabledforeground="#a3a3a3")
                self.Label7_6.configure(font=font16)
                self.Label7_6.configure(foreground="#1220e0")
                self.Label7_6.configure(highlightbackground="#d9d9d9")
                self.Label7_6.configure(highlightcolor="black")
                self.Label7_6.configure(text='''Total''')

                self.Label7_3 = tk.Label(top)
                self.Label7_3.place(relx=0.858, rely=0.251, height=26, width=68)
                self.Label7_3.configure(activebackground="#f9f9f9")
                self.Label7_3.configure(activeforeground="black")
                self.Label7_3.configure(background="#ffff24")
                self.Label7_3.configure(disabledforeground="#a3a3a3")
                self.Label7_3.configure(font=font16)
                self.Label7_3.configure(foreground="#1220e0")
                self.Label7_3.configure(highlightbackground="#d9d9d9")
                self.Label7_3.configure(highlightcolor="black")
                self.Label7_3.configure(text='''Healthtax''')

                self.Label7_4 = tk.Label(top)
                self.Label7_4.place(relx=0.943, rely=0.251, height=26, width=48)
                self.Label7_4.configure(activebackground="#f9f9f9")
                self.Label7_4.configure(activeforeground="black")
                self.Label7_4.configure(background="#ffff24")
                self.Label7_4.configure(disabledforeground="#a3a3a3")
                self.Label7_4.configure(font=font16)
                self.Label7_4.configure(foreground="#1220e0")
                self.Label7_4.configure(highlightbackground="#d9d9d9")
                self.Label7_4.configure(highlightcolor="black")
                self.Label7_4.configure(text='''Rest''')

                self.Label7_5 = tk.Label(top)
                self.Label7_5.place(relx=0.806, rely=0.216, height=26, width=68)
                self.Label7_5.configure(activebackground="#f9f9f9")
                self.Label7_5.configure(activeforeground="black")
                self.Label7_5.configure(background="#ffff24")
                self.Label7_5.configure(disabledforeground="#a3a3a3")
                self.Label7_5.configure(font=font16)
                self.Label7_5.configure(foreground="#1220e0")
                self.Label7_5.configure(highlightbackground="#d9d9d9")
                self.Label7_5.configure(highlightcolor="black")
                self.Label7_5.configure(text='''Paid''')

                self.TSeparator4 = ttk.Separator(top)
                self.TSeparator4.place(relx=0.715, rely=0.228, relheight=0.057)
                self.TSeparator4.configure(orient="vertical")

                self.TSeparator4_6 = ttk.Separator(top)
                self.TSeparator4_6.place(relx=0.936, rely=0.228, relheight=0.057)
                self.TSeparator4_6.configure(orient="vertical")

                self.viewtotal = tk.Button(top)
                self.viewtotal.place(relx=0.844, rely=0.155, height=53, width=158)
                self.viewtotal.configure(activebackground="#ececec")
                self.viewtotal.configure(activeforeground="#000000")
                self.viewtotal.configure(background="#ff0d0d")
                self.viewtotal.configure(disabledforeground="#a3a3a3")
                self.viewtotal.configure(font="-family {Rockwell} -size 13 -weight bold")
                self.viewtotal.configure(foreground="#ffffff")
                self.viewtotal.configure(highlightbackground="#d9d9d9")
                self.viewtotal.configure(highlightcolor="black")
                self.viewtotal.configure(pady="0")
                self.viewtotal.configure(takefocus="0")
                self.viewtotal.configure(text='''VIEW TOTAL''')
                self.viewtotal.configure(width=158,command=self.viewtotals)

                self.viewtotalm = tk.Button(top)
                self.viewtotalm.place(relx=0.544, rely=0.155, height=53, width=168)
                self.viewtotalm.configure(activebackground="#ececec")
                self.viewtotalm.configure(activeforeground="#000000")
                self.viewtotalm.configure(background="Green")
                self.viewtotalm.configure(disabledforeground="#a3a3a3")
                self.viewtotalm.configure(font="-family {Rockwell} -size 13 -weight bold")
                self.viewtotalm.configure(foreground="#ffffff")
                self.viewtotalm.configure(highlightbackground="#d9d9d9")
                self.viewtotalm.configure(highlightcolor="black")
                self.viewtotalm.configure(pady="0")
                self.viewtotalm.configure(takefocus="0")
                self.viewtotalm.configure(text='''EXPORT TO EXCEL''')
                self.viewtotalm.configure(width=158,command=self.exports)

                self.viewtotalad = tk.Button(top)
                self.viewtotalad.place(relx=0.374, rely=0.155, height=53, width=168)
                self.viewtotalad.configure(activebackground="#ececec")
                self.viewtotalad.configure(activeforeground="#000000")
                self.viewtotalad.configure(background="Blue")
                self.viewtotalad.configure(disabledforeground="#a3a3a3")
                self.viewtotalad.configure(font="-family {Rockwell} -size 13 -weight bold")
                self.viewtotalad.configure(foreground="#ffffff")
                self.viewtotalad.configure(highlightbackground="#d9d9d9")
                self.viewtotalad.configure(highlightcolor="black")
                self.viewtotalad.configure(pady="0")
                self.viewtotalad.configure(takefocus="0")
                self.viewtotalad.configure(text='''ADVANCE MENU''')
                self.viewtotalad.configure(width=158,command=self.advances)                


                self.txt_viewtotal = tk.Text(top)
                self.txt_viewtotal.place(relx=0.706, rely=0.155, relheight=0.053
                        , relwidth=0.121)
                self.txt_viewtotal.configure(background="#757575")
                self.txt_viewtotal.configure(font=font41)
                self.txt_viewtotal.configure(foreground="aqua")
                self.txt_viewtotal.configure(highlightbackground="#d9d9d9")
                self.txt_viewtotal.configure(highlightcolor="black")
                self.txt_viewtotal.configure(insertbackground="black")
                self.txt_viewtotal.configure(selectbackground="#c4c4c4")
                self.txt_viewtotal.configure(selectforeground="black")
                self.txt_viewtotal.configure(width=194)
                self.txt_viewtotal.configure(wrap='word')


                self.Label6 = tk.Label(top)
                self.Label6.place(relx=0.949, rely=0.034, height=44, width=44)
                self.Label6.configure(background="#d9d9d9")
                self.Label6.configure(disabledforeground="#a3a3a3")
                self.Label6.configure(foreground="#000000")
                self._img1 = tk.PhotoImage(file="./login3.png")
                self.Label6.configure(image=self._img1)
                self.Label6.configure(text='''Label''')

        # The following code is added to facilitate the Scrolled widgets you specified.
        class AutoScroll(object):
            '''Configure the scrollbars for a widget.'''

            def __init__(self, master):
                #  Rozen. Added the try-except clauses so that this class
                #  could be used for scrolled entry widget for which vertical
                #  scrolling is not supported. 5/7/14.
                try:
                    vsb = ttk.Scrollbar(master, orient='vertical', command=self.yview)
                except:
                    pass
                hsb = ttk.Scrollbar(master, orient='horizontal', command=self.xview)

                #self.configure(yscrollcommand=_autoscroll(vsb),
                #    xscrollcommand=_autoscroll(hsb))
                try:
                    self.configure(yscrollcommand=self._autoscroll(vsb))
                except:
                    pass
                self.configure(xscrollcommand=self._autoscroll(hsb))

                self.grid(column=0, row=0, sticky='nsew')
                try:
                    vsb.grid(column=1, row=0, sticky='ns')
                except:
                    pass
                hsb.grid(column=0, row=1, sticky='ew')

                master.grid_columnconfigure(0, weight=1)
                master.grid_rowconfigure(0, weight=1)

                # Copy geometry methods of master  (taken from ScrolledText.py)
                if py3:
                    methods = tk.Pack.__dict__.keys() | tk.Grid.__dict__.keys() \
                          | tk.Place.__dict__.keys()
                else:
                    methods = tk.Pack.__dict__.keys() + tk.Grid.__dict__.keys() \
                          + tk.Place.__dict__.keys()

                for meth in methods:
                    if meth[0] != '_' and meth not in ('config', 'configure'):
                        setattr(self, meth, getattr(master, meth))

            @staticmethod
            def _autoscroll(sbar):
                '''Hide and show scrollbar as needed.'''
                def wrapped(first, last):
                    first, last = float(first), float(last)
                    if first <= 0 and last >= 1:
                        sbar.grid_remove()
                    else:
                        sbar.grid()
                    sbar.set(first, last)
                return wrapped

            def __str__(self):
                return str(self.master)

        def _create_container(func):
            '''Creates a ttk Frame with a given master, and use this new frame to
            place the scrollbars and the widget.'''
            def wrapped(cls, master, **kw):
                container = ttk.Frame(master)
                container.bind('<Enter>', lambda e: _bound_to_mousewheel(e, container))
                container.bind('<Leave>', lambda e: _unbound_to_mousewheel(e, container))
                return func(cls, container, **kw)
            return wrapped

        class ScrolledListBox(AutoScroll, tk.Listbox):
            '''A standard Tkinter Text widget with scrollbars that will
            automatically show/hide as needed.'''
            @_create_container
            def __init__(self, master, **kw):
                tk.Listbox.__init__(self, master, **kw)
                AutoScroll.__init__(self, master)

        import platform
        def _bound_to_mousewheel(event, widget):
            child = widget.winfo_children()[0]
            if platform.system() == 'Windows' or platform.system() == 'Darwin':
                child.bind_all('<MouseWheel>', lambda e: _on_mousewheel(e, child))
                child.bind_all('<Shift-MouseWheel>', lambda e: _on_shiftmouse(e, child))
            else:
                child.bind_all('<Button-4>', lambda e: _on_mousewheel(e, child))
                child.bind_all('<Button-5>', lambda e: _on_mousewheel(e, child))
                child.bind_all('<Shift-Button-4>', lambda e: _on_shiftmouse(e, child))
                child.bind_all('<Shift-Button-5>', lambda e: _on_shiftmouse(e, child))

        def _unbound_to_mousewheel(event, widget):
            if platform.system() == 'Windows' or platform.system() == 'Darwin':
                widget.unbind_all('<MouseWheel>')
                widget.unbind_all('<Shift-MouseWheel>')
            else:
                widget.unbind_all('<Button-4>')
                widget.unbind_all('<Button-5>')
                widget.unbind_all('<Shift-Button-4>')
                widget.unbind_all('<Shift-Button-5>')

        def _on_mousewheel(event, widget):
            if platform.system() == 'Windows':
                widget.yview_scroll(-1*int(event.delta/120),'units')
            elif platform.system() == 'Darwin':
                widget.yview_scroll(-1*int(event.delta),'units')
            else:
                if event.num == 4:
                    widget.yview_scroll(-1, 'units')
                elif event.num == 5:
                    widget.yview_scroll(1, 'units')

        def _on_shiftmouse(event, widget):
            if platform.system() == 'Windows':
                widget.xview_scroll(-1*int(event.delta/120), 'units')
            elif platform.system() == 'Darwin':
                widget.xview_scroll(-1*int(event.delta), 'units')
            else:
                if event.num == 4:
                    widget.xview_scroll(-1, 'units')
                elif event.num == 5:
                    widget.xview_scroll(1, 'units')

        if __name__ == '__main__':
            vp_start_gui()





















    


























































    def extra1():

        try:
            import Tkinter as tk
        except ImportError:
            import tkinter as tk

        try:
            import ttk
            py3 = False
        except ImportError:
            import tkinter.ttk as ttk
            py3 = True

        import extramainpage_support
        import os.path

        def vp_start_gui():
            '''Starting point when module is the main routine.'''
            global val, w, root
            global prog_location
            prog_call = sys.argv[0]
            print ('prog_call = {}'.format(prog_call))
            prog_location = os.path.split(prog_call)[0]
            print ('prog_location = {}'.format(prog_location))
            sys.stdout.flush()
            root = tk.Tk()
            top = Toplevel1 (root)
            extramainpage_support.init(root, top)
            root.mainloop()

        w = None
        def create_Toplevel1(root, *args, **kwargs):
            '''Starting point when module is imported by another program.'''
            global w, w_win, rt
            global prog_location
            prog_call = sys.argv[0]
            print ('prog_call = {}'.format(prog_call))
            prog_location = os.path.split(prog_call)[0]
            print ('prog_location = {}'.format(prog_location))
            rt = root
            w = tk.Toplevel (root)
            top = Toplevel1 (w)
            extramainpage_support.init(w, top, *args, **kwargs)
            return (w, top)

        def destroy_Toplevel1():
            global w
            w.destroy()
            w = None






        def world1():

            try:
                import Tkinter as tk
            except ImportError:
                import tkinter as tk

            try:
                import ttk
                py3 = False
            except ImportError:
                import tkinter.ttk as ttk
                py3 = True

            import connect_support

            def vp_start_gui():
                '''Starting point when module is the main routine.'''
                global val, w, root
                root = tk.Tk()
                top = Toplevel1 (root)
                connect_support.init(root, top)
                root.mainloop()

            w = None
            def create_Toplevel1(root, *args, **kwargs):
                '''Starting point when module is imported by another program.'''
                global w, w_win, rt
                rt = root
                w = tk.Toplevel (root)
                top = Toplevel1 (w)
                connect_support.init(w, top, *args, **kwargs)
                return (w, top)

            def destroy_Toplevel1():
                global w
                w.destroy()
                w = None

            class Toplevel1:
                def exits(self):
                    msg=tkinter.messagebox.showinfo("etax-2019","Do You Want To Exit ?")
                    if msg:
                        os._exit(1)
                def backs(self):
                    root.destroy()
                    extra1()
                def viewnames(self):
                    vill=str(self.villagename.get());

                    if (vill == ""):
                        msg=tkinter.messagebox.showerror("etax2019","Please Enter villagename")

                    try :
                        mydb=mysql.connector.connect(host='remotemysql.com',user='NhpDbfVMT3',password='NXWSVJNu55',database='NhpDbfVMT3',port='3306')
                    except :
                        tkinter.messagebox.showerror('etax-2019','Failed to connect server, Please contact your administrator')
                    
                    mycursor=mydb.cursor()
                    query = ("SELECT idnumber, meternumber, wardnumber, name, housetax, healthtax, lighttax, watertax, total, reciptnumber, housetaxpaid, healthtaxpaid, lighttaxpaid, watertaxpaid, totalpaid, rest, datei FROM "+vill)
                    mycursor.execute(query)
                    for(idnumber, meternumber, wardnumber, name, housetax, healthtax, lighttax, watertax, total, reciptnumber, housetaxpaid, healthtaxpaid, lighttaxpaid, watertaxpaid, totalpaid, rest, datei) in mycursor:
                        s="{}   {}".format(idnumber, name)
                        self.box1o1.insert(0,s)



                def viewalls(self):
                    vill=str(self.villagename.get());
                    if (vill == ""):
                        msg=tkinter.messagebox.showerror("etax2019","Please Enter villagename")
                    try :
                        mydb=mysql.connector.connect(host='remotemysql.com',user='NhpDbfVMT3',password='NXWSVJNu55',database='NhpDbfVMT3',port='3306')
                    except :
                        tkinter.messagebox.showerror('etax-2019','Failed to connect server, Please contact your administrator')
                    mycursor=mydb.cursor()
                    query = ("SELECT idnumber, meternumber, wardnumber, name, housetax, healthtax, lighttax, watertax, total, reciptnumber, housetaxpaid, healthtaxpaid, lighttaxpaid, watertaxpaid, totalpaid, rest , datei FROM "+vill)
                    mycursor.execute(query)
                    for(idnumber, meternumber, wardnumber, name, housetax, healthtax, lighttax, watertax, total, reciptnumber, housetaxpaid, healthtaxpaid, lighttaxpaid, watertaxpaid, totalpaid, rest, datei) in mycursor:
                        s="{}        {}       {}       {}       {}       {}       {}       {}       {}       {}       {}       {}       {}       {}       {}    {}".format(idnumber, meternumber, wardnumber, housetax, healthtax, lighttax, watertax, total, reciptnumber, housetaxpaid, healthtaxpaid, lighttaxpaid, watertaxpaid, totalpaid, rest, datei)
                        self.box2o1.insert(0,s)
                def clearalls(self):
                    self.box2o1.delete(0,tk.END)
                    self.box1o1.delete(0,tk.END)

                def findss(self):
                    vill=str(self.villagename.get());
                    uid=str(self.Entry2.get());


                    if (vill == ""):
                        msg=tkinter.messagebox.showerror("etax2019","Please Enter villagename")


                    try :
                        mydb=mysql.connector.connect(host='remotemysql.com',user='NhpDbfVMT3',password='NXWSVJNu55',database='NhpDbfVMT3',port='3306')
                    except :
                        tkinter.messagebox.showerror('etax-2019','Failed to connect server, Please contact your administrator')
                    mycursor=mydb.cursor()
                    query = ("SELECT idnumber, meternumber, wardnumber, name, housetax, healthtax, lighttax, watertax, total, reciptnumber, housetaxpaid, healthtaxpaid, lighttaxpaid, watertaxpaid, totalpaid, rest, datei FROM %s WHERE idnumber= %s" %(vill,uid))
                    mycursor.execute(query)
                    for(idnumber, meternumber, wardnumber, name, housetax, healthtax, lighttax, watertax, total, reciptnumber, housetaxpaid, healthtaxpaid, lighttaxpaid, watertaxpaid, totalpaid, rest, datei) in mycursor:
                        s="{}        {}       {}       {}       {}       {}       {}       {}       {}       {}       {}       {}       {}       {}       {}    {}".format(idnumber, meternumber, wardnumber, housetax, healthtax, lighttax, watertax, total, reciptnumber, housetaxpaid, healthtaxpaid, lighttaxpaid, watertaxpaid, totalpaid, rest, datei)
                        self.box2o1.insert(0,s)

                    try :
                        mydb=mysql.connector.connect(host='remotemysql.com',user='NhpDbfVMT3',password='NXWSVJNu55',database='NhpDbfVMT3',port='3306')
                    except :
                        tkinter.messagebox.showerror('etax-2019','Failed to connect server, Please contact your administrator')
                    mycursor=mydb.cursor()
                    query = ("SELECT idnumber, meternumber, wardnumber, name, housetax, healthtax, lighttax, watertax, total, reciptnumber, housetaxpaid, healthtaxpaid, lighttaxpaid, watertaxpaid, totalpaid, rest FROM %s WHERE idnumber= %s" %(vill,uid))
                    mycursor.execute(query)
                    for(idnumber, meternumber, wardnumber, name, housetax, healthtax, lighttax, watertax, total, reciptnumber, housetaxpaid, healthtaxpaid, lighttaxpaid, watertaxpaid, totalpaid, rest) in mycursor:
                        s="{}   {}".format(idnumber, name)
                        self.box1o1.insert(0,s)




                def __init__(self, top=None):
                    '''This class configures and populates the toplevel window.
                       top is the toplevel containing window.'''
                    _bgcolor = '#d9d9d9'  # X11 color: 'gray85'
                    _fgcolor = '#000000'  # X11 color: 'black'
                    _compcolor = '#d9d9d9' # X11 color: 'gray85'
                    _ana1color = '#d9d9d9' # X11 color: 'gray85' 
                    _ana2color = '#ececec' # Closest X11 color: 'gray92' 
                    font10 = "-family {Rockwell Extra Bold} -size 12 -weight bold "  \
                        "-slant roman -underline 0 -overstrike 0"
                    font11 = "-family {Rockwell Extra Bold} -size 40 -weight bold "  \
                        "-slant roman -underline 0 -overstrike 0"
                    font12 = "-family Rockwell -size 15 -weight normal -slant "  \
                        "roman -underline 0 -overstrike 0"
                    font13 = "-family Rockwell -size 9 -weight normal -slant roman"  \
                        " -underline 0 -overstrike 0"
                    font14 = "-family Rockwell -size 12 -weight normal -slant "  \
                        "roman -underline 0 -overstrike 0"
                    font15 = "-family Rockwell -size 13 -weight bold -slant roman "  \
                        "-underline 0 -overstrike 0"
                    font20 = "-family {Britannic Bold} -size 48 -weight bold "  \
                        "-slant roman -underline 0 -overstrike 0"
                    font16 = "-family {Palatino Linotype} -size 11 -weight normal "  \
                        "-slant roman -underline 0 -overstrike 0"
                    font9 = "-family {Segoe Script} -size 12 -weight normal -slant"  \
                        " italic -underline 0 -overstrike 0"
                    font17 = "-family {Berlin Sans FB} -size 15"
                    self.style = ttk.Style()
                    if sys.platform == "win32":
                        self.style.theme_use('winnative')
                    self.style.configure('.',background=_bgcolor)
                    self.style.configure('.',foreground=_fgcolor)
                    self.style.configure('.',font="TkDefaultFont")
                    self.style.map('.',background=
                        [('selected', _compcolor), ('active',_ana2color)])

                    top.geometry("1538x878+41+60")
                    top.title("New Toplevel")
                    top.configure(background="#ffff24")
                    top.configure(highlightbackground="#d9d9d9")
                    top.configure(highlightcolor="black")

                    self.Label1 = tk.Label(top)
                    self.Label1.place(relx=0.013, rely=0.023, height=81, width=156)
                    self.Label1.configure(activebackground="#f9f9f9")
                    self.Label1.configure(activeforeground="black")
                    self.Label1.configure(background="#ffff24")
                    self.Label1.configure(disabledforeground="#a3a3a3")
                    self.Label1.configure(font=font20)
                    self.Label1.configure(foreground="#ff250d")
                    self.Label1.configure(highlightbackground="#d9d9d9")
                    self.Label1.configure(highlightcolor="black")
                    self.Label1.configure(text='''eTAX''')

                    self.Label1_1 = tk.Label(top)
                    self.Label1_1.place(relx=0.117, rely=0.023, height=81, width=156)
                    self.Label1_1.configure(activebackground="#f9f9f9")
                    self.Label1_1.configure(activeforeground="black")
                    self.Label1_1.configure(background="#ffff24")
                    self.Label1_1.configure(disabledforeground="#a3a3a3")
                    self.Label1_1.configure(font=font20)
                    self.Label1_1.configure(foreground="#2212ff")
                    self.Label1_1.configure(highlightbackground="#d9d9d9")
                    self.Label1_1.configure(highlightcolor="black")
                    self.Label1_1.configure(text='''2019''')

                    self.Label2 = tk.Label(top)
                    self.Label2.place(relx=0.072, rely=0.103, height=31, width=141)
                    self.Label2.configure(activebackground="#f9f9f9")
                    self.Label2.configure(activeforeground="black")
                    self.Label2.configure(background="#ffff24")
                    self.Label2.configure(disabledforeground="#a3a3a3")
                    self.Label2.configure(font=font9)
                    self.Label2.configure(foreground="#13c12a")
                    self.Label2.configure(highlightbackground="#d9d9d9")
                    self.Label2.configure(highlightcolor="black")
                    self.Label2.configure(text='''working for you''')

                    self.backbutton = tk.Button(top)
                    self.backbutton.place(relx=0.013, rely=0.159, height=44, width=97)
                    self.backbutton.configure(activebackground="#ececec")
                    self.backbutton.configure(activeforeground="#000000")
                    self.backbutton.configure(background="#120bd8")
                    self.backbutton.configure(disabledforeground="#a3a3a3")
                    self.backbutton.configure(font=font10)
                    self.backbutton.configure(foreground="#fcffff")
                    self.backbutton.configure(highlightbackground="#d9d9d9")
                    self.backbutton.configure(highlightcolor="black")
                    self.backbutton.configure(pady="0")
                    self.backbutton.configure(text='''Back''',command = self.backs)

                    self.exit = tk.Button(top)
                    self.exit.place(relx=0.104, rely=0.159, height=44, width=97)
                    self.exit.configure(activebackground="#ececec")
                    self.exit.configure(activeforeground="#000000")
                    self.exit.configure(background="#120bd8")
                    self.exit.configure(disabledforeground="#a3a3a3")
                    self.exit.configure(font=font10)
                    self.exit.configure(foreground="#fcffff")
                    self.exit.configure(highlightbackground="#d9d9d9")
                    self.exit.configure(highlightcolor="black")
                    self.exit.configure(pady="0")
                    self.exit.configure(command = self.exits, text='''Exit''')

                    self.Label3 = tk.Label(top)
                    self.Label3.place(relx=0.013, rely=0.9, height=21, width=56)
                    self.Label3.configure(activebackground="#f9f9f9")
                    self.Label3.configure(activeforeground="black")
                    self.Label3.configure(background="#ffff24")
                    self.Label3.configure(disabledforeground="#a3a3a3")
                    self.Label3.configure(foreground="#000000")
                    self.Label3.configure(highlightbackground="#d9d9d9")
                    self.Label3.configure(highlightcolor="black")
                    self.Label3.configure(text='''etax-2019''')

                    self.Label3_3 = tk.Label(top)
                    self.Label3_3.place(relx=0.013, rely=0.923, height=21, width=34)
                    self.Label3_3.configure(activebackground="#f9f9f9")
                    self.Label3_3.configure(activeforeground="black")
                    self.Label3_3.configure(background="#ffff24")
                    self.Label3_3.configure(disabledforeground="#a3a3a3")
                    self.Label3_3.configure(foreground="#000000")
                    self.Label3_3.configure(highlightbackground="#d9d9d9")
                    self.Label3_3.configure(highlightcolor="black")
                    self.Label3_3.configure(text='''v 1.0.2''')

                    self.Label3_4 = tk.Label(top)
                    self.Label3_4.place(relx=0.007, rely=0.968, height=21, width=134)
                    self.Label3_4.configure(activebackground="#f9f9f9")
                    self.Label3_4.configure(activeforeground="black")
                    self.Label3_4.configure(background="#ffff24")
                    self.Label3_4.configure(disabledforeground="#a3a3a3")
                    self.Label3_4.configure(foreground="#000000")
                    self.Label3_4.configure(highlightbackground="#d9d9d9")
                    self.Label3_4.configure(highlightcolor="black")
                    self.Label3_4.configure(text='''Working On Windows''')

                    self.Label3_1 = tk.Label(top)
                    self.Label3_1.place(relx=0.013, rely=0.945, height=21, width=164)
                    self.Label3_1.configure(activebackground="#f9f9f9")
                    self.Label3_1.configure(activeforeground="black")
                    self.Label3_1.configure(background="#ffff24")
                    self.Label3_1.configure(disabledforeground="#a3a3a3")
                    self.Label3_1.configure(foreground="#000000")
                    self.Label3_1.configure(highlightbackground="#d9d9d9")
                    self.Label3_1.configure(highlightcolor="black")
                    self.Label3_1.configure(text='''Connected to MySQL server 8.0''')

                    self.Label4 = tk.Label(top)
                    self.Label4.place(relx=0.316, rely=0.034, height=68, width=651)
                    self.Label4.configure(activebackground="#f9f9f9")
                    self.Label4.configure(activeforeground="black")
                    self.Label4.configure(background="#ffff24")
                    self.Label4.configure(disabledforeground="#36911a")
                    self.Label4.configure(font=font11)
                    self.Label4.configure(foreground="#36911a")
                    self.Label4.configure(highlightbackground="#d9d9d9")
                    self.Label4.configure(highlightcolor="black")
                    self.Label4.configure(text='''CONNECTING WORLD''')

                    self.Label5 = tk.Label(top)
                    self.Label5.place(relx=0.793, rely=0.034, height=28, width=192)
                    self.Label5.configure(activebackground="#f9f9f9")
                    self.Label5.configure(activeforeground="black")
                    self.Label5.configure(background="#ffff24")
                    self.Label5.configure(disabledforeground="#a3a3a3")
                    self.Label5.configure(font=font12)
                    self.Label5.configure(foreground="#000000")
                    self.Label5.configure(highlightbackground="#d9d9d9")
                    self.Label5.configure(highlightcolor="black")
                    self.Label5.configure(text='''Village : Kalamwadi''')

                    self.Label5_2 = tk.Label(top)
                    self.Label5_2.place(relx=0.813, rely=0.068, height=28, width=172)
                    self.Label5_2.configure(activebackground="#f9f9f9")
                    self.Label5_2.configure(activeforeground="black")
                    self.Label5_2.configure(background="#ffff24")
                    self.Label5_2.configure(disabledforeground="#a3a3a3")
                    self.Label5_2.configure(font=font12)
                    self.Label5_2.configure(foreground="#000000")
                    self.Label5_2.configure(highlightbackground="#d9d9d9")
                    self.Label5_2.configure(highlightcolor="black")
                    self.Label5_2.configure(text='''District : Sangli''')

                    self.Label5_3 = tk.Label(top)
                    self.Label5_3.place(relx=0.897, rely=0.923, height=28, width=172)
                    self.Label5_3.configure(activebackground="#f9f9f9")
                    self.Label5_3.configure(activeforeground="black")
                    self.Label5_3.configure(background="#ffff24")
                    self.Label5_3.configure(disabledforeground="#a3a3a3")
                    self.Label5_3.configure(font=font13)
                    self.Label5_3.configure(foreground="#000000")
                    self.Label5_3.configure(highlightbackground="#d9d9d9")
                    self.Label5_3.configure(highlightcolor="black")
                    self.Label5_3.configure(text='''Server  Status : Online''')

                    self.Label5_4 = tk.Label(top)
                    self.Label5_4.place(relx=0.904, rely=0.945, height=28, width=172)
                    self.Label5_4.configure(activebackground="#f9f9f9")
                    self.Label5_4.configure(activeforeground="black")
                    self.Label5_4.configure(background="#ffff24")
                    self.Label5_4.configure(disabledforeground="#a3a3a3")
                    self.Label5_4.configure(font=font13)
                    self.Label5_4.configure(foreground="#000000")
                    self.Label5_4.configure(highlightbackground="#d9d9d9")
                    self.Label5_4.configure(highlightcolor="black")
                    self.Label5_4.configure(text='''Host : localhost''')

                    self.Label5_5 = tk.Label(top)
                    self.Label5_5.place(relx=0.904, rely=0.968, height=28, width=172)
                    self.Label5_5.configure(activebackground="#f9f9f9")
                    self.Label5_5.configure(activeforeground="black")
                    self.Label5_5.configure(background="#ffff24")
                    self.Label5_5.configure(disabledforeground="#a3a3a3")
                    self.Label5_5.configure(font=font13)
                    self.Label5_5.configure(foreground="#000000")
                    self.Label5_5.configure(highlightbackground="#d9d9d9")
                    self.Label5_5.configure(highlightcolor="black")
                    self.Label5_5.configure(text='''Port : 3306''')

                    self.Label5_1 = tk.Label(top)
                    self.Label5_1.place(relx=0.91, rely=0.091, height=28, width=172)
                    self.Label5_1.configure(activebackground="#f9f9f9")
                    self.Label5_1.configure(activeforeground="black")
                    self.Label5_1.configure(background="#ffff24")
                    self.Label5_1.configure(disabledforeground="#a3a3a3")
                    self.Label5_1.configure(font=font14)
                    self.Label5_1.configure(foreground="#000000")
                    self.Label5_1.configure(highlightbackground="#d9d9d9")
                    self.Label5_1.configure(highlightcolor="black")
                    self.Label5_1.configure(text='''User : user''')

                    self.box1o1 = ScrolledListBox(top)
                    self.box1o1.place(relx=0.013, rely=0.285, relheight=0.598
                            , relwidth=0.274)
                    self.box1o1.configure(background="white")
                    self.box1o1.configure(disabledforeground="#a3a3a3")
                    self.box1o1.configure(font="TkFixedFont")
                    self.box1o1.configure(foreground="black")
                    self.box1o1.configure(highlightbackground="#d9d9d9")
                    self.box1o1.configure(highlightcolor="#d9d9d9")
                    self.box1o1.configure(selectbackground="#c4c4c4")
                    self.box1o1.configure(selectforeground="black")
                    self.box1o1.configure(width=10)

                    self.box2o1 = ScrolledListBox(top)
                    self.box2o1.place(relx=0.293, rely=0.285, relheight=0.598
                            , relwidth=0.703)
                    self.box2o1.configure(background="white")
                    self.box2o1.configure(disabledforeground="#a3a3a3")
                    self.box2o1.configure(font="TkFixedFont")
                    self.box2o1.configure(foreground="black")
                    self.box2o1.configure(highlightbackground="#d9d9d9")
                    self.box2o1.configure(highlightcolor="#d9d9d9")
                    self.box2o1.configure(selectbackground="#c4c4c4")
                    self.box2o1.configure(selectforeground="black")
                    self.box2o1.configure(width=10)

                    self.TSeparator1 = ttk.Separator(top)
                    self.TSeparator1.place(relx=0.923, rely=0.011, relheight=0.114)
                    self.TSeparator1.configure(orient="vertical")

                    self.TSeparator2 = ttk.Separator(top)
                    self.TSeparator2.place(relx=0.013, rely=0.137, relwidth=0.202)

                    self.TSeparator3 = ttk.Separator(top)
                    self.TSeparator3.place(relx=0.013, rely=0.228, relwidth=0.975)

                    self.TSeparator3_6 = ttk.Separator(top)
                    self.TSeparator3_6.place(relx=0.013, rely=0.894, relwidth=0.975)

                    self.viewbutton = tk.Button(top)
                    self.viewbutton.place(relx=0.364, rely=0.923, height=33, width=148)
                    self.viewbutton.configure(activebackground="#ececec")
                    self.viewbutton.configure(activeforeground="#000000")
                    self.viewbutton.configure(background="#2020d8")
                    self.viewbutton.configure(disabledforeground="#a3a3a3")
                    self.viewbutton.configure(font=font15)
                    self.viewbutton.configure(foreground="#ffffff")
                    self.viewbutton.configure(highlightbackground="#d9d9d9")
                    self.viewbutton.configure(highlightcolor="black")
                    self.viewbutton.configure(pady="0")
                    self.viewbutton.configure(takefocus="0")
                    self.viewbutton.configure(command = self.viewnames , text='''View all Names''')

                    self.viewbutton_8 = tk.Button(top)
                    self.viewbutton_8.place(relx=0.501, rely=0.923, height=33, width=148)
                    self.viewbutton_8.configure(activebackground="#ececec")
                    self.viewbutton_8.configure(activeforeground="#000000")
                    self.viewbutton_8.configure(background="#2020d8")
                    self.viewbutton_8.configure(disabledforeground="#a3a3a3")
                    self.viewbutton_8.configure(font=font15)
                    self.viewbutton_8.configure(foreground="#ffffff")
                    self.viewbutton_8.configure(highlightbackground="#d9d9d9")
                    self.viewbutton_8.configure(highlightcolor="black")
                    self.viewbutton_8.configure(pady="0")
                    self.viewbutton_8.configure(takefocus="0")
                    self.viewbutton_8.configure(command = self.viewalls , text='''View all Data''')

                    self.viewbutton_9 = tk.Button(top)
                    self.viewbutton_9.place(relx=0.637, rely=0.923, height=33, width=108)
                    self.viewbutton_9.configure(activebackground="#ececec")
                    self.viewbutton_9.configure(activeforeground="#000000")
                    self.viewbutton_9.configure(background="#2020d8")
                    self.viewbutton_9.configure(disabledforeground="#a3a3a3")
                    self.viewbutton_9.configure(font=font15)
                    self.viewbutton_9.configure(foreground="#ffffff")
                    self.viewbutton_9.configure(highlightbackground="#d9d9d9")
                    self.viewbutton_9.configure(highlightcolor="black")
                    self.viewbutton_9.configure(pady="0")
                    self.viewbutton_9.configure(takefocus="0")
                    self.viewbutton_9.configure(command = self.clearalls , text='''Clear all''')


                    self.Label7 = tk.Label(top)
                    self.Label7.place(relx=0.013, rely=0.251, height=26, width=78)
                    self.Label7.configure(activebackground="#f9f9f9")
                    self.Label7.configure(activeforeground="black")
                    self.Label7.configure(background="#ffff24")
                    self.Label7.configure(disabledforeground="#a3a3a3")
                    self.Label7.configure(font=font16)
                    self.Label7.configure(foreground="#1220e0")
                    self.Label7.configure(highlightbackground="#d9d9d9")
                    self.Label7.configure(highlightcolor="black")
                    self.Label7.configure(text='''Id number''')

                    self.menubar = tk.Menu(top,font="TkMenuFont",bg=_bgcolor,fg=_fgcolor)
                    top.configure(menu = self.menubar)

                    self.Label7_1 = tk.Label(top)
                    self.Label7_1.place(relx=0.104, rely=0.251, height=26, width=78)
                    self.Label7_1.configure(activebackground="#f9f9f9")
                    self.Label7_1.configure(activeforeground="black")
                    self.Label7_1.configure(background="#ffff24")
                    self.Label7_1.configure(disabledforeground="#a3a3a3")
                    self.Label7_1.configure(font=font16)
                    self.Label7_1.configure(foreground="#1220e0")
                    self.Label7_1.configure(highlightbackground="#d9d9d9")
                    self.Label7_1.configure(highlightcolor="black")
                    self.Label7_1.configure(text='''Name''')

                    self.Label7_2 = tk.Label(top)
                    self.Label7_2.place(relx=0.293, rely=0.251, height=26, width=68)
                    self.Label7_2.configure(activebackground="#f9f9f9")
                    self.Label7_2.configure(activeforeground="black")
                    self.Label7_2.configure(background="#ffff24")
                    self.Label7_2.configure(disabledforeground="#a3a3a3")
                    self.Label7_2.configure(font=font16)
                    self.Label7_2.configure(foreground="#1220e0")
                    self.Label7_2.configure(highlightbackground="#d9d9d9")
                    self.Label7_2.configure(highlightcolor="black")
                    self.Label7_2.configure(text='''Id number''')

                    self.Label7_3 = tk.Label(top)
                    self.Label7_3.place(relx=0.345, rely=0.251, height=26, width=68)
                    self.Label7_3.configure(activebackground="#f9f9f9")
                    self.Label7_3.configure(activeforeground="black")
                    self.Label7_3.configure(background="#ffff24")
                    self.Label7_3.configure(disabledforeground="#a3a3a3")
                    self.Label7_3.configure(font=font16)
                    self.Label7_3.configure(foreground="#1220e0")
                    self.Label7_3.configure(highlightbackground="#d9d9d9")
                    self.Label7_3.configure(highlightcolor="black")
                    self.Label7_3.configure(text='''Meter No.''')

                    self.Label7_4 = tk.Label(top)
                    self.Label7_4.place(relx=0.397, rely=0.251, height=26, width=68)
                    self.Label7_4.configure(activebackground="#f9f9f9")
                    self.Label7_4.configure(activeforeground="black")
                    self.Label7_4.configure(background="#ffff24")
                    self.Label7_4.configure(disabledforeground="#a3a3a3")
                    self.Label7_4.configure(font=font16)
                    self.Label7_4.configure(foreground="#1220e0")
                    self.Label7_4.configure(highlightbackground="#d9d9d9")
                    self.Label7_4.configure(highlightcolor="black")
                    self.Label7_4.configure(text='''Ward No.''')

                    self.Label7_1 = tk.Label(top)
                    self.Label7_1.place(relx=0.449, rely=0.251, height=26, width=58)
                    self.Label7_1.configure(activebackground="#f9f9f9")
                    self.Label7_1.configure(activeforeground="black")
                    self.Label7_1.configure(background="#ffff24")
                    self.Label7_1.configure(disabledforeground="#a3a3a3")
                    self.Label7_1.configure(font=font16)
                    self.Label7_1.configure(foreground="#1220e0")
                    self.Label7_1.configure(highlightbackground="#d9d9d9")
                    self.Label7_1.configure(highlightcolor="black")
                    self.Label7_1.configure(text='''Housetax''')

                    self.Label7_2 = tk.Label(top)
                    self.Label7_2.place(relx=0.494, rely=0.251, height=26, width=68)
                    self.Label7_2.configure(activebackground="#f9f9f9")
                    self.Label7_2.configure(activeforeground="black")
                    self.Label7_2.configure(background="#ffff24")
                    self.Label7_2.configure(disabledforeground="#a3a3a3")
                    self.Label7_2.configure(font=font16)
                    self.Label7_2.configure(foreground="#1220e0")
                    self.Label7_2.configure(highlightbackground="#d9d9d9")
                    self.Label7_2.configure(highlightcolor="black")
                    self.Label7_2.configure(text='''Healthtax''')

                    self.Label7_3 = tk.Label(top)
                    self.Label7_3.place(relx=0.54, rely=0.251, height=26, width=58)
                    self.Label7_3.configure(activebackground="#f9f9f9")
                    self.Label7_3.configure(activeforeground="black")
                    self.Label7_3.configure(background="#ffff24")
                    self.Label7_3.configure(disabledforeground="#a3a3a3")
                    self.Label7_3.configure(font=font16)
                    self.Label7_3.configure(foreground="#1220e0")
                    self.Label7_3.configure(highlightbackground="#d9d9d9")
                    self.Label7_3.configure(highlightcolor="black")
                    self.Label7_3.configure(text='''Lighttax''')

                    self.Label7_5 = tk.Label(top)
                    self.Label7_5.place(relx=0.579, rely=0.251, height=26, width=68)
                    self.Label7_5.configure(activebackground="#f9f9f9")
                    self.Label7_5.configure(activeforeground="black")
                    self.Label7_5.configure(background="#ffff24")
                    self.Label7_5.configure(disabledforeground="#a3a3a3")
                    self.Label7_5.configure(font=font16)
                    self.Label7_5.configure(foreground="#1220e0")
                    self.Label7_5.configure(highlightbackground="#d9d9d9")
                    self.Label7_5.configure(highlightcolor="black")
                    self.Label7_5.configure(text='''Watertax''')

                    self.Label7_5 = tk.Label(top)
                    self.Label7_5.place(relx=0.624, rely=0.251, height=26, width=48)
                    self.Label7_5.configure(activebackground="#f9f9f9")
                    self.Label7_5.configure(activeforeground="black")
                    self.Label7_5.configure(background="#ffff24")
                    self.Label7_5.configure(disabledforeground="#a3a3a3")
                    self.Label7_5.configure(font=font16)
                    self.Label7_5.configure(foreground="#1220e0")
                    self.Label7_5.configure(highlightbackground="#d9d9d9")
                    self.Label7_5.configure(highlightcolor="black")
                    self.Label7_5.configure(text='''Total''')

                    self.Label7_5 = tk.Label(top)
                    self.Label7_5.place(relx=0.663, rely=0.251, height=26, width=78)
                    self.Label7_5.configure(activebackground="#f9f9f9")
                    self.Label7_5.configure(activeforeground="black")
                    self.Label7_5.configure(background="#ffff24")
                    self.Label7_5.configure(disabledforeground="#a3a3a3")
                    self.Label7_5.configure(font=font16)
                    self.Label7_5.configure(foreground="#1220e0")
                    self.Label7_5.configure(highlightbackground="#d9d9d9")
                    self.Label7_5.configure(highlightcolor="black")
                    self.Label7_5.configure(text='''Reciept No.''')

                    self.Label7_6 = tk.Label(top)
                    self.Label7_6.place(relx=0.722, rely=0.251, height=26, width=68)
                    self.Label7_6.configure(activebackground="#f9f9f9")
                    self.Label7_6.configure(activeforeground="black")
                    self.Label7_6.configure(background="#ffff24")
                    self.Label7_6.configure(disabledforeground="#a3a3a3")
                    self.Label7_6.configure(font=font16)
                    self.Label7_6.configure(foreground="#1220e0")
                    self.Label7_6.configure(highlightbackground="#d9d9d9")
                    self.Label7_6.configure(highlightcolor="black")
                    self.Label7_6.configure(text='''Housetax''')
                    self.Label7_6.configure(width=68)

                    self.Label7_6 = tk.Label(top)
                    self.Label7_6.place(relx=0.767, rely=0.251, height=26, width=68)
                    self.Label7_6.configure(activebackground="#f9f9f9")
                    self.Label7_6.configure(activeforeground="black")
                    self.Label7_6.configure(background="#ffff24")
                    self.Label7_6.configure(disabledforeground="#a3a3a3")
                    self.Label7_6.configure(font=font16)
                    self.Label7_6.configure(foreground="#1220e0")
                    self.Label7_6.configure(highlightbackground="#d9d9d9")
                    self.Label7_6.configure(highlightcolor="black")
                    self.Label7_6.configure(text='''Lighttax''')

                    self.Label7_6 = tk.Label(top)
                    self.Label7_6.place(relx=0.813, rely=0.251, height=26, width=68)
                    self.Label7_6.configure(activebackground="#f9f9f9")
                    self.Label7_6.configure(activeforeground="black")
                    self.Label7_6.configure(background="#ffff24")
                    self.Label7_6.configure(disabledforeground="#a3a3a3")
                    self.Label7_6.configure(font=font16)
                    self.Label7_6.configure(foreground="#1220e0")
                    self.Label7_6.configure(highlightbackground="#d9d9d9")
                    self.Label7_6.configure(highlightcolor="black")
                    self.Label7_6.configure(text='''Watertax''')

                    self.Label7_6 = tk.Label(top)
                    self.Label7_6.place(relx=0.904, rely=0.251, height=26, width=48)
                    self.Label7_6.configure(activebackground="#f9f9f9")
                    self.Label7_6.configure(activeforeground="black")
                    self.Label7_6.configure(background="#ffff24")
                    self.Label7_6.configure(disabledforeground="#a3a3a3")
                    self.Label7_6.configure(font=font16)
                    self.Label7_6.configure(foreground="#1220e0")
                    self.Label7_6.configure(highlightbackground="#d9d9d9")
                    self.Label7_6.configure(highlightcolor="black")
                    self.Label7_6.configure(text='''Total''')

                    self.Label7_3 = tk.Label(top)
                    self.Label7_3.place(relx=0.858, rely=0.251, height=26, width=68)
                    self.Label7_3.configure(activebackground="#f9f9f9")
                    self.Label7_3.configure(activeforeground="black")
                    self.Label7_3.configure(background="#ffff24")
                    self.Label7_3.configure(disabledforeground="#a3a3a3")
                    self.Label7_3.configure(font=font16)
                    self.Label7_3.configure(foreground="#1220e0")
                    self.Label7_3.configure(highlightbackground="#d9d9d9")
                    self.Label7_3.configure(highlightcolor="black")
                    self.Label7_3.configure(text='''Healthtax''')

                    self.Label7_4 = tk.Label(top)
                    self.Label7_4.place(relx=0.943, rely=0.251, height=26, width=48)
                    self.Label7_4.configure(activebackground="#f9f9f9")
                    self.Label7_4.configure(activeforeground="black")
                    self.Label7_4.configure(background="#ffff24")
                    self.Label7_4.configure(disabledforeground="#a3a3a3")
                    self.Label7_4.configure(font=font16)
                    self.Label7_4.configure(foreground="#1220e0")
                    self.Label7_4.configure(highlightbackground="#d9d9d9")
                    self.Label7_4.configure(highlightcolor="black")
                    self.Label7_4.configure(text='''Rest''')

                    self.Label7_5 = tk.Label(top)
                    self.Label7_5.place(relx=0.806, rely=0.216, height=26, width=68)
                    self.Label7_5.configure(activebackground="#f9f9f9")
                    self.Label7_5.configure(activeforeground="black")
                    self.Label7_5.configure(background="#ffff24")
                    self.Label7_5.configure(disabledforeground="#a3a3a3")
                    self.Label7_5.configure(font=font16)
                    self.Label7_5.configure(foreground="#1220e0")
                    self.Label7_5.configure(highlightbackground="#d9d9d9")
                    self.Label7_5.configure(highlightcolor="black")
                    self.Label7_5.configure(text='''Paid''')

                    self.TSeparator4 = ttk.Separator(top)
                    self.TSeparator4.place(relx=0.715, rely=0.228, relheight=0.057)
                    self.TSeparator4.configure(orient="vertical")

                    self.TSeparator4_6 = ttk.Separator(top)
                    self.TSeparator4_6.place(relx=0.936, rely=0.228, relheight=0.057)
                    self.TSeparator4_6.configure(orient="vertical")


                    self.villagename = tk.Entry(top)
                    self.villagename.place(relx=0.375, rely=0.179, height=20, relwidth=0.153)
                    self.villagename.configure(background="white")
                    self.villagename.configure(disabledforeground="#a3a3a3")
                    self.villagename.configure(font="TkFixedFont")
                    self.villagename.configure(foreground="#1b1391")
                    self.villagename.configure(insertbackground="black")
                    self.villagename.configure(width=244)

                    self.Entry2 = tk.Entry(top)
                    self.Entry2.place(relx=0.675, rely=0.179,height=20, relwidth=0.146)
                    self.Entry2.configure(background="white")
                    self.Entry2.configure(disabledforeground="#a3a3a3")
                    self.Entry2.configure(font="TkFixedFont")
                    self.Entry2.configure(foreground="#000000")
                    self.Entry2.configure(insertbackground="black")
                    self.Entry2.configure(width=234)

                    self.Label7 = tk.Label(top)
                    self.Label7.place(relx=0.269, rely=0.179, height=21, width=154)
                    self.Label7.configure(background="#ffff24")
                    self.Label7.configure(disabledforeground="#a3a3a3")
                    self.Label7.configure(font=font17)
                    self.Label7.configure(foreground="#000000")
                    self.Label7.configure(text='''Village Name :''')
                    self.Label7.configure(width=154)

                    self.Label7_1 = tk.Label(top)
                    self.Label7_1.place(relx=0.575, rely=0.179, height=21, width=154)
                    self.Label7_1.configure(activebackground="#f9f9f9")
                    self.Label7_1.configure(activeforeground="black")
                    self.Label7_1.configure(background="#ffff24")
                    self.Label7_1.configure(disabledforeground="#a3a3a3")
                    self.Label7_1.configure(font="-family {Berlin Sans FB} -size 15")
                    self.Label7_1.configure(foreground="#000000")
                    self.Label7_1.configure(highlightbackground="#d9d9d9")
                    self.Label7_1.configure(highlightcolor="black")
                    self.Label7_1.configure(text='''UID Number  :''')
                    self.Label7_1.configure(width=154)

                    self.btn_find = tk.Button(top)
                    self.btn_find.place(relx=0.856, rely=0.167, height=34, width=97)
                    self.btn_find.configure(activebackground="#ececec")
                    self.btn_find.configure(activeforeground="#000000")
                    self.btn_find.configure(background="#ff330a")
                    self.btn_find.configure(disabledforeground="#a3a3a3")
                    self.btn_find.configure(font="-family {Rockwell Extra Bold} -size 12 -weight bold")
                    self.btn_find.configure(foreground="#fcffff")
                    self.btn_find.configure(highlightbackground="#d9d9d9")
                    self.btn_find.configure(highlightcolor="black")
                    self.btn_find.configure(pady="0")
                    self.btn_find.configure(text='''FIND''')
                    self.btn_find.configure(width=97,command=self.findss)
                    
                    """self.Label6 = tk.Label(top)
                    self.Label6.place(relx=0.949, rely=0.034, height=44, width=44)
                    self.Label6.configure(background="#d9d9d9")
                    self.Label6.configure(disabledforeground="#a3a3a3")
                    self.Label6.configure(foreground="#000000")
                    self._img1 = tk.PhotoImage(file="./login3.png")
                    self.Label6.configure(image=self._img1)
                    self.Label6.configure(text='''Label''')"""

            # The following code is added to facilitate the Scrolled widgets you specified.
            class AutoScroll(object):
                '''Configure the scrollbars for a widget.'''

                def __init__(self, master):
                    #  Rozen. Added the try-except clauses so that this class
                    #  could be used for scrolled entry widget for which vertical
                    #  scrolling is not supported. 5/7/14.
                    try:
                        vsb = ttk.Scrollbar(master, orient='vertical', command=self.yview)
                    except:
                        pass
                    hsb = ttk.Scrollbar(master, orient='horizontal', command=self.xview)

                    #self.configure(yscrollcommand=_autoscroll(vsb),
                    #    xscrollcommand=_autoscroll(hsb))
                    try:
                        self.configure(yscrollcommand=self._autoscroll(vsb))
                    except:
                        pass
                    self.configure(xscrollcommand=self._autoscroll(hsb))

                    self.grid(column=0, row=0, sticky='nsew')
                    try:
                        vsb.grid(column=1, row=0, sticky='ns')
                    except:
                        pass
                    hsb.grid(column=0, row=1, sticky='ew')

                    master.grid_columnconfigure(0, weight=1)
                    master.grid_rowconfigure(0, weight=1)

                    # Copy geometry methods of master  (taken from ScrolledText.py)
                    if py3:
                        methods = tk.Pack.__dict__.keys() | tk.Grid.__dict__.keys() \
                              | tk.Place.__dict__.keys()
                    else:
                        methods = tk.Pack.__dict__.keys() + tk.Grid.__dict__.keys() \
                              + tk.Place.__dict__.keys()

                    for meth in methods:
                        if meth[0] != '_' and meth not in ('config', 'configure'):
                            setattr(self, meth, getattr(master, meth))

                @staticmethod
                def _autoscroll(sbar):
                    '''Hide and show scrollbar as needed.'''
                    def wrapped(first, last):
                        first, last = float(first), float(last)
                        if first <= 0 and last >= 1:
                            sbar.grid_remove()
                        else:
                            sbar.grid()
                        sbar.set(first, last)
                    return wrapped

                def __str__(self):
                    return str(self.master)

            def _create_container(func):
                '''Creates a ttk Frame with a given master, and use this new frame to
                place the scrollbars and the widget.'''
                def wrapped(cls, master, **kw):
                    container = ttk.Frame(master)
                    container.bind('<Enter>', lambda e: _bound_to_mousewheel(e, container))
                    container.bind('<Leave>', lambda e: _unbound_to_mousewheel(e, container))
                    return func(cls, container, **kw)
                return wrapped

            class ScrolledListBox(AutoScroll, tk.Listbox):
                '''A standard Tkinter Text widget with scrollbars that will
                automatically show/hide as needed.'''
                @_create_container
                def __init__(self, master, **kw):
                    tk.Listbox.__init__(self, master, **kw)
                    AutoScroll.__init__(self, master)

            import platform
            def _bound_to_mousewheel(event, widget):
                child = widget.winfo_children()[0]
                if platform.system() == 'Windows' or platform.system() == 'Darwin':
                    child.bind_all('<MouseWheel>', lambda e: _on_mousewheel(e, child))
                    child.bind_all('<Shift-MouseWheel>', lambda e: _on_shiftmouse(e, child))
                else:
                    child.bind_all('<Button-4>', lambda e: _on_mousewheel(e, child))
                    child.bind_all('<Button-5>', lambda e: _on_mousewheel(e, child))
                    child.bind_all('<Shift-Button-4>', lambda e: _on_shiftmouse(e, child))
                    child.bind_all('<Shift-Button-5>', lambda e: _on_shiftmouse(e, child))

            def _unbound_to_mousewheel(event, widget):
                if platform.system() == 'Windows' or platform.system() == 'Darwin':
                    widget.unbind_all('<MouseWheel>')
                    widget.unbind_all('<Shift-MouseWheel>')
                else:
                    widget.unbind_all('<Button-4>')
                    widget.unbind_all('<Button-5>')
                    widget.unbind_all('<Shift-Button-4>')
                    widget.unbind_all('<Shift-Button-5>')

            def _on_mousewheel(event, widget):
                if platform.system() == 'Windows':
                    widget.yview_scroll(-1*int(event.delta/120),'units')
                elif platform.system() == 'Darwin':
                    widget.yview_scroll(-1*int(event.delta),'units')
                else:
                    if event.num == 4:
                        widget.yview_scroll(-1, 'units')
                    elif event.num == 5:
                        widget.yview_scroll(1, 'units')

            def _on_shiftmouse(event, widget):
                if platform.system() == 'Windows':
                    widget.xview_scroll(-1*int(event.delta/120), 'units')
                elif platform.system() == 'Darwin':
                    widget.xview_scroll(-1*int(event.delta), 'units')
                else:
                    if event.num == 4:
                        widget.xview_scroll(-1, 'units')
                    elif event.num == 5:
                        widget.xview_scroll(1, 'units')

            if __name__ == '__main__':
                vp_start_gui()





























        def calculator1():
            import time
            area=float(input("Enter Area In square meter:\n\n"))
            print("Choose The type of Building\n\n1.Hut or mud Building\n2.Stone or Brick Building Built in clay \n 3.Stone, bricks Building built in lime or cement.\n 4.RCC Building ")
            ch1=int(input())
            if  ch1 == 1:
                taxrate=0.30
                maxtaxrate=0.75
                redirec=5320
            elif ch1==2:
                taxrate=0.60
                maxtaxrate=0.120
                redirec=9120
            elif  ch1 == 3:
                taxrate=0.75
                maxtaxrate=0.150
                redirec=12920
            elif ch1==4:
                taxrate=0.120
                maxtaxrate=0.200
                redirec=15200
            print("What is the use of Building\n\n\n\n1.Residential Use\n2.Industrial Use \n3.Commertial Use")
            ch2=int(input())
            if  ch2==1:
                bhtax=1
            elif ch2==2:
                bhtax=1.20
            elif ch2==3:
                taxrate=1.25
            print("What is the age of Building..... in years")
            age=int(input())
            if age <= 2:
                drate=1
                maxdrate=1
            elif age>2 and age<=5:
                drate=0.95
                maxdrate=0.95
            elif age>5 and age<=10:
                drate=0.85
                maxdrate=0.9
            elif age>10 and age<=20:
                drate=0.75
                maxdrate=0.8
            elif age>20 and age<=30:
                drate=0.60
                maxdrate=0.7
            elif age>30 and age<=40:
                drate=0.45
                maxdrate=0.6
            elif age>40 and age<=50:
                drate=0.30
                maxdrate=0.5
            elif age>50 and age<=60:
                drate=0.20
                maxdrate=0.4
            elif age>60:
                drate=0.15
                maxdrate=0.3


            taxa1=area*1130

            taxb1=area*redirec*drate*bhtax
            taxc1=(taxa1+taxb1)/1000
            tax1=taxc1*taxrate

            maxtaxb1=area*redirec*maxdrate*bhtax
            maxtaxc1=(taxa1+maxtaxb1)/1000
            maxtax1=maxtaxc1*maxtaxrate
            print("Building Tax is from ",tax1,"To",maxtax1)

            if area <= 28.00:
                light=10
                health=10
            elif area >28.0 and area<60.0:
                light=20
                health=20
            elif area >60.0 :
                light=25
                health=25
            print("Healthtax : ",health,"\n Lighttax : ",light)
            time.sleep(10)
            main()






















        def online1():
            import sys

            try:
                import Tkinter as tk
            except ImportError:
                import tkinter as tk
            import webbrowser
            import tkinter
            from tkinter import messagebox
            import smtplib


            try:
                import ttk
                py3 = False
            except ImportError:
                import tkinter.ttk as ttk
                py3 = True

            import onlinesupport_support
            import os.path

            def vp_start_gui():
                '''Starting point when module is the main routine.'''
                global val, w, root
                global prog_location
                prog_call = sys.argv[0]
                print ('prog_call = {}'.format(prog_call))
                prog_location = os.path.split(prog_call)[0]
                print ('prog_location = {}'.format(prog_location))
                sys.stdout.flush()
                root = tk.Tk()
                onlinesupport_support.set_Tk_var()
                top = e_TAX_2019 (root)
                onlinesupport_support.init(root, top)
                root.mainloop()

            w = None
            def create_e_TAX_2019(root, *args, **kwargs):
                '''Starting point when module is imported by another program.'''
                global w, w_win, rt
                global prog_location
                prog_call = sys.argv[0]
                print ('prog_call = {}'.format(prog_call))
                prog_location = os.path.split(prog_call)[0]
                print ('prog_location = {}'.format(prog_location))
                rt = root
                w = tk.Toplevel (root)
                onlinesupport_support.set_Tk_var()
                top = e_TAX_2019 (w)
                onlinesupport_support.init(w, top, *args, **kwargs)
                return (w, top)

            def destroy_e_TAX_2019():
                global w
                w.destroy()
                w = None

            class e_TAX_2019:
                def facebooks(self):
                    new =2
                    url= "https://www.facebook.com/pranesh.kulkarni.359"
                    webbrowser.open(url,new=new);

                def calls(self):
                    new =2
                    url= "https://api.whatsapp.com/send?phone=918956795667&text=Hi+User+here&fbclid=IwAR05YmoAuyw0dzyDMetLj3Zcya3QrJxfOW6sVmo4ydcqaaLf019RR4fNT0M"
                    webbrowser.open(url,new=new);

                def websites(self):
                    new =2
                    url= "https://etax20194.webnode.com/"
                    webbrowser.open(url,new=new);

                def submit2s(self):
                    r1=str(self.spin1.get())
                    r2=str(self.spin2.get())
                    r3=str(self.spin3.get())
                    r4=str(self.spin4.get())

                    rev=str(self.txt_review.get())
                    tot= r1+r2+r3+r4+rev
                    mail= smtplib.SMTP('smtp.gmail.com',587)
                    mail.ehlo()
                    mail.starttls()
                    mail.login('etaxsupp2019@gmail.com','Pass@123')
                    mail.sendmail('etaxsupp2019@gmail.com','etaxsupp2019@gmail.com',tot)
                    mail.close()
                    messagebox.showinfo("e-TAX 2019","Thank You For Rating")


                def submit1s(self):
                    a=str(self.txt_village.get())
                    b=str(self.txt_username.get())
                    c=str(self.txt_productid.get())
                    d=str(self.txt_query.get())

                    tota = a+b+c+d
                    mail= smtplib.SMTP('smtp.gmail.com',587)
                    mail.ehlo()
                    mail.starttls()
                    mail.login('etaxsupp2019@gmail.com','Pass@123')
                    mail.sendmail('etaxsupp2019@gmail.com','etaxsupp2019@gmail.com',tota)
                    mail.close()
                    messagebox.showinfo("e-TAX 2019","Bug Reported.")

                def backs(self):
                    root.destroy()
                    extra1()


                def exits(self):
                    msg=tkinter.messagebox.askyesno("e-TAX 2019","Do You Want To EXIT ?")
                    if msg:
                        os._exit(1)





                def __init__(self, top=None):
                    '''This class configures and populates the toplevel window.
                       top is the toplevel containing window.'''
                    _bgcolor = '#d9d9d9'  # X11 color: 'gray85'
                    _fgcolor = '#000000'  # X11 color: 'black'
                    _compcolor = '#d9d9d9' # X11 color: 'gray85'
                    _ana1color = '#d9d9d9' # X11 color: 'gray85'
                    _ana2color = '#ececec' # Closest X11 color: 'gray92'

                    top.geometry("1162x780+138+34")
                    top.title("e-TAX 2019")
                    top.configure(background="#727272")
                    top.configure(highlightbackground="#d9d9d9")
                    top.configure(highlightcolor="black")

                    self.Frame1 = tk.Frame(top)
                    self.Frame1.place(relx=0.164, rely=0.026, relheight=0.186
                            , relwidth=0.667)
                    self.Frame1.configure(relief='ridge')
                    self.Frame1.configure(borderwidth="10")
                    self.Frame1.configure(relief='ridge')
                    self.Frame1.configure(background="#595959")
                    self.Frame1.configure(highlightbackground="#d9d9d9")
                    self.Frame1.configure(highlightcolor="black")
                    self.Frame1.configure(width=775)

                    self.Label1 = tk.Label(self.Frame1)
                    self.Label1.place(relx=0.039, rely=0.207, height=59, width=196)
                    self.Label1.configure(activebackground="#f9f9f9")
                    self.Label1.configure(activeforeground="black")
                    self.Label1.configure(background="#595959")
                    self.Label1.configure(disabledforeground="#a3a3a3")
                    self.Label1.configure(font="-family {Rockwell Extra} -size 40 -weight bold")
                    self.Label1.configure(foreground="#0d4dff")
                    self.Label1.configure(highlightbackground="#d9d9d9")
                    self.Label1.configure(highlightcolor="black")
                    self.Label1.configure(text='''eTAX''')

                    self.Label1_1 = tk.Label(self.Frame1)
                    self.Label1_1.place(relx=0.271, rely=0.207, height=59, width=146)
                    self.Label1_1.configure(activebackground="#f9f9f9")
                    self.Label1_1.configure(activeforeground="black")
                    self.Label1_1.configure(background="#595959")
                    self.Label1_1.configure(disabledforeground="#a3a3a3")
                    self.Label1_1.configure(font="-family {Rockwell Extra} -size 40 -weight bold")
                    self.Label1_1.configure(foreground="#ff2b0a")
                    self.Label1_1.configure(highlightbackground="#d9d9d9")
                    self.Label1_1.configure(highlightcolor="black")
                    self.Label1_1.configure(text='''2019''')

                    self.Label2 = tk.Label(self.Frame1)
                    self.Label2.place(relx=0.142, rely=0.621, height=31, width=184)
                    self.Label2.configure(activebackground="#f9f9f9")
                    self.Label2.configure(activeforeground="black")
                    self.Label2.configure(background="#595959")
                    self.Label2.configure(disabledforeground="#a3a3a3")
                    self.Label2.configure(font="-family {Rage Italic} -size 19 -slant italic")
                    self.Label2.configure(foreground="#f7ff0a")
                    self.Label2.configure(highlightbackground="#d9d9d9")
                    self.Label2.configure(highlightcolor="black")
                    self.Label2.configure(text='''working for you''')

                    self.Label1_2 = tk.Label(self.Frame1)
                    self.Label1_2.place(relx=0.465, rely=0.207, height=69, width=356)
                    self.Label1_2.configure(activebackground="#f9f9f9")
                    self.Label1_2.configure(activeforeground="black")
                    self.Label1_2.configure(background="#595959")
                    self.Label1_2.configure(disabledforeground="#a3a3a3")
                    self.Label1_2.configure(font="-family {Rockwell Extra} -size 28 -weight bold")
                    self.Label1_2.configure(foreground="#5faa14")
                    self.Label1_2.configure(highlightbackground="#d9d9d9")
                    self.Label1_2.configure(highlightcolor="black")
                    self.Label1_2.configure(text='''Online Support''')

                    self.Frame2 = tk.Frame(top)
                    self.Frame2.place(relx=0.052, rely=0.256, relheight=0.455, relwidth=0.46)

                    self.Frame2.configure(relief='ridge')
                    self.Frame2.configure(borderwidth="10")
                    self.Frame2.configure(relief='ridge')
                    self.Frame2.configure(background="#595959")
                    self.Frame2.configure(highlightbackground="#d9d9d9")
                    self.Frame2.configure(highlightcolor="black")
                    self.Frame2.configure(width=535)

                    self.Label6 = tk.Label(self.Frame2)
                    self.Label6.place(relx=0.224, rely=0.056, height=49, width=284)
                    self.Label6.configure(activebackground="#f9f9f9")
                    self.Label6.configure(activeforeground="black")
                    self.Label6.configure(background="#595959")
                    self.Label6.configure(disabledforeground="#a3a3a3")
                    self.Label6.configure(font="-family {Franklin Gothic Demi Cond} -size 36")
                    self.Label6.configure(foreground="#ff9e1f")
                    self.Label6.configure(highlightbackground="#d9d9d9")
                    self.Label6.configure(highlightcolor="black")
                    self.Label6.configure(text='''REPORT A BUG''')

                    self.Label6_5 = tk.Label(self.Frame2)
                    self.Label6_5.place(relx=0.056, rely=0.254, height=39, width=94)
                    self.Label6_5.configure(activebackground="#f9f9f9")
                    self.Label6_5.configure(activeforeground="black")
                    self.Label6_5.configure(background="#595959")
                    self.Label6_5.configure(disabledforeground="#a3a3a3")
                    self.Label6_5.configure(font="-family {Franklin Gothic Demi Cond} -size 16")
                    self.Label6_5.configure(foreground="aqua")
                    self.Label6_5.configure(highlightbackground="#d9d9d9")
                    self.Label6_5.configure(highlightcolor="black")
                    self.Label6_5.configure(text='''Village :''')

                    self.Label6_6 = tk.Label(self.Frame2)
                    self.Label6_6.place(relx=0.056, rely=0.366, height=39, width=94)
                    self.Label6_6.configure(activebackground="#f9f9f9")
                    self.Label6_6.configure(activeforeground="black")
                    self.Label6_6.configure(background="#595959")
                    self.Label6_6.configure(disabledforeground="#a3a3a3")
                    self.Label6_6.configure(font="-family {Franklin Gothic Demi Cond} -size 16")
                    self.Label6_6.configure(foreground="aqua")
                    self.Label6_6.configure(highlightbackground="#d9d9d9")
                    self.Label6_6.configure(highlightcolor="black")
                    self.Label6_6.configure(text='''User Name :''')

                    self.Label6_1 = tk.Label(self.Frame2)
                    self.Label6_1.place(relx=0.056, rely=0.592, height=39, width=94)
                    self.Label6_1.configure(activebackground="#f9f9f9")
                    self.Label6_1.configure(activeforeground="black")
                    self.Label6_1.configure(background="#595959")
                    self.Label6_1.configure(disabledforeground="#a3a3a3")
                    self.Label6_1.configure(font="-family {Franklin Gothic Demi Cond} -size 16")
                    self.Label6_1.configure(foreground="aqua")
                    self.Label6_1.configure(highlightbackground="#d9d9d9")
                    self.Label6_1.configure(highlightcolor="black")
                    self.Label6_1.configure(text='''Query :''')

                    self.Label6_2 = tk.Label(self.Frame2)
                    self.Label6_2.place(relx=0.056, rely=0.479, height=39, width=94)
                    self.Label6_2.configure(activebackground="#f9f9f9")
                    self.Label6_2.configure(activeforeground="black")
                    self.Label6_2.configure(background="#595959")
                    self.Label6_2.configure(disabledforeground="#a3a3a3")
                    self.Label6_2.configure(font="-family {Franklin Gothic Demi Cond} -size 16")
                    self.Label6_2.configure(foreground="aqua")
                    self.Label6_2.configure(highlightbackground="#d9d9d9")
                    self.Label6_2.configure(highlightcolor="black")
                    self.Label6_2.configure(text='''Product ID :''')

                    self.txt_village = tk.Entry(self.Frame2)
                    self.txt_village.place(relx=0.28, rely=0.282,height=20, relwidth=0.643)
                    self.txt_village.configure(background="white")
                    self.txt_village.configure(disabledforeground="#a3a3a3")
                    self.txt_village.configure(font="TkFixedFont")
                    self.txt_village.configure(foreground="#000000")
                    self.txt_village.configure(highlightbackground="#d9d9d9")
                    self.txt_village.configure(highlightcolor="black")
                    self.txt_village.configure(insertbackground="black")
                    self.txt_village.configure(selectbackground="#c4c4c4")
                    self.txt_village.configure(selectforeground="black")

                    self.txt_username = tk.Entry(self.Frame2)
                    self.txt_username.place(relx=0.28, rely=0.394, height=20, relwidth=0.643)

                    self.txt_username.configure(background="white")
                    self.txt_username.configure(disabledforeground="#a3a3a3")
                    self.txt_username.configure(font="TkFixedFont")
                    self.txt_username.configure(foreground="#000000")
                    self.txt_username.configure(highlightbackground="#d9d9d9")
                    self.txt_username.configure(highlightcolor="black")
                    self.txt_username.configure(insertbackground="black")
                    self.txt_username.configure(selectbackground="#c4c4c4")
                    self.txt_username.configure(selectforeground="black")

                    self.txt_productid = tk.Entry(self.Frame2)
                    self.txt_productid.place(relx=0.28, rely=0.507, height=20
                            , relwidth=0.643)
                    self.txt_productid.configure(background="white")
                    self.txt_productid.configure(disabledforeground="#a3a3a3")
                    self.txt_productid.configure(font="TkFixedFont")
                    self.txt_productid.configure(foreground="#000000")
                    self.txt_productid.configure(highlightbackground="#d9d9d9")
                    self.txt_productid.configure(highlightcolor="black")
                    self.txt_productid.configure(insertbackground="black")
                    self.txt_productid.configure(selectbackground="#c4c4c4")
                    self.txt_productid.configure(selectforeground="black")

                    self.txt_query = tk.Entry(self.Frame2)
                    self.txt_query.place(relx=0.28, rely=0.62,height=70, relwidth=0.643)
                    self.txt_query.configure(background="white")
                    self.txt_query.configure(disabledforeground="#a3a3a3")
                    self.txt_query.configure(font="TkFixedFont")
                    self.txt_query.configure(foreground="#000000")
                    self.txt_query.configure(highlightbackground="#d9d9d9")
                    self.txt_query.configure(highlightcolor="black")
                    self.txt_query.configure(insertbackground="black")
                    self.txt_query.configure(selectbackground="#c4c4c4")
                    self.txt_query.configure(selectforeground="black")

                    self.btn_submit1 = tk.Button(self.Frame2)
                    self.btn_submit1.place(relx=0.729, rely=0.845, height=34, width=80)
                    self.btn_submit1.configure(activebackground="#ececec")
                    self.btn_submit1.configure(activeforeground="#000000")
                    self.btn_submit1.configure(background="#ff350d")
                    self.btn_submit1.configure(borderwidth="10")
                    self.btn_submit1.configure(disabledforeground="#a3a3a3")
                    self.btn_submit1.configure(font="-family {Swatch it} -size 11 -weight bold")
                    self.btn_submit1.configure(foreground="#000000")
                    self.btn_submit1.configure(highlightbackground="#d9d9d9")
                    self.btn_submit1.configure(highlightcolor="black")
                    self.btn_submit1.configure(pady="0")
                    self.btn_submit1.configure(text='''Submit''',command= self.submit1s)

                    self.Frame2_1 = tk.Frame(top)
                    self.Frame2_1.place(relx=0.525, rely=0.256, relheight=0.455
                            , relwidth=0.443)
                    self.Frame2_1.configure(relief='ridge')
                    self.Frame2_1.configure(borderwidth="10")
                    self.Frame2_1.configure(relief='ridge')
                    self.Frame2_1.configure(background="#595959")
                    self.Frame2_1.configure(highlightbackground="#d9d9d9")
                    self.Frame2_1.configure(highlightcolor="black")
                    self.Frame2_1.configure(width=515)

                    self.Label6_11 = tk.Label(self.Frame2_1)
                    self.Label6_11.place(relx=0.214, rely=0.028, height=49, width=284)
                    self.Label6_11.configure(activebackground="#f9f9f9")
                    self.Label6_11.configure(activeforeground="black")
                    self.Label6_11.configure(background="#595959")
                    self.Label6_11.configure(disabledforeground="#a3a3a3")
                    self.Label6_11.configure(font="-family {Franklin Gothic Demi Cond} -size 36")
                    self.Label6_11.configure(foreground="#ff9e1f")
                    self.Label6_11.configure(highlightbackground="#d9d9d9")
                    self.Label6_11.configure(highlightcolor="black")
                    self.Label6_11.configure(text='''Give Rating''')

                    self.Label6_3 = tk.Label(self.Frame2_1)
                    self.Label6_3.place(relx=0.117, rely=0.225, height=39, width=94)
                    self.Label6_3.configure(activebackground="#f9f9f9")
                    self.Label6_3.configure(activeforeground="black")
                    self.Label6_3.configure(background="#595959")
                    self.Label6_3.configure(disabledforeground="#a3a3a3")
                    self.Label6_3.configure(font="-family {Franklin Gothic Demi Cond} -size 16")
                    self.Label6_3.configure(foreground="aqua")
                    self.Label6_3.configure(highlightbackground="#d9d9d9")
                    self.Label6_3.configure(highlightcolor="black")
                    self.Label6_3.configure(text='''UI :''')

                    self.Label6_7 = tk.Label(self.Frame2_1)
                    self.Label6_7.place(relx=0.058, rely=0.31, height=39, width=174)
                    self.Label6_7.configure(activebackground="#f9f9f9")
                    self.Label6_7.configure(activeforeground="black")
                    self.Label6_7.configure(background="#595959")
                    self.Label6_7.configure(disabledforeground="#a3a3a3")
                    self.Label6_7.configure(font="-family {Franklin Gothic Demi Cond} -size 16")
                    self.Label6_7.configure(foreground="aqua")
                    self.Label6_7.configure(highlightbackground="#d9d9d9")
                    self.Label6_7.configure(highlightcolor="black")
                    self.Label6_7.configure(text='''Server Connectivity :''')

                    self.Label6_4 = tk.Label(self.Frame2_1)
                    self.Label6_4.place(relx=0.019, rely=0.394, height=39, width=204)
                    self.Label6_4.configure(activebackground="#f9f9f9")
                    self.Label6_4.configure(activeforeground="black")
                    self.Label6_4.configure(background="#595959")
                    self.Label6_4.configure(disabledforeground="#a3a3a3")
                    self.Label6_4.configure(font="-family {Franklin Gothic Demi Cond} -size 16")
                    self.Label6_4.configure(foreground="aqua")
                    self.Label6_4.configure(highlightbackground="#d9d9d9")
                    self.Label6_4.configure(highlightcolor="black")
                    self.Label6_4.configure(text='''Costomer Care Service :''')

                    self.Label6_5 = tk.Label(self.Frame2_1)
                    self.Label6_5.place(relx=0.058, rely=0.479, height=39, width=154)
                    self.Label6_5.configure(activebackground="#f9f9f9")
                    self.Label6_5.configure(activeforeground="black")
                    self.Label6_5.configure(background="#595959")
                    self.Label6_5.configure(disabledforeground="#a3a3a3")
                    self.Label6_5.configure(font="-family {Franklin Gothic Demi Cond} -size 16")
                    self.Label6_5.configure(foreground="aqua")
                    self.Label6_5.configure(highlightbackground="#d9d9d9")
                    self.Label6_5.configure(highlightcolor="black")
                    self.Label6_5.configure(text='''Overall Ratings:''')

                    self.spin1 = tk.Spinbox(self.Frame2_1, from_=1.0, to=5.0)
                    self.spin1.place(relx=0.544, rely=0.254, relheight=0.054, relwidth=0.146)

                    self.spin1.configure(activebackground="#f9f9f9")
                    self.spin1.configure(background="white")
                    self.spin1.configure(buttonbackground="#a3a5a8")
                    self.spin1.configure(disabledforeground="#a3a3a3")
                    self.spin1.configure(foreground="black")
                    self.spin1.configure(highlightbackground="black")
                    self.spin1.configure(highlightcolor="black")
                    self.spin1.configure(insertbackground="black")
                    self.spin1.configure(selectbackground="#c4c4c4")
                    self.spin1.configure(selectforeground="black")
                    self.spin1.configure(textvariable=onlinesupport_support.spinbox1)

                    self.Label5_4 = tk.Label(self.Frame2_1)
                    self.Label5_4.place(relx=0.35, rely=0.169, height=22, width=152)
                    self.Label5_4.configure(activebackground="#f9f9f9")
                    self.Label5_4.configure(activeforeground="black")
                    self.Label5_4.configure(background="#595959")
                    self.Label5_4.configure(disabledforeground="#a3a3a3")
                    self.Label5_4.configure(font="-family {Rockwell} -size 11 -slant italic")
                    self.Label5_4.configure(foreground="#f7ff14")
                    self.Label5_4.configure(highlightbackground="#d9d9d9")
                    self.Label5_4.configure(highlightcolor="black")
                    self.Label5_4.configure(text='''Give Out of 5''')

                    self.spin4 = tk.Spinbox(self.Frame2_1, from_=1.0, to=5.0)
                    self.spin4.place(relx=0.544, rely=0.507, relheight=0.054, relwidth=0.146)

                    self.spin4.configure(activebackground="#f9f9f9")
                    self.spin4.configure(background="white")
                    self.spin4.configure(buttonbackground="#a3a5a8")
                    self.spin4.configure(disabledforeground="#a3a3a3")
                    self.spin4.configure(foreground="black")
                    self.spin4.configure(highlightbackground="black")
                    self.spin4.configure(highlightcolor="black")
                    self.spin4.configure(insertbackground="black")
                    self.spin4.configure(selectbackground="#c4c4c4")
                    self.spin4.configure(selectforeground="black")
                    self.spin4.configure(textvariable=onlinesupport_support.spinbox4)

                    self.spin3 = tk.Spinbox(self.Frame2_1, from_=1.0, to=5.0)
                    self.spin3.place(relx=0.544, rely=0.423, relheight=0.054, relwidth=0.146)

                    self.spin3.configure(activebackground="#f9f9f9")
                    self.spin3.configure(background="white")
                    self.spin3.configure(buttonbackground="#a3a5a8")
                    self.spin3.configure(disabledforeground="#a3a3a3")
                    self.spin3.configure(foreground="black")
                    self.spin3.configure(highlightbackground="black")
                    self.spin3.configure(highlightcolor="black")
                    self.spin3.configure(insertbackground="black")
                    self.spin3.configure(selectbackground="#c4c4c4")
                    self.spin3.configure(selectforeground="black")
                    self.spin3.configure(textvariable=onlinesupport_support.spinbox3)

                    self.spin2 = tk.Spinbox(self.Frame2_1, from_=1.0, to=5.0)
                    self.spin2.place(relx=0.544, rely=0.338, relheight=0.054, relwidth=0.146)

                    self.spin2.configure(activebackground="#f9f9f9")
                    self.spin2.configure(background="white")
                    self.spin2.configure(buttonbackground="#a3a5a8")
                    self.spin2.configure(disabledforeground="#a3a3a3")
                    self.spin2.configure(foreground="black")
                    self.spin2.configure(highlightbackground="black")
                    self.spin2.configure(highlightcolor="black")
                    self.spin2.configure(insertbackground="black")
                    self.spin2.configure(selectbackground="#c4c4c4")
                    self.spin2.configure(selectforeground="black")
                    self.spin2.configure(textvariable=onlinesupport_support.spinbox2)

                    self.Label6_9 = tk.Label(self.Frame2_1)
                    self.Label6_9.place(relx=0.252, rely=0.592, height=49, width=284)
                    self.Label6_9.configure(activebackground="#f9f9f9")
                    self.Label6_9.configure(activeforeground="black")
                    self.Label6_9.configure(background="#595959")
                    self.Label6_9.configure(disabledforeground="#a3a3a3")
                    self.Label6_9.configure(font="-family {Franklin Gothic Demi Cond} -size 36")
                    self.Label6_9.configure(foreground="#ff9e1f")
                    self.Label6_9.configure(highlightbackground="#d9d9d9")
                    self.Label6_9.configure(highlightcolor="black")
                    self.Label6_9.configure(text='''Write A Review''')

                    self.txt_review = tk.Entry(self.Frame2_1)
                    self.txt_review.place(relx=0.078, rely=0.732,height=30, relwidth=0.804)
                    self.txt_review.configure(background="white")
                    self.txt_review.configure(disabledforeground="#a3a3a3")
                    self.txt_review.configure(font="TkFixedFont")
                    self.txt_review.configure(foreground="#000000")
                    self.txt_review.configure(highlightbackground="#d9d9d9")
                    self.txt_review.configure(highlightcolor="black")
                    self.txt_review.configure(insertbackground="black")
                    self.txt_review.configure(selectbackground="#c4c4c4")
                    self.txt_review.configure(selectforeground="black")

                    self.btn_submit2 = tk.Button(self.Frame2_1)
                    self.btn_submit2.place(relx=0.408, rely=0.845, height=34, width=80)
                    self.btn_submit2.configure(activebackground="#ececec")
                    self.btn_submit2.configure(activeforeground="#000000")
                    self.btn_submit2.configure(background="#ff350d")
                    self.btn_submit2.configure(borderwidth="10")
                    self.btn_submit2.configure(disabledforeground="#a3a3a3")
                    self.btn_submit2.configure(font="-family {Swatch it} -size 11 -weight bold")
                    self.btn_submit2.configure(foreground="#000000")
                    self.btn_submit2.configure(highlightbackground="#d9d9d9")
                    self.btn_submit2.configure(highlightcolor="black")
                    self.btn_submit2.configure(pady="0")
                    self.btn_submit2.configure(text='''Submit''',command= self.submit2s)

                    self.Label3 = tk.Label(top)
                    self.Label3.place(relx=0.034, rely=0.731, height=176, width=304)
                    self.Label3.configure(activebackground="#f9f9f9")
                    self.Label3.configure(activeforeground="black")
                    self.Label3.configure(background="#d9d9d9")
                    self.Label3.configure(disabledforeground="#a3a3a3")
                    self.Label3.configure(foreground="#000000")
                    self.Label3.configure(highlightbackground="#d9d9d9")
                    self.Label3.configure(highlightcolor="black")
                    photo_location = os.path.join(prog_location,"./online2.png")
                    self._img0 = tk.PhotoImage(file=photo_location)
                    self.Label3.configure(image=self._img0)
                    self.Label3.configure(text='''Label''')

                    self.Frame2_2 = tk.Frame(top)
                    self.Frame2_2.place(relx=0.336, rely=0.731, relheight=0.224
                            , relwidth=0.331)
                    self.Frame2_2.configure(relief='ridge')
                    self.Frame2_2.configure(borderwidth="10")
                    self.Frame2_2.configure(relief='ridge')
                    self.Frame2_2.configure(background="#595959")
                    self.Frame2_2.configure(highlightbackground="#d9d9d9")
                    self.Frame2_2.configure(highlightcolor="black")
                    self.Frame2_2.configure(width=385)

                    self.Label4 = tk.Label(self.Frame2_2)
                    self.Label4.place(relx=0.39, rely=0.114, height=23, width=99)
                    self.Label4.configure(activebackground="#f9f9f9")
                    self.Label4.configure(activeforeground="black")
                    self.Label4.configure(background="#595959")
                    self.Label4.configure(disabledforeground="#a3a3a3")
                    self.Label4.configure(font="-family {SF Distant Galaxy} -size 13")
                    self.Label4.configure(foreground="#ffff12")
                    self.Label4.configure(highlightbackground="#d9d9d9")
                    self.Label4.configure(highlightcolor="black")
                    self.Label4.configure(text='''Contact''')

                    self.btn_facebook = tk.Button(self.Frame2_2)
                    self.btn_facebook.place(relx=0.104, rely=0.343, height=91, width=96)
                    self.btn_facebook.configure(activebackground="#ececec")
                    self.btn_facebook.configure(activeforeground="#000000")
                    self.btn_facebook.configure(background="#d9d9d9")
                    self.btn_facebook.configure(borderwidth="10")
                    self.btn_facebook.configure(disabledforeground="#a3a3a3")
                    self.btn_facebook.configure(foreground="#000000")
                    self.btn_facebook.configure(highlightbackground="#d9d9d9")
                    self.btn_facebook.configure(highlightcolor="black")
                    photo_location = os.path.join(prog_location,"./facebook-logo-png-5a35528eaa4f08.7998622015134439826976.png")
                    self._img1 = tk.PhotoImage(file=photo_location)
                    self.btn_facebook.configure(image=self._img1)
                    self.btn_facebook.configure(overrelief="raised")
                    self.btn_facebook.configure(pady="0")
                    self.btn_facebook.configure(text='''Button''',command= self.facebooks)

                    self.btn_website = tk.Button(self.Frame2_2)
                    self.btn_website.place(relx=0.416, rely=0.343, height=86, width=86)
                    self.btn_website.configure(activebackground="#ececec")
                    self.btn_website.configure(activeforeground="#000000")
                    self.btn_website.configure(background="#d9d9d9")
                    self.btn_website.configure(borderwidth="10")
                    self.btn_website.configure(disabledforeground="#a3a3a3")
                    self.btn_website.configure(foreground="#000000")
                    self.btn_website.configure(highlightbackground="#d9d9d9")
                    self.btn_website.configure(highlightcolor="black")
                    photo_location = os.path.join(prog_location,"./icon_weebly.com_bot_02e090.png")
                    self._img2 = tk.PhotoImage(file=photo_location)
                    self.btn_website.configure(image=self._img2)
                    self.btn_website.configure(pady="0")
                    self.btn_website.configure(text='''Button''',command= self.websites)

                    self.btn_call = tk.Button(self.Frame2_2)
                    self.btn_call.place(relx=0.701, rely=0.343, height=92, width=91)
                    self.btn_call.configure(activebackground="#ececec")
                    self.btn_call.configure(activeforeground="#000000")
                    self.btn_call.configure(background="#d9d9d9")
                    self.btn_call.configure(borderwidth="10")
                    self.btn_call.configure(disabledforeground="#a3a3a3")
                    self.btn_call.configure(foreground="#000000")
                    self.btn_call.configure(highlightbackground="#d9d9d9")
                    self.btn_call.configure(highlightcolor="black")
                    photo_location = os.path.join(prog_location,"./2019-04-25 14_03_03-Downloads.png")
                    self._img3 = tk.PhotoImage(file=photo_location)
                    self.btn_call.configure(image=self._img3)
                    self.btn_call.configure(pady="0")
                    self.btn_call.configure(text='''Button''',command= self.calls)

                    self.Frame2_3 = tk.Frame(top)
                    self.Frame2_3.place(relx=0.68, rely=0.731, relheight=0.25
                            , relwidth=0.297)
                    self.Frame2_3.configure(relief='ridge')
                    self.Frame2_3.configure(borderwidth="10")
                    self.Frame2_3.configure(relief='ridge')
                    self.Frame2_3.configure(background="#595959")
                    self.Frame2_3.configure(highlightbackground="#d9d9d9")
                    self.Frame2_3.configure(highlightcolor="black")
                    self.Frame2_3.configure(width=345)

                    self.Label5 = tk.Label(self.Frame2_3)
                    self.Label5.place(relx=0.232, rely=0.051, height=32, width=172)
                    self.Label5.configure(activebackground="#f9f9f9")
                    self.Label5.configure(activeforeground="black")
                    self.Label5.configure(background="#595959")
                    self.Label5.configure(disabledforeground="#a3a3a3")
                    self.Label5.configure(font="-family {Rockwell} -size 17 -slant italic")
                    self.Label5.configure(foreground="#1fff0f")
                    self.Label5.configure(highlightbackground="#d9d9d9")
                    self.Label5.configure(highlightcolor="black")
                    self.Label5.configure(text='''::: Developers :::''')

                    self.Label5_3 = tk.Label(self.Frame2_3)
                    self.Label5_3.place(relx=0.261, rely=0.205, height=32, width=152)
                    self.Label5_3.configure(activebackground="#f9f9f9")
                    self.Label5_3.configure(activeforeground="black")
                    self.Label5_3.configure(background="#595959")
                    self.Label5_3.configure(disabledforeground="#a3a3a3")
                    self.Label5_3.configure(font="-family {Rockwell} -size 11 -slant italic")
                    self.Label5_3.configure(foreground="aqua")
                    self.Label5_3.configure(highlightbackground="#d9d9d9")
                    self.Label5_3.configure(highlightcolor="black")
                    self.Label5_3.configure(text='''Pranesh Kulkarni''')

                    self.Label5_5 = tk.Label(self.Frame2_3)
                    self.Label5_5.place(relx=0.261, rely=0.564, height=22, width=152)
                    self.Label5_5.configure(activebackground="#f9f9f9")
                    self.Label5_5.configure(activeforeground="black")
                    self.Label5_5.configure(background="#595959")
                    self.Label5_5.configure(disabledforeground="#a3a3a3")
                    self.Label5_5.configure(font="-family {Rockwell} -size 11 -slant italic")
                    self.Label5_5.configure(foreground="aqua")
                    self.Label5_5.configure(highlightbackground="#d9d9d9")
                    self.Label5_5.configure(highlightcolor="black")
                    self.Label5_5.configure(text='''Prathamesh Dhasade''')

                    self.Label5_5 = tk.Label(self.Frame2_3)
                    self.Label5_5.place(relx=0.261, rely=0.667, height=22, width=152)
                    self.Label5_5.configure(activebackground="#f9f9f9")
                    self.Label5_5.configure(activeforeground="black")
                    self.Label5_5.configure(background="#595959")
                    self.Label5_5.configure(disabledforeground="#a3a3a3")
                    self.Label5_5.configure(font="-family {Rockwell} -size 11 -slant italic")
                    self.Label5_5.configure(foreground="aqua")
                    self.Label5_5.configure(highlightbackground="#d9d9d9")
                    self.Label5_5.configure(highlightcolor="black")
                    self.Label5_5.configure(text='''Anjali Gayakwad''')

                    self.Label5_5 = tk.Label(self.Frame2_3)
                    self.Label5_5.place(relx=0.261, rely=0.769, height=32, width=152)
                    self.Label5_5.configure(activebackground="#f9f9f9")
                    self.Label5_5.configure(activeforeground="black")
                    self.Label5_5.configure(background="#595959")
                    self.Label5_5.configure(disabledforeground="#a3a3a3")
                    self.Label5_5.configure(font="-family {Rockwell} -size 11 -slant italic")
                    self.Label5_5.configure(foreground="aqua")
                    self.Label5_5.configure(highlightbackground="#d9d9d9")
                    self.Label5_5.configure(highlightcolor="black")
                    self.Label5_5.configure(text='''Shraddha Kshirsagar''')

                    self.Label5_5 = tk.Label(self.Frame2_3)
                    self.Label5_5.place(relx=0.261, rely=0.462, height=22, width=152)
                    self.Label5_5.configure(activebackground="#f9f9f9")
                    self.Label5_5.configure(activeforeground="black")
                    self.Label5_5.configure(background="#595959")
                    self.Label5_5.configure(disabledforeground="#a3a3a3")
                    self.Label5_5.configure(font="-family {Rockwell} -size 11 -slant italic")
                    self.Label5_5.configure(foreground="aqua")
                    self.Label5_5.configure(highlightbackground="#d9d9d9")
                    self.Label5_5.configure(highlightcolor="black")
                    self.Label5_5.configure(text='''Ankita Chirame''')

                    self.Label5_5 = tk.Label(self.Frame2_3)
                    self.Label5_5.place(relx=0.261, rely=0.359, height=22, width=152)
                    self.Label5_5.configure(activebackground="#f9f9f9")
                    self.Label5_5.configure(activeforeground="black")
                    self.Label5_5.configure(background="#595959")
                    self.Label5_5.configure(disabledforeground="#a3a3a3")
                    self.Label5_5.configure(font="-family {Rockwell} -size 11 -slant italic")
                    self.Label5_5.configure(foreground="aqua")
                    self.Label5_5.configure(highlightbackground="#d9d9d9")
                    self.Label5_5.configure(highlightcolor="black")
                    self.Label5_5.configure(text='''Anamika Rathod''')

                    self.btn_submit2_11 = tk.Button(top)
                    self.btn_submit2_11.place(relx=0.869, rely=0.115, height=44, width=90)
                    self.btn_submit2_11.configure(activebackground="#ececec")
                    self.btn_submit2_11.configure(activeforeground="#000000")
                    self.btn_submit2_11.configure(background="#3045ff")
                    self.btn_submit2_11.configure(borderwidth="10")
                    self.btn_submit2_11.configure(disabledforeground="#a3a3a3")
                    self.btn_submit2_11.configure(font="-family {Swatch it} -size 11 -weight bold")
                    self.btn_submit2_11.configure(foreground="#000000")
                    self.btn_submit2_11.configure(highlightbackground="#d9d9d9")
                    self.btn_submit2_11.configure(highlightcolor="black",command= self.exits)
                    self.btn_submit2_11.configure(pady="0")
                    self.btn_submit2_11.configure(text='''Exit''')

                    self.btn_back = tk.Button(top)
                    self.btn_back.place(relx=0.869, rely=0.051, height=44, width=90)
                    self.btn_back.configure(activebackground="#ececec")
                    self.btn_back.configure(activeforeground="#000000")
                    self.btn_back.configure(background="#2ba30d")
                    self.btn_back.configure(borderwidth="10")
                    self.btn_back.configure(disabledforeground="#a3a3a3")
                    self.btn_back.configure(font="-family {Swatch it} -size 11 -weight bold")
                    self.btn_back.configure(foreground="#000000")
                    self.btn_back.configure(highlightbackground="#d9d9d9")
                    self.btn_back.configure(highlightcolor="black")
                    self.btn_back.configure(pady="0")
                    self.btn_back.configure(text='''Back''',command = self.backs)

            if __name__ == '__main__':
                vp_start_gui()




































        def update1():
            import sys
            import tkinter
            from tkinter import messagebox
            try:
                import Tkinter as tk
            except ImportError:
                import tkinter as tk

            try:
                import ttk
                py3 = False
            except ImportError:
                import tkinter.ttk as ttk
                py3 = True

            import modificationmainpage_support
            import os.path

            def vp_start_gui():
                '''Starting point when module is the main routine.'''
                global val, w, root
                global prog_location
                prog_call = sys.argv[0]
                print ('prog_call = {}'.format(prog_call))
                prog_location = os.path.split(prog_call)[0]
                print ('prog_location = {}'.format(prog_location))
                sys.stdout.flush()
                root = tk.Tk()
                top = e_TAX_2019 (root)
                modificationmainpage_support.init(root, top)
                root.mainloop()

            w = None
            def create_e_TAX_2019(root, *args, **kwargs):
                '''Starting point when module is imported by another program.'''
                global w, w_win, rt
                global prog_location
                prog_call = sys.argv[0]
                print ('prog_call = {}'.format(prog_call))
                prog_location = os.path.split(prog_call)[0]
                print ('prog_location = {}'.format(prog_location))
                rt = root
                w = tk.Toplevel (root)
                top = e_TAX_2019 (w)
                modificationmainpage_support.init(w, top, *args, **kwargs)
                return (w, top)

            def destroy_e_TAX_2019():
                global w
                w.destroy()
                w = None





























            def access1():

                import sys
                import tkinter
                import time
                from tkinter import messagebox
                import smtplib

                try:
                    import Tkinter as tk
                except ImportError:
                    import tkinter as tk

                try:
                    import ttk
                    py3 = False
                except ImportError:
                    import tkinter.ttk as ttk
                    py3 = True

                import permit_support

                def vp_start_gui():
                    '''Starting point when module is the main routine.'''
                    global val, w, root
                    root = tk.Tk()
                    top = e_TAX_2019 (root)
                    permit_support.init(root, top)
                    root.mainloop()

                w = None
                def create_e_TAX_2019(root, *args, **kwargs):
                    '''Starting point when module is imported by another program.'''
                    global w, w_win, rt
                    rt = root
                    w = tk.Toplevel (root)
                    top = e_TAX_2019 (w)
                    permit_support.init(w, top, *args, **kwargs)
                    return (w, top)

                def destroy_e_TAX_2019():
                    global w
                    w.destroy()
                    w = None

                class e_TAX_2019:
                    def exits(self):
                        msg=tkinter.messagebox.askyesno("e-TAX 2019","Do You Want To EXIT ?")
                        if msg :
                            os._exit(1)
                    def backs(self):
                        root.destroy()
                        update1()

                    def submits(self):
                        p=str(self.txt_productkey.get());
                        v=str(self.txt_village.get());
                        u=str(self.txt_username.get());
                        n=str(self.txt_noofentries.get());
                        r=str(self.txt_reason.get());
                        t=str(time.asctime(time.localtime(time.time())))


                        totcontaint = str("Hi there is a new request for updating e-TAX 2019 database \n\n\n here are DETAILS ::::::\n\n\nproduct key :"+p+"\n village name :"+v+"\n Username :"+u+"\n No. of entries to modify :"+n+"\n Reason to modify data :"+r+"\n\n\n\n\n Requested on :"+t)
                        mail= smtplib.SMTP('smtp.gmail.com',587)
                        mail.ehlo()
                        mail.starttls()
                        mail.login('etaxsupp2019@gmail.com','Pass@123')
                        mail.sendmail('etaxsupp2019@gmail.com','kulkarnipranesh1767@gmail.com',totcontaint)
                        mail.close()
                        messagebox.showinfo("e-TAX 2019","Requested for updating Data. Wait for Confirmation from Administrator. After recieving Confirmation Email From Administrator You can modify database through respective modules")



                    def __init__(self, top=None):
                        '''This class configures and populates the toplevel window.
                           top is the toplevel containing window.'''
                        _bgcolor = '#d9d9d9'  # X11 color: 'gray85'
                        _fgcolor = '#000000'  # X11 color: 'black'
                        _compcolor = '#d9d9d9' # X11 color: 'gray85'
                        _ana1color = '#d9d9d9' # X11 color: 'gray85'
                        _ana2color = '#ececec' # Closest X11 color: 'gray92'
                        font12 = "-family {GoudyHandtooled BT} -size 22"
                        font15 = "-family {System} -size 10 -weight bold"
                        font16 = "-family {Segoe UI} -size 12"

                        top.geometry("990x650+273+127")
                        top.title("e-TAX 2019")
                        top.configure(background="#727272")
                        top.configure(highlightbackground="#d9d9d9")
                        top.configure(highlightcolor="black")

                        self.Frame1 = tk.Frame(top)
                        self.Frame1.place(relx=0.141, rely=0.046, relheight=0.223
                                , relwidth=0.712)
                        self.Frame1.configure(relief='ridge')
                        self.Frame1.configure(borderwidth="10")
                        self.Frame1.configure(relief='ridge')
                        self.Frame1.configure(background="#595959")
                        self.Frame1.configure(highlightbackground="#d9d9d9")
                        self.Frame1.configure(highlightcolor="black")
                        self.Frame1.configure(width=705)

                        self.Label1 = tk.Label(self.Frame1)
                        self.Label1.place(relx=0.043, rely=0.207, height=59, width=196)
                        self.Label1.configure(activebackground="#f9f9f9")
                        self.Label1.configure(activeforeground="black")
                        self.Label1.configure(background="#595959")
                        self.Label1.configure(disabledforeground="#a3a3a3")
                        self.Label1.configure(font="-family {Rockwell Extra} -size 40 -weight bold")
                        self.Label1.configure(foreground="#0d4dff")
                        self.Label1.configure(highlightbackground="#d9d9d9")
                        self.Label1.configure(highlightcolor="black")
                        self.Label1.configure(text='''eTAX''')

                        self.Label1_1 = tk.Label(self.Frame1)
                        self.Label1_1.place(relx=0.298, rely=0.207, height=59, width=146)
                        self.Label1_1.configure(activebackground="#f9f9f9")
                        self.Label1_1.configure(activeforeground="black")
                        self.Label1_1.configure(background="#595959")
                        self.Label1_1.configure(disabledforeground="#a3a3a3")
                        self.Label1_1.configure(font="-family {Rockwell Extra} -size 40 -weight bold")
                        self.Label1_1.configure(foreground="#ff2b0a")
                        self.Label1_1.configure(highlightbackground="#d9d9d9")
                        self.Label1_1.configure(highlightcolor="black")
                        self.Label1_1.configure(text='''2019''')

                        self.Label2 = tk.Label(self.Frame1)
                        self.Label2.place(relx=0.156, rely=0.621, height=31, width=184)
                        self.Label2.configure(activebackground="#f9f9f9")
                        self.Label2.configure(activeforeground="black")
                        self.Label2.configure(background="#595959")
                        self.Label2.configure(disabledforeground="#a3a3a3")
                        self.Label2.configure(font="-family {Rage Italic} -size 19 -slant italic")
                        self.Label2.configure(foreground="#f7ff0a")
                        self.Label2.configure(highlightbackground="#d9d9d9")
                        self.Label2.configure(highlightcolor="black")
                        self.Label2.configure(text='''working for you''')

                        self.Label1_2 = tk.Label(self.Frame1)
                        self.Label1_2.place(relx=0.525, rely=0.207, height=79, width=306)
                        self.Label1_2.configure(activebackground="#f9f9f9")
                        self.Label1_2.configure(activeforeground="black")
                        self.Label1_2.configure(background="#595959")
                        self.Label1_2.configure(disabledforeground="#a3a3a3")
                        self.Label1_2.configure(font="-family {Rockwell Extra} -size 28 -weight bold")
                        self.Label1_2.configure(foreground="#5faa14")
                        self.Label1_2.configure(highlightbackground="#d9d9d9")
                        self.Label1_2.configure(highlightcolor="black")
                        self.Label1_2.configure(text='''PERMISSION''')
                        self.Label1_2.configure(width=306)

                        self.Frame1_1 = tk.Frame(top)
                        self.Frame1_1.place(relx=0.141, rely=0.308, relheight=0.577
                                , relwidth=0.712)
                        self.Frame1_1.configure(relief='ridge')
                        self.Frame1_1.configure(borderwidth="10")
                        self.Frame1_1.configure(relief='ridge')
                        self.Frame1_1.configure(background="#595959")
                        self.Frame1_1.configure(highlightbackground="#d9d9d9")
                        self.Frame1_1.configure(highlightcolor="black")
                        self.Frame1_1.configure(width=705)

                        self.Label3 = tk.Label(self.Frame1_1)
                        self.Label3.place(relx=0.156, rely=0.08, height=41, width=446)
                        self.Label3.configure(background="#595959")
                        self.Label3.configure(disabledforeground="#a3a3a3")
                        self.Label3.configure(font=font12)
                        self.Label3.configure(foreground="#fbf2ff")
                        self.Label3.configure(text='''FILL YOUR PRODUCT DETAILS''')

                        self.Label4 = tk.Label(self.Frame1_1)
                        self.Label4.place(relx=0.142, rely=0.24, height=22, width=102)
                        self.Label4.configure(background="#595959")
                        self.Label4.configure(disabledforeground="#a3a3a3")
                        self.Label4.configure(font=font15)
                        self.Label4.configure(foreground="AQUA")
                        self.Label4.configure(text='''Product Key :''')
                        self.Label4.configure(width=102)

                        self.Label4_5 = tk.Label(self.Frame1_1)
                        self.Label4_5.place(relx=0.142, rely=0.32, height=22, width=102)
                        self.Label4_5.configure(activebackground="#f9f9f9")
                        self.Label4_5.configure(activeforeground="black")
                        self.Label4_5.configure(background="#595959")
                        self.Label4_5.configure(disabledforeground="#a3a3a3")
                        self.Label4_5.configure(font="-family {System} -size 10 -weight bold")
                        self.Label4_5.configure(foreground="AQUA")
                        self.Label4_5.configure(highlightbackground="#d9d9d9")
                        self.Label4_5.configure(highlightcolor="black")
                        self.Label4_5.configure(text='''Village :''')

                        self.Label4_6 = tk.Label(self.Frame1_1)
                        self.Label4_6.place(relx=0.142, rely=0.4, height=22, width=102)
                        self.Label4_6.configure(activebackground="#f9f9f9")
                        self.Label4_6.configure(activeforeground="black")
                        self.Label4_6.configure(background="#595959")
                        self.Label4_6.configure(disabledforeground="#a3a3a3")
                        self.Label4_6.configure(font="-family {System} -size 10 -weight bold")
                        self.Label4_6.configure(foreground="AQUA")
                        self.Label4_6.configure(highlightbackground="#d9d9d9")
                        self.Label4_6.configure(highlightcolor="black")
                        self.Label4_6.configure(text='''Username :''')

                        self.Label4_7 = tk.Label(self.Frame1_1)
                        self.Label4_7.place(relx=0.085, rely=0.48, height=22, width=202)
                        self.Label4_7.configure(activebackground="#f9f9f9")
                        self.Label4_7.configure(activeforeground="black")
                        self.Label4_7.configure(background="#595959")
                        self.Label4_7.configure(disabledforeground="#a3a3a3")
                        self.Label4_7.configure(font="-family {System} -size 10 -weight bold")
                        self.Label4_7.configure(foreground="AQUA")
                        self.Label4_7.configure(highlightbackground="#d9d9d9")
                        self.Label4_7.configure(highlightcolor="black")
                        self.Label4_7.configure(text='''No. of entries to update :''')
                        self.Label4_7.configure(width=202)

                        self.Label4_8 = tk.Label(self.Frame1_1)
                        self.Label4_8.place(relx=0.099, rely=0.56, height=22, width=172)
                        self.Label4_8.configure(activebackground="#f9f9f9")
                        self.Label4_8.configure(activeforeground="black")
                        self.Label4_8.configure(background="#595959")
                        self.Label4_8.configure(disabledforeground="#a3a3a3")
                        self.Label4_8.configure(font="-family {System} -size 10 -weight bold")
                        self.Label4_8.configure(foreground="AQUA")
                        self.Label4_8.configure(highlightbackground="#d9d9d9")
                        self.Label4_8.configure(highlightcolor="black")
                        self.Label4_8.configure(text='''Reason to modify data :''')
                        self.Label4_8.configure(width=172)

                        self.txt_productkey = tk.Entry(self.Frame1_1)
                        self.txt_productkey.place(relx=0.397, rely=0.24, height=20
                                , relwidth=0.445)
                        self.txt_productkey.configure(background="white")
                        self.txt_productkey.configure(disabledforeground="#a3a3a3")
                        self.txt_productkey.configure(font="TkFixedFont")
                        self.txt_productkey.configure(foreground="#000000")
                        self.txt_productkey.configure(insertbackground="black")
                        self.txt_productkey.configure(width=314)

                        self.txt_reason = tk.Entry(self.Frame1_1)
                        self.txt_reason.place(relx=0.397, rely=0.56,height=20, relwidth=0.445)
                        self.txt_reason.configure(background="white")
                        self.txt_reason.configure(disabledforeground="#a3a3a3")
                        self.txt_reason.configure(font="TkFixedFont")
                        self.txt_reason.configure(foreground="#000000")
                        self.txt_reason.configure(highlightbackground="#d9d9d9")
                        self.txt_reason.configure(highlightcolor="black")
                        self.txt_reason.configure(insertbackground="black")
                        self.txt_reason.configure(selectbackground="#c4c4c4")
                        self.txt_reason.configure(selectforeground="black")

                        self.txt_noofentries = tk.Entry(self.Frame1_1)
                        self.txt_noofentries.place(relx=0.397, rely=0.48, height=20
                                , relwidth=0.445)
                        self.txt_noofentries.configure(background="white")
                        self.txt_noofentries.configure(disabledforeground="#a3a3a3")
                        self.txt_noofentries.configure(font="TkFixedFont")
                        self.txt_noofentries.configure(foreground="#000000")
                        self.txt_noofentries.configure(highlightbackground="#d9d9d9")
                        self.txt_noofentries.configure(highlightcolor="black")
                        self.txt_noofentries.configure(insertbackground="black")
                        self.txt_noofentries.configure(selectbackground="#c4c4c4")
                        self.txt_noofentries.configure(selectforeground="black")

                        self.txt_username = tk.Entry(self.Frame1_1)
                        self.txt_username.place(relx=0.397, rely=0.4,height=20, relwidth=0.445)
                        self.txt_username.configure(background="white")
                        self.txt_username.configure(disabledforeground="#a3a3a3")
                        self.txt_username.configure(font="TkFixedFont")
                        self.txt_username.configure(foreground="#000000")
                        self.txt_username.configure(highlightbackground="#d9d9d9")
                        self.txt_username.configure(highlightcolor="black")
                        self.txt_username.configure(insertbackground="black")
                        self.txt_username.configure(selectbackground="#c4c4c4")
                        self.txt_username.configure(selectforeground="black")

                        self.txt_village = tk.Entry(self.Frame1_1)
                        self.txt_village.place(relx=0.397, rely=0.32,height=20, relwidth=0.445)
                        self.txt_village.configure(background="white")
                        self.txt_village.configure(disabledforeground="#a3a3a3")
                        self.txt_village.configure(font="TkFixedFont")
                        self.txt_village.configure(foreground="#000000")
                        self.txt_village.configure(highlightbackground="#d9d9d9")
                        self.txt_village.configure(highlightcolor="black")
                        self.txt_village.configure(insertbackground="black")
                        self.txt_village.configure(selectbackground="#c4c4c4")
                        self.txt_village.configure(selectforeground="black")

                        self.btn_submit = tk.Button(self.Frame1_1)
                        self.btn_submit.place(relx=0.454, rely=0.667, height=44, width=227)
                        self.btn_submit.configure(activebackground="#ececec")
                        self.btn_submit.configure(activeforeground="#000000")
                        self.btn_submit.configure(background="#3028a0")
                        self.btn_submit.configure(borderwidth="10")
                        self.btn_submit.configure(disabledforeground="#a3a3a3")
                        self.btn_submit.configure(font=font16)
                        self.btn_submit.configure(foreground="#ffffff")
                        self.btn_submit.configure(highlightbackground="#d9d9d9")
                        self.btn_submit.configure(highlightcolor="black")
                        self.btn_submit.configure(pady="0")
                        self.btn_submit.configure(text='''Ask for permission''',command = self.submits)
                        self.btn_submit.configure(width=227)

                        self.btn_exit = tk.Button(top)
                        self.btn_exit.place(relx=0.879, rely=0.169, height=44, width=107)
                        self.btn_exit.configure(activebackground="#ececec")
                        self.btn_exit.configure(activeforeground="#000000")
                        self.btn_exit.configure(background="#ff002b")
                        self.btn_exit.configure(borderwidth="10")
                        self.btn_exit.configure(disabledforeground="#a3a3a3")
                        self.btn_exit.configure(font="-family {Segoe UI} -size 12")
                        self.btn_exit.configure(foreground="#ffffff")
                        self.btn_exit.configure(highlightbackground="#d9d9d9")
                        self.btn_exit.configure(highlightcolor="black")
                        self.btn_exit.configure(pady="0")
                        self.btn_exit.configure(text='''EXIT''',command = self.exits)
                        self.btn_exit.configure(width=107)

                        self.btn_back = tk.Button(top)
                        self.btn_back.place(relx=0.879, rely=0.077, height=44, width=107)
                        self.btn_back.configure(activebackground="#ececec")
                        self.btn_back.configure(activeforeground="#000000")
                        self.btn_back.configure(background="#39a324")
                        self.btn_back.configure(borderwidth="10")
                        self.btn_back.configure(disabledforeground="#a3a3a3")
                        self.btn_back.configure(font="-family {Segoe UI} -size 12")
                        self.btn_back.configure(foreground="#ffffff")
                        self.btn_back.configure(highlightbackground="#d9d9d9")
                        self.btn_back.configure(highlightcolor="black")
                        self.btn_back.configure(pady="0")
                        self.btn_back.configure(text='''BACK''',command = self.backs)

                if __name__ == '__main__':
                    vp_start_gui()










































            def modify1():

                import sys
                import tkinter
                from tkinter import messagebox
                import mysql.connector
                import dbConnect
                import mysql.connector
                from dbConnect import DBConnect
                import time
                import smtplib


                try:
                    import Tkinter as tk
                except ImportError:
                    import tkinter as tk

                try:
                    import ttk
                    py3 = False
                except ImportError:
                    import tkinter.ttk as ttk
                    py3 = True

                import modifymod_support

                def vp_start_gui():
                    '''Starting point when module is the main routine.'''
                    global val, w, root
                    root = tk.Tk()
                    top = Toplevel1 (root)
                    modifymod_support.init(root, top)
                    root.mainloop()

                w = None
                def create_Toplevel1(root, *args, **kwargs):
                    '''Starting point when module is imported by another program.'''
                    global w, w_win, rt
                    rt = root
                    w = tk.Toplevel (root)
                    top = Toplevel1 (w)
                    modifymod_support.init(w, top, *args, **kwargs)
                    return (w, top)

                def destroy_Toplevel1():
                    global w
                    w.destroy()
                    w = None

                class Toplevel1:

                    def exits(self):
                        msg=tkinter.messagebox.askyesno('etax-2019','Do You Want To Exit ?');
                        if(msg):
                            os._exit(1)

                    def backs(self):
                        root.destroy()
                        update1()



                    def submit4(self):
                        a=str(self.txt_idnumber.get());
                        b=str(self.txt_name.get());
                        c=str(self.txt_meternumber.get());
                        d=str(self.txt_wardnumber.get());
                        e=str(self.txt_house.get());
                        f=str(self.txt_health.get());
                        g=str(self.txt_light.get());
                        h=str(self.txt_water.get());
                        i=str(self.txt_total.get());
                        j=str(self.txt_reciept.get());
                        k=str(self.txt_housepaid.get());
                        l=str(self.txt_healthpaid.get());
                        m=str(self.txt_lightpaid.get());
                        n=str(self.txt_waterpaid.get());
                        o=str(self.txt_totalpaid.get());
                        v=str(self.txt_village.get());



                        localtime=str(time.asctime(time.localtime(time.time())))

                        t=int(i)
                        u=int(o)
                        p=str(t-u)


                        database = DBConnect(host='remotemysql.com',user='NhpDbfVMT3',password='NXWSVJNu55',database='NhpDbfVMT3',port='3306')
                        new_user = {'village':v,'idnumber': a,'meternumber': c,'wardnumber': d,'name': b,'housetax': e,'healthtax': f,'lighttax': g,'watertax': h,'total': i,'reciptnumber':j,'housetaxpaid':k,'healthtaxpaid':l,'lighttaxpaid':m,'watertaxpaid':n,'totalpaid':o,'rest':p}
                        database.insert(new_user,'updateddata')
                        database.commit()

                        content1="WARNING   !!!!!! \n\n\n     Hi there is a New request to modify eTAX 2019 database \n Here are details: \n Village Name : "+v+"\n\n"
                        content2="\nReciept Number :"+j+"\nName :"+b+"\nID Number :"+a+"\nMeter Number :"+c+"\nWard Number :"+d+"\nHouse Tax :"+e+"\nHealth Tax :"+f+"\nLight Tax :"+g+"\nWater Tax :"+h+"\nTotal Ammount of tax to be paid :"+i+"\nPaid House Tax :"+k+"\nPaid Health Tax :"+l+"\nPaid Light Tax :"+m+"\nPiad Water Tax :"+n+"\nTotal Tax paid  :"+o
                        totcontaint=content1+content2 


                        mail= smtplib.SMTP('smtp.gmail.com',587)
                        mail.ehlo()
                        mail.starttls()
                        mail.login('etaxsupp2019@gmail.com','Pass@123')
                        mail.sendmail('etaxsupp2019@gmail.com','kulkarnipranesh1767@gmail.com',totcontaint)
                        mail.close()
                        messagebox.showinfo("etax-2019","Data Modified Successfully")
                        root.destroy()
                        extra1()





                    def __init__(self, top=None):
                        '''This class configures and populates the toplevel window.
                           top is the toplevel containing window.'''
                        _bgcolor = '#d9d9d9'  # X11 color: 'gray85'
                        _fgcolor = '#000000'  # X11 color: 'black'
                        _compcolor = '#d9d9d9' # X11 color: 'gray85'
                        _ana1color = '#d9d9d9' # X11 color: 'gray85'
                        _ana2color = '#ececec' # Closest X11 color: 'gray92'

                        top.geometry("1043x729+179+114")
                        top.title("New Toplevel")
                        top.configure(background="#515154")
                        top.configure(highlightbackground="#d9d9d9")
                        top.configure(highlightcolor="black")

                        self.Label1 = tk.Label(top)
                        self.Label1.place(relx=0.019, rely=0.027, height=81, width=156)
                        self.Label1.configure(activebackground="#f9f9f9")
                        self.Label1.configure(activeforeground="black")
                        self.Label1.configure(background="#515154")
                        self.Label1.configure(disabledforeground="#a3a3a3")
                        self.Label1.configure(font="-family {Britannic Bold} -size 48 -weight bold")
                        self.Label1.configure(foreground="#ff250d")
                        self.Label1.configure(highlightbackground="#d9d9d9")
                        self.Label1.configure(highlightcolor="black")
                        self.Label1.configure(text='''eTAX''')

                        self.Label1_1 = tk.Label(top)
                        self.Label1_1.place(relx=0.173, rely=0.027, height=81, width=156)
                        self.Label1_1.configure(activebackground="#f9f9f9")
                        self.Label1_1.configure(activeforeground="black")
                        self.Label1_1.configure(background="#515154")
                        self.Label1_1.configure(disabledforeground="#a3a3a3")
                        self.Label1_1.configure(font="-family {Britannic Bold} -size 48 -weight bold")
                        self.Label1_1.configure(foreground="#2212ff")
                        self.Label1_1.configure(highlightbackground="#d9d9d9")
                        self.Label1_1.configure(highlightcolor="black")
                        self.Label1_1.configure(text='''2019''')

                        self.Label2 = tk.Label(top)
                        self.Label2.place(relx=0.105, rely=0.123, height=31, width=141)
                        self.Label2.configure(activebackground="#f9f9f9")
                        self.Label2.configure(activeforeground="black")
                        self.Label2.configure(background="#515154")
                        self.Label2.configure(disabledforeground="#a3a3a3")
                        self.Label2.configure(font="-family {Segoe Script} -size 12 -slant italic")
                        self.Label2.configure(foreground="#f7ff0d")
                        self.Label2.configure(highlightbackground="#d9d9d9")
                        self.Label2.configure(highlightcolor="black")
                        self.Label2.configure(text='''working for you''')

                        self.backbutton = tk.Button(top)
                        self.backbutton.place(relx=0.058, rely=0.192, height=44, width=97)
                        self.backbutton.configure(activebackground="#ececec")
                        self.backbutton.configure(activeforeground="#000000")
                        self.backbutton.configure(background="#120bd8")
                        self.backbutton.configure(borderwidth="10")
                        self.backbutton.configure(disabledforeground="#a3a3a3")
                        self.backbutton.configure(font="-family {Rockwell Extra Bold} -size 12 -weight bold")
                        self.backbutton.configure(foreground="#fcffff")
                        self.backbutton.configure(highlightbackground="#d9d9d9")
                        self.backbutton.configure(highlightcolor="black")
                        self.backbutton.configure(pady="0")
                        self.backbutton.configure(text='''Back''',command = self.backs)

                        self.exit = tk.Button(top)
                        self.exit.place(relx=0.163, rely=0.192, height=44, width=97)
                        self.exit.configure(activebackground="#ececec")
                        self.exit.configure(activeforeground="#000000")
                        self.exit.configure(background="#120bd8")
                        self.exit.configure(borderwidth="10")
                        self.exit.configure(disabledforeground="#a3a3a3")
                        self.exit.configure(font="-family {Rockwell Extra Bold} -size 12 -weight bold")
                        self.exit.configure(foreground="#fcffff")
                        self.exit.configure(highlightbackground="#d9d9d9")
                        self.exit.configure(highlightcolor="black")
                        self.exit.configure(pady="0")
                        self.exit.configure(text='''Exit''',command = self.exits)

                        self.Label4 = tk.Label(top)
                        self.Label4.place(relx=0.326, rely=0.041, height=68, width=699)
                        self.Label4.configure(activebackground="#f9f9f9")
                        self.Label4.configure(activeforeground="black")
                        self.Label4.configure(background="#515154")
                        self.Label4.configure(disabledforeground="#36911a")
                        self.Label4.configure(font="-family {Rockwell Extra Bold} -size 40 -weight bold")
                        self.Label4.configure(foreground="#fff71c")
                        self.Label4.configure(highlightbackground="#d9d9d9")
                        self.Label4.configure(highlightcolor="black")
                        self.Label4.configure(text='''DATA MODIFICATION''')
                        self.Label4.configure(width=699)

                        self.Frame1 = tk.Frame(top)
                        self.Frame1.place(relx=0.24, rely=0.274, relheight=0.569, relwidth=0.292)

                        self.Frame1.configure(relief='ridge')
                        self.Frame1.configure(borderwidth="10")
                        self.Frame1.configure(relief='ridge')
                        self.Frame1.configure(background="#50e82a")
                        self.Frame1.configure(highlightbackground="#d9d9d9")
                        self.Frame1.configure(highlightcolor="black")
                        self.Frame1.configure(width=305)

                        self.txt_idnumber = tk.Entry(self.Frame1)
                        self.txt_idnumber.place(relx=0.426, rely=0.096, height=20
                                , relwidth=0.538)
                        self.txt_idnumber.configure(background="white")
                        self.txt_idnumber.configure(disabledforeground="#a3a3a3")
                        self.txt_idnumber.configure(font="TkFixedFont")
                        self.txt_idnumber.configure(foreground="#000000")
                        self.txt_idnumber.configure(highlightbackground="#d9d9d9")
                        self.txt_idnumber.configure(highlightcolor="black")
                        self.txt_idnumber.configure(insertbackground="black")
                        self.txt_idnumber.configure(selectbackground="#c4c4c4")
                        self.txt_idnumber.configure(selectforeground="black")

                        self.Label7_1 = tk.Label(self.Frame1)
                        self.Label7_1.place(relx=0.033, rely=0.169, height=39, width=106)
                        self.Label7_1.configure(activebackground="#f9f9f9")
                        self.Label7_1.configure(activeforeground="black")
                        self.Label7_1.configure(background="#50e82a")
                        self.Label7_1.configure(disabledforeground="#a3a3a3")
                        self.Label7_1.configure(font="-family {Plantagenet Cherokee} -size 13")
                        self.Label7_1.configure(foreground="#000000")
                        self.Label7_1.configure(highlightbackground="#d9d9d9")
                        self.Label7_1.configure(highlightcolor="black")
                        self.Label7_1.configure(text='''Name :''')

                        self.Label7_2 = tk.Label(self.Frame1)
                        self.Label7_2.place(relx=0.033, rely=0.289, height=29, width=116)
                        self.Label7_2.configure(activebackground="#f9f9f9")
                        self.Label7_2.configure(activeforeground="black")
                        self.Label7_2.configure(background="#50e82a")
                        self.Label7_2.configure(disabledforeground="#a3a3a3")
                        self.Label7_2.configure(font="-family {Plantagenet Cherokee} -size 13")
                        self.Label7_2.configure(foreground="#000000")
                        self.Label7_2.configure(highlightbackground="#d9d9d9")
                        self.Label7_2.configure(highlightcolor="black")
                        self.Label7_2.configure(text='''Meter Number :''')

                        self.Label7_3 = tk.Label(self.Frame1)
                        self.Label7_3.place(relx=0.033, rely=0.361, height=39, width=116)
                        self.Label7_3.configure(activebackground="#f9f9f9")
                        self.Label7_3.configure(activeforeground="black")
                        self.Label7_3.configure(background="#50e82a")
                        self.Label7_3.configure(disabledforeground="#a3a3a3")
                        self.Label7_3.configure(font="-family {Plantagenet Cherokee} -size 13")
                        self.Label7_3.configure(foreground="#000000")
                        self.Label7_3.configure(highlightbackground="#d9d9d9")
                        self.Label7_3.configure(highlightcolor="black")
                        self.Label7_3.configure(text='''Ward Number :''')

                        self.Label7_4 = tk.Label(self.Frame1)
                        self.Label7_4.place(relx=0.033, rely=0.458, height=39, width=106)
                        self.Label7_4.configure(activebackground="#f9f9f9")
                        self.Label7_4.configure(activeforeground="black")
                        self.Label7_4.configure(background="#50e82a")
                        self.Label7_4.configure(disabledforeground="#a3a3a3")
                        self.Label7_4.configure(font="-family {Plantagenet Cherokee} -size 13")
                        self.Label7_4.configure(foreground="#000000")
                        self.Label7_4.configure(highlightbackground="#d9d9d9")
                        self.Label7_4.configure(highlightcolor="black")
                        self.Label7_4.configure(text='''House Tax :''')

                        self.Label7_5 = tk.Label(self.Frame1)
                        self.Label7_5.place(relx=0.033, rely=0.554, height=39, width=106)
                        self.Label7_5.configure(activebackground="#f9f9f9")
                        self.Label7_5.configure(activeforeground="black")
                        self.Label7_5.configure(background="#50e82a")
                        self.Label7_5.configure(disabledforeground="#a3a3a3")
                        self.Label7_5.configure(font="-family {Plantagenet Cherokee} -size 13")
                        self.Label7_5.configure(foreground="#000000")
                        self.Label7_5.configure(highlightbackground="#d9d9d9")
                        self.Label7_5.configure(highlightcolor="black")
                        self.Label7_5.configure(text='''Health Tax :''')

                        self.Label7_6 = tk.Label(self.Frame1)
                        self.Label7_6.place(relx=0.033, rely=0.651, height=39, width=106)
                        self.Label7_6.configure(activebackground="#f9f9f9")
                        self.Label7_6.configure(activeforeground="black")
                        self.Label7_6.configure(background="#50e82a")
                        self.Label7_6.configure(disabledforeground="#a3a3a3")
                        self.Label7_6.configure(font="-family {Plantagenet Cherokee} -size 13")
                        self.Label7_6.configure(foreground="#000000")
                        self.Label7_6.configure(highlightbackground="#d9d9d9")
                        self.Label7_6.configure(highlightcolor="black")
                        self.Label7_6.configure(text='''Light Tax :''')

                        self.Label7_7 = tk.Label(self.Frame1)
                        self.Label7_7.place(relx=0.033, rely=0.747, height=39, width=106)
                        self.Label7_7.configure(activebackground="#f9f9f9")
                        self.Label7_7.configure(activeforeground="black")
                        self.Label7_7.configure(background="#50e82a")
                        self.Label7_7.configure(disabledforeground="#a3a3a3")
                        self.Label7_7.configure(font="-family {Plantagenet Cherokee} -size 13")
                        self.Label7_7.configure(foreground="#000000")
                        self.Label7_7.configure(highlightbackground="#d9d9d9")
                        self.Label7_7.configure(highlightcolor="black")
                        self.Label7_7.configure(text='''Water Tax :''')

                        self.Label7_8 = tk.Label(self.Frame1)
                        self.Label7_8.place(relx=0.033, rely=0.843, height=39, width=106)
                        self.Label7_8.configure(activebackground="#f9f9f9")
                        self.Label7_8.configure(activeforeground="black")
                        self.Label7_8.configure(background="#50e82a")
                        self.Label7_8.configure(disabledforeground="#a3a3a3")
                        self.Label7_8.configure(font="-family {Plantagenet Cherokee} -size 13")
                        self.Label7_8.configure(foreground="#000000")
                        self.Label7_8.configure(highlightbackground="#d9d9d9")
                        self.Label7_8.configure(highlightcolor="black")
                        self.Label7_8.configure(text='''Total :''')

                        self.txt_name = tk.Entry(self.Frame1)
                        self.txt_name.place(relx=0.426, rely=0.193,height=20, relwidth=0.538)
                        self.txt_name.configure(background="white")
                        self.txt_name.configure(disabledforeground="#a3a3a3")
                        self.txt_name.configure(font="TkFixedFont")
                        self.txt_name.configure(foreground="#000000")
                        self.txt_name.configure(highlightbackground="#d9d9d9")
                        self.txt_name.configure(highlightcolor="black")
                        self.txt_name.configure(insertbackground="black")
                        self.txt_name.configure(selectbackground="#c4c4c4")
                        self.txt_name.configure(selectforeground="black")

                        self.txt_meternumber = tk.Entry(self.Frame1)
                        self.txt_meternumber.place(relx=0.426, rely=0.289, height=20
                                , relwidth=0.505)
                        self.txt_meternumber.configure(background="white")
                        self.txt_meternumber.configure(disabledforeground="#a3a3a3")
                        self.txt_meternumber.configure(font="TkFixedFont")
                        self.txt_meternumber.configure(foreground="#000000")
                        self.txt_meternumber.configure(highlightbackground="#d9d9d9")
                        self.txt_meternumber.configure(highlightcolor="black")
                        self.txt_meternumber.configure(insertbackground="black")
                        self.txt_meternumber.configure(selectbackground="#c4c4c4")
                        self.txt_meternumber.configure(selectforeground="black")

                        self.txt_wardnumber = tk.Entry(self.Frame1)
                        self.txt_wardnumber.place(relx=0.426, rely=0.361, height=20
                                , relwidth=0.505)
                        self.txt_wardnumber.configure(background="white")
                        self.txt_wardnumber.configure(disabledforeground="#a3a3a3")
                        self.txt_wardnumber.configure(font="TkFixedFont")
                        self.txt_wardnumber.configure(foreground="#000000")
                        self.txt_wardnumber.configure(highlightbackground="#d9d9d9")
                        self.txt_wardnumber.configure(highlightcolor="black")
                        self.txt_wardnumber.configure(insertbackground="black")
                        self.txt_wardnumber.configure(selectbackground="#c4c4c4")
                        self.txt_wardnumber.configure(selectforeground="black")

                        self.txt_house = tk.Entry(self.Frame1)
                        self.txt_house.place(relx=0.393, rely=0.482,height=20, relwidth=0.538)
                        self.txt_house.configure(background="white")
                        self.txt_house.configure(disabledforeground="#a3a3a3")
                        self.txt_house.configure(font="TkFixedFont")
                        self.txt_house.configure(foreground="#000000")
                        self.txt_house.configure(highlightbackground="#d9d9d9")
                        self.txt_house.configure(highlightcolor="black")
                        self.txt_house.configure(insertbackground="black")
                        self.txt_house.configure(selectbackground="#c4c4c4")
                        self.txt_house.configure(selectforeground="black")

                        self.txt_total = tk.Entry(self.Frame1)
                        self.txt_total.place(relx=0.393, rely=0.867,height=20, relwidth=0.538)
                        self.txt_total.configure(background="white")
                        self.txt_total.configure(disabledforeground="#a3a3a3")
                        self.txt_total.configure(font="TkFixedFont")
                        self.txt_total.configure(foreground="#000000")
                        self.txt_total.configure(highlightbackground="#d9d9d9")
                        self.txt_total.configure(highlightcolor="black")
                        self.txt_total.configure(insertbackground="black")
                        self.txt_total.configure(selectbackground="#c4c4c4")
                        self.txt_total.configure(selectforeground="black")

                        self.txt_light = tk.Entry(self.Frame1)
                        self.txt_light.place(relx=0.393, rely=0.675,height=20, relwidth=0.538)
                        self.txt_light.configure(background="white")
                        self.txt_light.configure(disabledforeground="#a3a3a3")
                        self.txt_light.configure(font="TkFixedFont")
                        self.txt_light.configure(foreground="#000000")
                        self.txt_light.configure(highlightbackground="#d9d9d9")
                        self.txt_light.configure(highlightcolor="black")
                        self.txt_light.configure(insertbackground="black")
                        self.txt_light.configure(selectbackground="#c4c4c4")
                        self.txt_light.configure(selectforeground="black")

                        self.txt_health = tk.Entry(self.Frame1)
                        self.txt_health.place(relx=0.393, rely=0.578,height=20, relwidth=0.538)
                        self.txt_health.configure(background="white")
                        self.txt_health.configure(disabledforeground="#a3a3a3")
                        self.txt_health.configure(font="TkFixedFont")
                        self.txt_health.configure(foreground="#000000")
                        self.txt_health.configure(highlightbackground="#d9d9d9")
                        self.txt_health.configure(highlightcolor="black")
                        self.txt_health.configure(insertbackground="black")
                        self.txt_health.configure(selectbackground="#c4c4c4")
                        self.txt_health.configure(selectforeground="black")

                        self.txt_water = tk.Entry(self.Frame1)
                        self.txt_water.place(relx=0.393, rely=0.771,height=20, relwidth=0.538)
                        self.txt_water.configure(background="white")
                        self.txt_water.configure(disabledforeground="#a3a3a3")
                        self.txt_water.configure(font="TkFixedFont")
                        self.txt_water.configure(foreground="#000000")
                        self.txt_water.configure(highlightbackground="#d9d9d9")
                        self.txt_water.configure(highlightcolor="black")
                        self.txt_water.configure(insertbackground="black")
                        self.txt_water.configure(selectbackground="#c4c4c4")
                        self.txt_water.configure(selectforeground="black")

                        self.Label7 = tk.Label(self.Frame1)
                        self.Label7.place(relx=0.066, rely=0.072, height=39, width=106)
                        self.Label7.configure(activebackground="#f9f9f9")
                        self.Label7.configure(activeforeground="black")
                        self.Label7.configure(background="#50e82a")
                        self.Label7.configure(disabledforeground="#a3a3a3")
                        self.Label7.configure(font="-family {Plantagenet Cherokee} -size 13")
                        self.Label7.configure(foreground="#000000")
                        self.Label7.configure(highlightbackground="#d9d9d9")
                        self.Label7.configure(highlightcolor="black")
                        self.Label7.configure(text='''ID Number :''')

                        self.Frame1_20 = tk.Frame(top)
                        self.Frame1_20.place(relx=0.537, rely=0.37, relheight=0.473
                                , relwidth=0.292)
                        self.Frame1_20.configure(relief='ridge')
                        self.Frame1_20.configure(borderwidth="10")
                        self.Frame1_20.configure(relief='ridge')
                        self.Frame1_20.configure(background="#50e82a")
                        self.Frame1_20.configure(highlightbackground="#d9d9d9")
                        self.Frame1_20.configure(highlightcolor="black")

                        self.Label7_21 = tk.Label(self.Frame1_20)
                        self.Label7_21.place(relx=0.033, rely=0.087, height=39, width=126)
                        self.Label7_21.configure(activebackground="#f9f9f9")
                        self.Label7_21.configure(activeforeground="black")
                        self.Label7_21.configure(background="#50e82a")
                        self.Label7_21.configure(disabledforeground="#a3a3a3")
                        self.Label7_21.configure(font="-family {Plantagenet Cherokee} -size 13")
                        self.Label7_21.configure(foreground="#000000")
                        self.Label7_21.configure(highlightbackground="#d9d9d9")
                        self.Label7_21.configure(highlightcolor="black")
                        self.Label7_21.configure(text='''Reciept Number :''')

                        self.txt_reciept = tk.Entry(self.Frame1_20)
                        self.txt_reciept.place(relx=0.492, rely=0.116, height=20, relwidth=0.439)

                        self.txt_reciept.configure(background="white")
                        self.txt_reciept.configure(disabledforeground="#a3a3a3")
                        self.txt_reciept.configure(font="TkFixedFont")
                        self.txt_reciept.configure(foreground="#000000")
                        self.txt_reciept.configure(highlightbackground="#d9d9d9")
                        self.txt_reciept.configure(highlightcolor="black")
                        self.txt_reciept.configure(insertbackground="black")
                        self.txt_reciept.configure(selectbackground="#c4c4c4")
                        self.txt_reciept.configure(selectforeground="black")

                        self.Label7_2 = tk.Label(self.Frame1_20)
                        self.Label7_2.place(relx=0.033, rely=0.203, height=39, width=126)
                        self.Label7_2.configure(activebackground="#f9f9f9")
                        self.Label7_2.configure(activeforeground="black")
                        self.Label7_2.configure(background="#50e82a")
                        self.Label7_2.configure(disabledforeground="#a3a3a3")
                        self.Label7_2.configure(font="-family {Plantagenet Cherokee} -size 13")
                        self.Label7_2.configure(foreground="#000000")
                        self.Label7_2.configure(highlightbackground="#d9d9d9")
                        self.Label7_2.configure(highlightcolor="black")
                        self.Label7_2.configure(text='''Housetax(Paid) :''')

                        self.Label7_3 = tk.Label(self.Frame1_20)
                        self.Label7_3.place(relx=0.033, rely=0.319, height=39, width=126)
                        self.Label7_3.configure(activebackground="#f9f9f9")
                        self.Label7_3.configure(activeforeground="black")
                        self.Label7_3.configure(background="#50e82a")
                        self.Label7_3.configure(disabledforeground="#a3a3a3")
                        self.Label7_3.configure(font="-family {Plantagenet Cherokee} -size 13")
                        self.Label7_3.configure(foreground="#000000")
                        self.Label7_3.configure(highlightbackground="#d9d9d9")
                        self.Label7_3.configure(highlightcolor="black")
                        self.Label7_3.configure(text='''Healthtax (Paid) :''')

                        self.Label7_4 = tk.Label(self.Frame1_20)
                        self.Label7_4.place(relx=0.066, rely=0.435, height=39, width=116)
                        self.Label7_4.configure(activebackground="#f9f9f9")
                        self.Label7_4.configure(activeforeground="black")
                        self.Label7_4.configure(background="#50e82a")
                        self.Label7_4.configure(disabledforeground="#a3a3a3")
                        self.Label7_4.configure(font="-family {Plantagenet Cherokee} -size 13")
                        self.Label7_4.configure(foreground="#000000")
                        self.Label7_4.configure(highlightbackground="#d9d9d9")
                        self.Label7_4.configure(highlightcolor="black")
                        self.Label7_4.configure(text='''Lighttax(Paid) :''')

                        self.Label7_5 = tk.Label(self.Frame1_20)
                        self.Label7_5.place(relx=0.033, rely=0.551, height=39, width=116)
                        self.Label7_5.configure(activebackground="#f9f9f9")
                        self.Label7_5.configure(activeforeground="black")
                        self.Label7_5.configure(background="#50e82a")
                        self.Label7_5.configure(disabledforeground="#a3a3a3")
                        self.Label7_5.configure(font="-family {Plantagenet Cherokee} -size 13")
                        self.Label7_5.configure(foreground="#000000")
                        self.Label7_5.configure(highlightbackground="#d9d9d9")
                        self.Label7_5.configure(highlightcolor="black")
                        self.Label7_5.configure(text='''Watertax(Paid) :''')

                        self.Label7_6 = tk.Label(self.Frame1_20)
                        self.Label7_6.place(relx=0.033, rely=0.667, height=39, width=106)
                        self.Label7_6.configure(activebackground="#f9f9f9")
                        self.Label7_6.configure(activeforeground="black")
                        self.Label7_6.configure(background="#50e82a")
                        self.Label7_6.configure(disabledforeground="#a3a3a3")
                        self.Label7_6.configure(font="-family {Plantagenet Cherokee} -size 13")
                        self.Label7_6.configure(foreground="#000000")
                        self.Label7_6.configure(highlightbackground="#d9d9d9")
                        self.Label7_6.configure(highlightcolor="black")
                        self.Label7_6.configure(text='''Total (Paid) :''')

                        self.txt_housepaid = tk.Entry(self.Frame1_20)
                        self.txt_housepaid.place(relx=0.492, rely=0.232, height=20
                                , relwidth=0.439)
                        self.txt_housepaid.configure(background="white")
                        self.txt_housepaid.configure(disabledforeground="#a3a3a3")
                        self.txt_housepaid.configure(font="TkFixedFont")
                        self.txt_housepaid.configure(foreground="#000000")
                        self.txt_housepaid.configure(highlightbackground="#d9d9d9")
                        self.txt_housepaid.configure(highlightcolor="black")
                        self.txt_housepaid.configure(insertbackground="black")
                        self.txt_housepaid.configure(selectbackground="#c4c4c4")
                        self.txt_housepaid.configure(selectforeground="black")

                        self.txt_healthpaid = tk.Entry(self.Frame1_20)
                        self.txt_healthpaid.place(relx=0.492, rely=0.348, height=20
                                , relwidth=0.439)
                        self.txt_healthpaid.configure(background="white")
                        self.txt_healthpaid.configure(disabledforeground="#a3a3a3")
                        self.txt_healthpaid.configure(font="TkFixedFont")
                        self.txt_healthpaid.configure(foreground="#000000")
                        self.txt_healthpaid.configure(highlightbackground="#d9d9d9")
                        self.txt_healthpaid.configure(highlightcolor="black")
                        self.txt_healthpaid.configure(insertbackground="black")
                        self.txt_healthpaid.configure(selectbackground="#c4c4c4")
                        self.txt_healthpaid.configure(selectforeground="black")

                        self.txt_lightpaid = tk.Entry(self.Frame1_20)
                        self.txt_lightpaid.place(relx=0.492, rely=0.464, height=20
                                , relwidth=0.439)
                        self.txt_lightpaid.configure(background="white")
                        self.txt_lightpaid.configure(disabledforeground="#a3a3a3")
                        self.txt_lightpaid.configure(font="TkFixedFont")
                        self.txt_lightpaid.configure(foreground="#000000")
                        self.txt_lightpaid.configure(highlightbackground="#d9d9d9")
                        self.txt_lightpaid.configure(highlightcolor="black")
                        self.txt_lightpaid.configure(insertbackground="black")
                        self.txt_lightpaid.configure(selectbackground="#c4c4c4")
                        self.txt_lightpaid.configure(selectforeground="black")

                        self.txt_waterpaid = tk.Entry(self.Frame1_20)
                        self.txt_waterpaid.place(relx=0.492, rely=0.58, height=20
                                , relwidth=0.439)
                        self.txt_waterpaid.configure(background="white")
                        self.txt_waterpaid.configure(disabledforeground="#a3a3a3")
                        self.txt_waterpaid.configure(font="TkFixedFont")
                        self.txt_waterpaid.configure(foreground="#000000")
                        self.txt_waterpaid.configure(highlightbackground="#d9d9d9")
                        self.txt_waterpaid.configure(highlightcolor="black")
                        self.txt_waterpaid.configure(insertbackground="black")
                        self.txt_waterpaid.configure(selectbackground="#c4c4c4")
                        self.txt_waterpaid.configure(selectforeground="black")

                        self.txt_totalpaid = tk.Entry(self.Frame1_20)
                        self.txt_totalpaid.place(relx=0.492, rely=0.696, height=20
                                , relwidth=0.439)
                        self.txt_totalpaid.configure(background="white")
                        self.txt_totalpaid.configure(disabledforeground="#a3a3a3")
                        self.txt_totalpaid.configure(font="TkFixedFont")
                        self.txt_totalpaid.configure(foreground="#000000")
                        self.txt_totalpaid.configure(highlightbackground="#d9d9d9")
                        self.txt_totalpaid.configure(highlightcolor="black")
                        self.txt_totalpaid.configure(insertbackground="black")
                        self.txt_totalpaid.configure(selectbackground="#c4c4c4")
                        self.txt_totalpaid.configure(selectforeground="black")

                        self.btn_submit = tk.Button(self.Frame1_20)
                        self.btn_submit.place(relx=0.328, rely=0.812, height=44, width=97)
                        self.btn_submit.configure(activebackground="#ececec")
                        self.btn_submit.configure(activeforeground="#000000")
                        self.btn_submit.configure(background="#ff350d")
                        self.btn_submit.configure(disabledforeground="#a3a3a3")
                        self.btn_submit.configure(font="-family {Segoe UI} -size 13")
                        self.btn_submit.configure(foreground="#ffffff")
                        self.btn_submit.configure(highlightbackground="#d9d9d9")
                        self.btn_submit.configure(highlightcolor="black")
                        self.btn_submit.configure(pady="0")
                        self.btn_submit.configure(text='''SUBMIT''',command = self.submit4)

                        self.Frame1_21 = tk.Frame(top)
                        self.Frame1_21.place(relx=0.537, rely=0.274, relheight=0.089
                                , relwidth=0.292)
                        self.Frame1_21.configure(relief='ridge')
                        self.Frame1_21.configure(borderwidth="10")
                        self.Frame1_21.configure(relief='ridge')
                        self.Frame1_21.configure(background="#50e82a")
                        self.Frame1_21.configure(highlightbackground="#d9d9d9")
                        self.Frame1_21.configure(highlightcolor="black")
                        self.Frame1_21.configure(width=305)

                        self.Label7_22 = tk.Label(self.Frame1_21)
                        self.Label7_22.place(relx=0.033, rely=0.154, height=39, width=126)
                        self.Label7_22.configure(activebackground="#f9f9f9")
                        self.Label7_22.configure(activeforeground="black")
                        self.Label7_22.configure(background="#50e82a")
                        self.Label7_22.configure(disabledforeground="#a3a3a3")
                        self.Label7_22.configure(font="-family {Plantagenet Cherokee} -size 13")
                        self.Label7_22.configure(foreground="#000000")
                        self.Label7_22.configure(highlightbackground="#d9d9d9")
                        self.Label7_22.configure(highlightcolor="black")
                        self.Label7_22.configure(text='''Village Name :''')

                        self.txt_village = tk.Entry(self.Frame1_21)
                        self.txt_village.place(relx=0.492, rely=0.308, height=20, relwidth=0.439)

                        self.txt_village.configure(background="white")
                        self.txt_village.configure(disabledforeground="#a3a3a3")
                        self.txt_village.configure(font="TkFixedFont")
                        self.txt_village.configure(foreground="#000000")
                        self.txt_village.configure(highlightbackground="#d9d9d9")
                        self.txt_village.configure(highlightcolor="black")
                        self.txt_village.configure(insertbackground="black")
                        self.txt_village.configure(selectbackground="#c4c4c4")
                        self.txt_village.configure(selectforeground="black")

                if __name__ == '__main__':
                    vp_start_gui()























































            def delete1():

                import sys
                import tkinter
                from tkinter import messagebox
                import mysql.connector
                import time
                import dbConnect
                from dbConnect import DBConnect
                import smtplib

                try:
                    import Tkinter as tk
                except ImportError:
                    import tkinter as tk

                try:
                    import ttk
                    py3 = False
                except ImportError:
                    import tkinter.ttk as ttk
                    py3 = True

                import deletedata_support
                import os.path

                def vp_start_gui():
                    '''Starting point when module is the main routine.'''
                    global val, w, root
                    global prog_location
                    prog_call = sys.argv[0]
                    print ('prog_call = {}'.format(prog_call))
                    prog_location = os.path.split(prog_call)[0]
                    print ('prog_location = {}'.format(prog_location))
                    sys.stdout.flush()
                    root = tk.Tk()
                    top = e_TAX_2019 (root)
                    deletedata_support.init(root, top)
                    root.mainloop()

                w = None
                def create_e_TAX_2019(root, *args, **kwargs):
                    '''Starting point when module is imported by another program.'''
                    global w, w_win, rt
                    global prog_location
                    prog_call = sys.argv[0]
                    print ('prog_call = {}'.format(prog_call))
                    prog_location = os.path.split(prog_call)[0]
                    print ('prog_location = {}'.format(prog_location))
                    rt = root
                    w = tk.Toplevel (root)
                    top = e_TAX_2019 (w)
                    deletedata_support.init(w, top, *args, **kwargs)
                    return (w, top)

                def destroy_e_TAX_2019():
                    global w
                    w.destroy()
                    w = None

                class e_TAX_2019:
                    def exits(self):
                        msg=tkinter.messagebox.askyesno("e-TAX 2019","Do You Want To EXIT ?")
                        if msg :
                            os._exit(1)


                    def backs(self):
                        root.destroy()
                        update1()


                    def submit2(self):
                        v=str(self.txt_villagename.get());
                        n=str(self.txt_name.get());
                        u=str(self.txt_uidnumber.get());

                        t=str(time.asctime(time.localtime(time.time())))
                        try:
                            mydb=mysql.connector.connect(host='remotemysql.com',user='NhpDbfVMT3',password='NXWSVJNu55',database='NhpDbfVMT3',port='3306')
                        except:
                            tkinter.messagebox.showerror("e-TAX 2019","Error While Connecting to the database")
                        mycursor=mydb.cursor()
                        query=("INSERT INTO deleteddata (village, uidnumber, name)VALUES(%s,%s,%s)")
                        datac=(v,u,n)
                        mycursor.execute(query,datac)
                        mydb.commit()

                        totcontaint=str("Warning !!!!! \n\n New Request To Delete Entry \n\n Some Entries has been changed \n\n Here are the details : \n\n\n Village Name :"+v+"\n UID Number :"+u+"\n Name Of individual"+n+"\n Data deleted on :"+t)
                        mail= smtplib.SMTP('smtp.gmail.com',587)
                        mail.ehlo()
                        mail.starttls()
                        mail.login('etaxsupp2019@gmail.com','Pass@123')
                        mail.sendmail('etaxsupp2019@gmail.com','kulkarnipranesh1767@gmail.com',totcontaint)
                        mail.close()
                        messagebox.showinfo("e-TAX 2019","Successfully deleted The Entry")
                        root.destroy()
                        extra1()


                    def __init__(self, top=None):
                        '''This class configures and populates the toplevel window.
                           top is the toplevel containing window.'''
                        _bgcolor = '#d9d9d9'  # X11 color: 'gray85'
                        _fgcolor = '#000000'  # X11 color: 'black'
                        _compcolor = '#d9d9d9' # X11 color: 'gray85'
                        _ana1color = '#d9d9d9' # X11 color: 'gray85'
                        _ana2color = '#ececec' # Closest X11 color: 'gray92'
                        font12 = "-family {Swatch it} -size 16"
                        font13 = "-family {Segoe UI Black} -size 13 -weight bold"
                        font14 = "-family {Shonar Bangla} -size 22"

                        top.geometry("843x469+274+133")
                        top.title("e-TAX 2019")
                        top.configure(background="#727272")
                        top.configure(highlightbackground="#d9d9d9")
                        top.configure(highlightcolor="black")

                        self.Frame1 = tk.Frame(top)
                        self.Frame1.place(relx=0.012, rely=0.021, relheight=0.309
                                , relwidth=0.836)
                        self.Frame1.configure(relief='ridge')
                        self.Frame1.configure(borderwidth="10")
                        self.Frame1.configure(relief='ridge')
                        self.Frame1.configure(background="#595959")
                        self.Frame1.configure(highlightbackground="#d9d9d9")
                        self.Frame1.configure(highlightcolor="black")
                        self.Frame1.configure(width=705)

                        self.Label1 = tk.Label(self.Frame1)
                        self.Label1.place(relx=0.043, rely=0.207, height=59, width=196)
                        self.Label1.configure(activebackground="#f9f9f9")
                        self.Label1.configure(activeforeground="black")
                        self.Label1.configure(background="#595959")
                        self.Label1.configure(disabledforeground="#a3a3a3")
                        self.Label1.configure(font="-family {Rockwell Extra} -size 40 -weight bold")
                        self.Label1.configure(foreground="#0d4dff")
                        self.Label1.configure(highlightbackground="#d9d9d9")
                        self.Label1.configure(highlightcolor="black")
                        self.Label1.configure(text='''eTAX''')

                        self.Label1_1 = tk.Label(self.Frame1)
                        self.Label1_1.place(relx=0.298, rely=0.207, height=59, width=146)
                        self.Label1_1.configure(activebackground="#f9f9f9")
                        self.Label1_1.configure(activeforeground="black")
                        self.Label1_1.configure(background="#595959")
                        self.Label1_1.configure(disabledforeground="#a3a3a3")
                        self.Label1_1.configure(font="-family {Rockwell Extra} -size 40 -weight bold")
                        self.Label1_1.configure(foreground="#ff2b0a")
                        self.Label1_1.configure(highlightbackground="#d9d9d9")
                        self.Label1_1.configure(highlightcolor="black")
                        self.Label1_1.configure(text='''2019''')

                        self.Label2 = tk.Label(self.Frame1)
                        self.Label2.place(relx=0.156, rely=0.621, height=31, width=184)
                        self.Label2.configure(activebackground="#f9f9f9")
                        self.Label2.configure(activeforeground="black")
                        self.Label2.configure(background="#595959")
                        self.Label2.configure(disabledforeground="#a3a3a3")
                        self.Label2.configure(font="-family {Rage Italic} -size 19 -slant italic")
                        self.Label2.configure(foreground="#f7ff0a")
                        self.Label2.configure(highlightbackground="#d9d9d9")
                        self.Label2.configure(highlightcolor="black")
                        self.Label2.configure(text='''working for you''')

                        self.Label1_2 = tk.Label(self.Frame1)
                        self.Label1_2.place(relx=0.525, rely=0.207, height=69, width=306)
                        self.Label1_2.configure(activebackground="#f9f9f9")
                        self.Label1_2.configure(activeforeground="black")
                        self.Label1_2.configure(background="#595959")
                        self.Label1_2.configure(disabledforeground="#a3a3a3")
                        self.Label1_2.configure(font="-family {Rockwell Extra} -size 28 -weight bold")
                        self.Label1_2.configure(foreground="#5faa14")
                        self.Label1_2.configure(highlightbackground="#d9d9d9")
                        self.Label1_2.configure(highlightcolor="black")
                        self.Label1_2.configure(text='''delete DATA''')
                        self.Label1_2.configure(width=306)

                        self.menubar = tk.Menu(top,font="TkMenuFont",bg=_bgcolor,fg=_fgcolor)
                        top.configure(menu = self.menubar)

                        self.Label3 = tk.Label(top)
                        self.Label3.place(relx=0.107, rely=0.448, height=39, width=186)
                        self.Label3.configure(background="#727272")
                        self.Label3.configure(disabledforeground="#a3a3a3")
                        self.Label3.configure(font=font12)
                        self.Label3.configure(foreground="#f7ff1c")
                        self.Label3.configure(text='''Village Name :''')
                        self.Label3.configure(width=186)

                        self.Label3_1 = tk.Label(top)
                        self.Label3_1.place(relx=0.107, rely=0.554, height=29, width=186)
                        self.Label3_1.configure(activebackground="#f9f9f9")
                        self.Label3_1.configure(activeforeground="black")
                        self.Label3_1.configure(background="#727272")
                        self.Label3_1.configure(disabledforeground="#a3a3a3")
                        self.Label3_1.configure(font="-family {Swatch it} -size 16")
                        self.Label3_1.configure(foreground="#f7ff1c")
                        self.Label3_1.configure(highlightbackground="#d9d9d9")
                        self.Label3_1.configure(highlightcolor="black")
                        self.Label3_1.configure(text='''UID Number :''')

                        self.Label3_2 = tk.Label(top)
                        self.Label3_2.place(relx=0.142, rely=0.64, height=39, width=116)
                        self.Label3_2.configure(activebackground="#f9f9f9")
                        self.Label3_2.configure(activeforeground="black")
                        self.Label3_2.configure(background="#727272")
                        self.Label3_2.configure(disabledforeground="#a3a3a3")
                        self.Label3_2.configure(font="-family {Swatch it} -size 16")
                        self.Label3_2.configure(foreground="#f7ff1c")
                        self.Label3_2.configure(highlightbackground="#d9d9d9")
                        self.Label3_2.configure(highlightcolor="black")
                        self.Label3_2.configure(text='''Name  :''')
                        self.Label3_2.configure(width=116)

                        self.txt_villagename = tk.Entry(top)
                        self.txt_villagename.place(relx=0.368, rely=0.469, height=20
                                , relwidth=0.361)
                        self.txt_villagename.configure(background="white")
                        self.txt_villagename.configure(disabledforeground="#a3a3a3")
                        self.txt_villagename.configure(font="TkFixedFont")
                        self.txt_villagename.configure(foreground="#000000")
                        self.txt_villagename.configure(insertbackground="black")
                        self.txt_villagename.configure(width=304)

                        self.txt_uidnumber = tk.Entry(top)
                        self.txt_uidnumber.place(relx=0.368, rely=0.576, height=20
                                , relwidth=0.361)
                        self.txt_uidnumber.configure(background="white")
                        self.txt_uidnumber.configure(disabledforeground="#a3a3a3")
                        self.txt_uidnumber.configure(font="TkFixedFont")
                        self.txt_uidnumber.configure(foreground="#000000")
                        self.txt_uidnumber.configure(highlightbackground="#d9d9d9")
                        self.txt_uidnumber.configure(highlightcolor="black")
                        self.txt_uidnumber.configure(insertbackground="black")
                        self.txt_uidnumber.configure(selectbackground="#c4c4c4")
                        self.txt_uidnumber.configure(selectforeground="black")

                        self.txt_name = tk.Entry(top)
                        self.txt_name.place(relx=0.368, rely=0.661,height=20, relwidth=0.361)
                        self.txt_name.configure(background="white")
                        self.txt_name.configure(disabledforeground="#a3a3a3")
                        self.txt_name.configure(font="TkFixedFont")
                        self.txt_name.configure(foreground="#000000")
                        self.txt_name.configure(highlightbackground="#d9d9d9")
                        self.txt_name.configure(highlightcolor="black")
                        self.txt_name.configure(insertbackground="black")
                        self.txt_name.configure(selectbackground="#c4c4c4")
                        self.txt_name.configure(selectforeground="black")

                        self.Label4 = tk.Label(top)
                        self.Label4.place(relx=0.771, rely=0.49, height=100, width=174)
                        self.Label4.configure(background="#d9d9d9")
                        self.Label4.configure(borderwidth="10")
                        self.Label4.configure(disabledforeground="#a3a3a3")
                        self.Label4.configure(foreground="#000000")
                        photo_location = os.path.join(prog_location,"./delete1.png")
                        self._img0 = tk.PhotoImage(file=photo_location)
                        self.Label4.configure(image=self._img0)
                        self.Label4.configure(relief='raised')
                        self.Label4.configure(text='''Label''')
                        self.Label4.configure(width=174)

                        self.btn_submit = tk.Button(top)
                        self.btn_submit.place(relx=0.427, rely=0.832, height=54, width=194)
                        self.btn_submit.configure(activebackground="#ececec")
                        self.btn_submit.configure(activeforeground="#000000")
                        self.btn_submit.configure(background="#3920d8")
                        self.btn_submit.configure(borderwidth="10")
                        self.btn_submit.configure(disabledforeground="#a3a3a3")
                        self.btn_submit.configure(font=font13)
                        self.btn_submit.configure(foreground="#ffffff")
                        self.btn_submit.configure(highlightbackground="#d9d9d9")
                        self.btn_submit.configure(highlightcolor="black")
                        self.btn_submit.configure(pady="0")
                        self.btn_submit.configure(text='''I Wish To Continue''',command = self.submit2)

                        self.Label5 = tk.Label(top)
                        self.Label5.place(relx=0.095, rely=0.725, height=43, width=560)
                        self.Label5.configure(background="#727272")
                        self.Label5.configure(disabledforeground="#a3a3a3")
                        self.Label5.configure(font=font14)
                        self.Label5.configure(foreground="#5bff3b")
                        self.Label5.configure(text='''Disclaimer :  You Can Not recover the data once deleted''')

                        self.btn_exit = tk.Button(top)
                        self.btn_exit.place(relx=0.866, rely=0.192, height=44, width=104)
                        self.btn_exit.configure(activebackground="#ececec")
                        self.btn_exit.configure(activeforeground="#000000")
                        self.btn_exit.configure(background="#ff3f0f")
                        self.btn_exit.configure(borderwidth="10")
                        self.btn_exit.configure(disabledforeground="#a3a3a3")
                        self.btn_exit.configure(font="-family {Segoe UI Black} -size 13 -weight bold")
                        self.btn_exit.configure(foreground="#ffffff")
                        self.btn_exit.configure(highlightbackground="#d9d9d9")
                        self.btn_exit.configure(highlightcolor="black")
                        self.btn_exit.configure(pady="0")
                        self.btn_exit.configure(text='''EXIT''',command = self.exits)
                        self.btn_exit.configure(width=104)

                        self.btn_back = tk.Button(top)
                        self.btn_back.place(relx=0.866, rely=0.064, height=44, width=104)
                        self.btn_back.configure(activebackground="#ececec")
                        self.btn_back.configure(activeforeground="#000000")
                        self.btn_back.configure(background="#3ba825")
                        self.btn_back.configure(borderwidth="10")
                        self.btn_back.configure(disabledforeground="#a3a3a3")
                        self.btn_back.configure(font="-family {Segoe UI Black} -size 13 -weight bold")
                        self.btn_back.configure(foreground="#ffffff")
                        self.btn_back.configure(highlightbackground="#d9d9d9")
                        self.btn_back.configure(highlightcolor="black")
                        self.btn_back.configure(pady="0")
                        self.btn_back.configure(text='''BACK''',command = self.backs)

                if __name__ == '__main__':
                    vp_start_gui()



















































            class e_TAX_2019:


                def backs(self):
                    root.destroy()
                    extra1()

                def exits(self):
                    msg=tkinter.messagebox.askyesno("eTAX 2019","Do you want to exit")
                    if msg :
                        os._exit(1)

                def accesss(self):
                    root.destroy()
                    access1()


                def deletes(self):
                    root.destroy()
                    delete1()

                def modifys(self):
                    root.destroy()
                    modify1()

                def __init__(self, top=None):
                    '''This class configures and populates the toplevel window.
                       top is the toplevel containing window.'''
                    _bgcolor = '#d9d9d9'  # X11 color: 'gray85'
                    _fgcolor = '#000000'  # X11 color: 'black'
                    _compcolor = '#d9d9d9' # X11 color: 'gray85'
                    _ana1color = '#d9d9d9' # X11 color: 'gray85'
                    _ana2color = '#ececec' # Closest X11 color: 'gray92'
                    font10 = "-family {Rockwell Extra} -size 40 -weight bold"
                    font11 = "-family {Rage Italic} -size 19 -slant italic"
                    font12 = "-family {Rockwell Extra} -size 28 -weight bold"
                    font15 = "-family {Segoe UI Black} -size 23 -weight bold"
                    font16 = "-family {Postmaster} -size 13"

                    top.geometry("990x650+274+133")
                    top.title("e-TAX 2019")
                    top.configure(background="#727272")

                    self.Frame1 = tk.Frame(top)
                    self.Frame1.place(relx=0.091, rely=0.046, relheight=0.223
                            , relwidth=0.843)
                    self.Frame1.configure(relief='ridge')
                    self.Frame1.configure(borderwidth="10")
                    self.Frame1.configure(relief='ridge')
                    self.Frame1.configure(background="#595959")
                    self.Frame1.configure(width=835)

                    self.Label1 = tk.Label(self.Frame1)
                    self.Label1.place(relx=0.036, rely=0.207, height=59, width=196)
                    self.Label1.configure(background="#595959")
                    self.Label1.configure(disabledforeground="#a3a3a3")
                    self.Label1.configure(font=font10)
                    self.Label1.configure(foreground="#0d4dff")
                    self.Label1.configure(text='''eTAX''')
                    self.Label1.configure(width=196)

                    self.Label1_1 = tk.Label(self.Frame1)
                    self.Label1_1.place(relx=0.251, rely=0.207, height=59, width=146)
                    self.Label1_1.configure(activebackground="#f9f9f9")
                    self.Label1_1.configure(activeforeground="black")
                    self.Label1_1.configure(background="#595959")
                    self.Label1_1.configure(disabledforeground="#a3a3a3")
                    self.Label1_1.configure(font="-family {Rockwell Extra} -size 40 -weight bold")
                    self.Label1_1.configure(foreground="#ff2b0a")
                    self.Label1_1.configure(highlightbackground="#d9d9d9")
                    self.Label1_1.configure(highlightcolor="black")
                    self.Label1_1.configure(text='''2019''')
                    self.Label1_1.configure(width=146)

                    self.Label2 = tk.Label(self.Frame1)
                    self.Label2.place(relx=0.132, rely=0.621, height=31, width=184)
                    self.Label2.configure(background="#595959")
                    self.Label2.configure(disabledforeground="#a3a3a3")
                    self.Label2.configure(font=font11)
                    self.Label2.configure(foreground="#f7ff0a")
                    self.Label2.configure(text='''working for you''')
                    self.Label2.configure(width=184)

                    self.Label1_2 = tk.Label(self.Frame1)
                    self.Label1_2.place(relx=0.431, rely=0.207, height=69, width=446)
                    self.Label1_2.configure(activebackground="#f9f9f9")
                    self.Label1_2.configure(activeforeground="black")
                    self.Label1_2.configure(background="#595959")
                    self.Label1_2.configure(disabledforeground="#a3a3a3")
                    self.Label1_2.configure(font=font12)
                    self.Label1_2.configure(foreground="#5faa14")
                    self.Label1_2.configure(highlightbackground="#d9d9d9")
                    self.Label1_2.configure(highlightcolor="black")
                    self.Label1_2.configure(text='''UPDATE DATABASE''')
                    self.Label1_2.configure(width=446)

                    self.Frame2 = tk.Frame(top)
                    self.Frame2.place(relx=0.01, rely=0.369, relheight=0.531, relwidth=0.328)

                    self.Frame2.configure(relief='ridge')
                    self.Frame2.configure(borderwidth="10")
                    self.Frame2.configure(relief='ridge')
                    self.Frame2.configure(background="#595959")
                    self.Frame2.configure(width=325)

                    self.Label3 = tk.Label(self.Frame2)
                    self.Label3.place(relx=0.123, rely=0.638, height=45, width=234)
                    self.Label3.configure(background="#595959")
                    self.Label3.configure(disabledforeground="#a3a3a3")
                    self.Label3.configure(font=font15)
                    self.Label3.configure(foreground="#efff0f")
                    self.Label3.configure(text='''MODIFY DATA''')
                    self.Label3.configure(width=234)

                    self.Label4 = tk.Label(self.Frame2)
                    self.Label4.place(relx=0.154, rely=0.783, height=22, width=213)
                    self.Label4.configure(background="#595959")
                    self.Label4.configure(disabledforeground="#a3a3a3")
                    self.Label4.configure(font=font16)
                    self.Label4.configure(foreground="#4cad1c")
                    self.Label4.configure(text='''CLICK HERE TO UPDATE''')

                    self.Label4_3 = tk.Label(self.Frame2)
                    self.Label4_3.place(relx=0.338, rely=0.841, height=32, width=83)
                    self.Label4_3.configure(activebackground="#f9f9f9")
                    self.Label4_3.configure(activeforeground="black")
                    self.Label4_3.configure(background="#595959")
                    self.Label4_3.configure(disabledforeground="#a3a3a3")
                    self.Label4_3.configure(font="-family {Postmaster} -size 13")
                    self.Label4_3.configure(foreground="#4cad1c")
                    self.Label4_3.configure(highlightbackground="#d9d9d9")
                    self.Label4_3.configure(highlightcolor="black")
                    self.Label4_3.configure(text='''DATA''')
                    self.Label4_3.configure(width=83)

                    self.btn_modify = tk.Button(self.Frame2)
                    self.btn_modify.place(relx=0.215, rely=0.058, height=186, width=186)
                    self.btn_modify.configure(activebackground="#ececec")
                    self.btn_modify.configure(activeforeground="#000000")
                    self.btn_modify.configure(background="#d9d9d9")
                    self.btn_modify.configure(disabledforeground="#a3a3a3")
                    self.btn_modify.configure(foreground="#000000")
                    self.btn_modify.configure(highlightbackground="#d9d9d9")
                    self.btn_modify.configure(highlightcolor="black")
                    photo_location = os.path.join(prog_location,"./updatepic1.png")
                    self._img0 = tk.PhotoImage(file=photo_location)
                    self.btn_modify.configure(image=self._img0)
                    self.btn_modify.configure(pady="0")
                    self.btn_modify.configure(text='''Button''',command= self.modifys)
                    self.btn_modify.configure(width=186)

                    self.Frame2_4 = tk.Frame(top)
                    self.Frame2_4.place(relx=0.333, rely=0.369, relheight=0.531
                            , relwidth=0.328)
                    self.Frame2_4.configure(relief='ridge')
                    self.Frame2_4.configure(borderwidth="10")
                    self.Frame2_4.configure(relief='ridge')
                    self.Frame2_4.configure(background="#595959")
                    self.Frame2_4.configure(highlightbackground="#d9d9d9")
                    self.Frame2_4.configure(highlightcolor="black")
                    self.Frame2_4.configure(width=325)

                    self.Label3_5 = tk.Label(self.Frame2_4)
                    self.Label3_5.place(relx=0.123, rely=0.638, height=45, width=234)
                    self.Label3_5.configure(activebackground="#f9f9f9")
                    self.Label3_5.configure(activeforeground="black")
                    self.Label3_5.configure(background="#595959")
                    self.Label3_5.configure(disabledforeground="#a3a3a3")
                    self.Label3_5.configure(font="-family {Segoe UI Black} -size 23 -weight bold")
                    self.Label3_5.configure(foreground="#efff0f")
                    self.Label3_5.configure(highlightbackground="#d9d9d9")
                    self.Label3_5.configure(highlightcolor="black")
                    self.Label3_5.configure(text='''DELETE DATA''')

                    self.Label4_6 = tk.Label(self.Frame2_4)
                    self.Label4_6.place(relx=0.154, rely=0.783, height=22, width=209)
                    self.Label4_6.configure(activebackground="#f9f9f9")
                    self.Label4_6.configure(activeforeground="black")
                    self.Label4_6.configure(background="#595959")
                    self.Label4_6.configure(disabledforeground="#a3a3a3")
                    self.Label4_6.configure(font="-family {Postmaster} -size 13")
                    self.Label4_6.configure(foreground="#4cad1c")
                    self.Label4_6.configure(highlightbackground="#d9d9d9")
                    self.Label4_6.configure(highlightcolor="black")
                    self.Label4_6.configure(text='''CLICK HERE TO DELETE''')

                    self.Label4_4 = tk.Label(self.Frame2_4)
                    self.Label4_4.place(relx=0.338, rely=0.841, height=32, width=83)
                    self.Label4_4.configure(activebackground="#f9f9f9")
                    self.Label4_4.configure(activeforeground="black")
                    self.Label4_4.configure(background="#595959")
                    self.Label4_4.configure(disabledforeground="#a3a3a3")
                    self.Label4_4.configure(font="-family {Postmaster} -size 13")
                    self.Label4_4.configure(foreground="#4cad1c")
                    self.Label4_4.configure(highlightbackground="#d9d9d9")
                    self.Label4_4.configure(highlightcolor="black")
                    self.Label4_4.configure(text='''DATA''')

                    self.btn_delete = tk.Button(self.Frame2_4)
                    self.btn_delete.place(relx=0.185, rely=0.174, height=122, width=206)
                    self.btn_delete.configure(activebackground="#ececec")
                    self.btn_delete.configure(activeforeground="#000000")
                    self.btn_delete.configure(background="#d9d9d9")
                    self.btn_delete.configure(disabledforeground="#a3a3a3")
                    self.btn_delete.configure(foreground="#000000")
                    self.btn_delete.configure(highlightbackground="#d9d9d9")
                    self.btn_delete.configure(highlightcolor="black")
                    photo_location = os.path.join(prog_location,"./delete1.png")
                    self._img1 = tk.PhotoImage(file=photo_location)
                    self.btn_delete.configure(image=self._img1)
                    self.btn_delete.configure(pady="0")
                    self.btn_delete.configure(text='''Button''',command= self.deletes)

                    self.Frame2_5 = tk.Frame(top)
                    self.Frame2_5.place(relx=0.657, rely=0.369, relheight=0.531
                            , relwidth=0.328)
                    self.Frame2_5.configure(relief='ridge')
                    self.Frame2_5.configure(borderwidth="10")
                    self.Frame2_5.configure(relief='ridge')
                    self.Frame2_5.configure(background="#595959")
                    self.Frame2_5.configure(highlightbackground="#d9d9d9")
                    self.Frame2_5.configure(highlightcolor="black")
                    self.Frame2_5.configure(width=325)

                    self.Label3_6 = tk.Label(self.Frame2_5)
                    self.Label3_6.place(relx=0.123, rely=0.638, height=45, width=234)
                    self.Label3_6.configure(activebackground="#f9f9f9")
                    self.Label3_6.configure(activeforeground="black")
                    self.Label3_6.configure(background="#595959")
                    self.Label3_6.configure(disabledforeground="#a3a3a3")
                    self.Label3_6.configure(font="-family {Segoe UI Black} -size 23 -weight bold")
                    self.Label3_6.configure(foreground="#efff0f")
                    self.Label3_6.configure(highlightbackground="#d9d9d9")
                    self.Label3_6.configure(highlightcolor="black")
                    self.Label3_6.configure(text='''PERMISSION''')

                    self.Label4_7 = tk.Label(self.Frame2_5)
                    self.Label4_7.place(relx=0.154, rely=0.783, height=22, width=219)
                    self.Label4_7.configure(activebackground="#f9f9f9")
                    self.Label4_7.configure(activeforeground="black")
                    self.Label4_7.configure(background="#595959")
                    self.Label4_7.configure(disabledforeground="#a3a3a3")
                    self.Label4_7.configure(font="-family {Postmaster} -size 13")
                    self.Label4_7.configure(foreground="#4cad1c")
                    self.Label4_7.configure(highlightbackground="#d9d9d9")
                    self.Label4_7.configure(highlightcolor="black")
                    self.Label4_7.configure(text='''CLICK HERE TO ASK FOR''')

                    self.Label4_4 = tk.Label(self.Frame2_5)
                    self.Label4_4.place(relx=0.338, rely=0.841, height=32, width=103)
                    self.Label4_4.configure(activebackground="#f9f9f9")
                    self.Label4_4.configure(activeforeground="black")
                    self.Label4_4.configure(background="#595959")
                    self.Label4_4.configure(disabledforeground="#a3a3a3")
                    self.Label4_4.configure(font="-family {Postmaster} -size 13")
                    self.Label4_4.configure(foreground="#4cad1c")
                    self.Label4_4.configure(highlightbackground="#d9d9d9")
                    self.Label4_4.configure(highlightcolor="black")
                    self.Label4_4.configure(text='''PERMISSION''')
                    self.Label4_4.configure(width=103)

                    self.btn_access = tk.Button(self.Frame2_5)
                    self.btn_access.place(relx=0.185, rely=0.174, height=139, width=206)
                    self.btn_access.configure(activebackground="#ececec")
                    self.btn_access.configure(activeforeground="#000000")
                    self.btn_access.configure(background="#d9d9d9")
                    self.btn_access.configure(disabledforeground="#a3a3a3")
                    self.btn_access.configure(foreground="#000000")
                    self.btn_access.configure(highlightbackground="#d9d9d9")
                    self.btn_access.configure(highlightcolor="black")
                    photo_location = os.path.join(prog_location,"./access.png")
                    self._img2 = tk.PhotoImage(file=photo_location)
                    self.btn_access.configure(image=self._img2)
                    self.btn_access.configure(pady="0")
                    self.btn_access.configure(text='''Button''',command= self.accesss)

                    self.btn_exit = tk.Button(top)
                    self.btn_exit.place(relx=0.869, rely=0.292, height=40, width=90)
                    self.btn_exit.configure(activebackground="#ececec")
                    self.btn_exit.configure(activeforeground="#000000")
                    self.btn_exit.configure(background="#264aff")
                    self.btn_exit.configure(borderwidth="10")
                    self.btn_exit.configure(disabledforeground="#a3a3a3")
                    self.btn_exit.configure(foreground="#ffffff")
                    self.btn_exit.configure(highlightbackground="#d9d9d9")
                    self.btn_exit.configure(highlightcolor="black")
                    self.btn_exit.configure(pady="0")
                    self.btn_exit.configure(text='''EXIT''')
                    self.btn_exit.configure(width=90,command= self.exits)

                    self.btn_back = tk.Button(top)
                    self.btn_back.place(relx=0.747, rely=0.292, height=40, width=90)
                    self.btn_back.configure(activebackground="#ececec")
                    self.btn_back.configure(activeforeground="#000000")
                    self.btn_back.configure(background="#264aff")
                    self.btn_back.configure(borderwidth="10")
                    self.btn_back.configure(disabledforeground="#a3a3a3")
                    self.btn_back.configure(foreground="#ffffff")
                    self.btn_back.configure(highlightbackground="#d9d9d9")
                    self.btn_back.configure(highlightcolor="black")
                    self.btn_back.configure(pady="0")
                    self.btn_back.configure(text='''BACK''',command= self.backs)

            if __name__ == '__main__':
                vp_start_gui()









































        class Toplevel1:
            def backs(self):
                root.destroy()
                main()


            def exits(self):
                msg=tkinter.messagebox.askyesno("Main","Do You Want To Exit ?")
                if msg:
                    os._exit(1)



            def worlds(self):
                root.destroy()
                world1()

            def calculators(self):
                root.destroy()
                calculator1()

            def onlines(self):
                root.destroy()
                online1()

            def updates(self):
                root.destroy()
                update1()


            def __init__(self, top=None):
                '''This class configures and populates the toplevel window.
                   top is the toplevel containing window.'''
                _bgcolor = '#d9d9d9'  # X11 color: 'gray85'
                _fgcolor = '#000000'  # X11 color: 'black'
                _compcolor = '#d9d9d9' # X11 color: 'gray85'
                _ana1color = '#d9d9d9' # X11 color: 'gray85'
                _ana2color = '#ececec' # Closest X11 color: 'gray92'
                font10 = "-family {Harlow Solid Italic} -size 40 -slant italic"  \
                    ""
                font12 = "-family {BadBlocks TT} -size 28 -weight bold"
                font13 = "-family {BadBlocks TT} -size 24 -weight bold"
                font14 = "-family {Serifa BT} -size 14 -weight bold"
                font15 = "-family {Sky Skunk} -size 12"
                font9 = "-family {BankGothic Md BT} -size 40 -weight bold"
                self.style = ttk.Style()
                if sys.platform == "win32":
                    self.style.theme_use('winnative')
                self.style.configure('.',background=_bgcolor)
                self.style.configure('.',foreground=_fgcolor)
                self.style.configure('.',font="TkDefaultFont")
                self.style.map('.',background=
                    [('selected', _compcolor), ('active',_ana2color)])

                top.geometry("992x628+286+152")
                top.title("eTAX- 2019")
                top.configure(background="#939393")

                self.menubar = tk.Menu(top,font="TkMenuFont",bg=_bgcolor,fg=_fgcolor)
                top.configure(menu = self.menubar)

                self.Frame1 = tk.Frame(top)
                self.Frame1.place(relx=0.02, rely=0.048, relheight=0.135, relwidth=0.64)
                self.Frame1.configure(relief='ridge')
                self.Frame1.configure(borderwidth="10")
                self.Frame1.configure(relief='ridge')
                self.Frame1.configure(background="#595959")
                self.Frame1.configure(width=635)

                self.Label1 = tk.Label(self.Frame1)
                self.Label1.place(relx=0.11, rely=0.235, height=41, width=197)
                self.Label1.configure(background="#595959")
                self.Label1.configure(disabledforeground="#a3a3a3")
                self.Label1.configure(font=font9)
                self.Label1.configure(foreground="#ff2a00")
                self.Label1.configure(text='''e-TAX''')
                self.Label1.configure(width=197)

                self.Label1_3 = tk.Label(self.Frame1)
                self.Label1_3.place(relx=0.409, rely=0.235, height=41, width=177)
                self.Label1_3.configure(activebackground="#f9f9f9")
                self.Label1_3.configure(activeforeground="black")
                self.Label1_3.configure(background="#595959")
                self.Label1_3.configure(disabledforeground="#a3a3a3")
                self.Label1_3.configure(font="-family {BankGothic Md BT} -size 40 -weight bold")
                self.Label1_3.configure(foreground="#143bff")
                self.Label1_3.configure(highlightbackground="#d9d9d9")
                self.Label1_3.configure(highlightcolor="black")
                self.Label1_3.configure(text='''2019''')
                self.Label1_3.configure(width=177)

                self.Label2 = tk.Label(self.Frame1)
                self.Label2.place(relx=0.709, rely=0.235, height=43, width=158)
                self.Label2.configure(background="#595959")
                self.Label2.configure(disabledforeground="#a3a3a3")
                self.Label2.configure(font=font10)
                self.Label2.configure(foreground="#f0ff1c")
                self.Label2.configure(text='''extra''')
                self.Label2.configure(width=158)

                self.Frame1_1 = tk.Frame(top)
                self.Frame1_1.place(relx=0.03, rely=0.239, relheight=0.693
                        , relwidth=0.247)
                self.Frame1_1.configure(relief='ridge')
                self.Frame1_1.configure(borderwidth="10")
                self.Frame1_1.configure(relief='ridge')
                self.Frame1_1.configure(background="#595959")
                self.Frame1_1.configure(highlightbackground="#d9d9d9")
                self.Frame1_1.configure(highlightcolor="black")
                self.Frame1_1.configure(width=245)

                self.Label3 = tk.Label(self.Frame1_1)
                self.Label3.place(relx=0.163, rely=0.069, height=39, width=168)
                self.Label3.configure(background="#595959")
                self.Label3.configure(disabledforeground="#a3a3a3")
                self.Label3.configure(font=font12)
                self.Label3.configure(foreground="#82ffff")
                self.Label3.configure(text='''UPDATE''')
                self.Label3.configure(width=168)

                self.Label3_5 = tk.Label(self.Frame1_1)
                self.Label3_5.place(relx=0.204, rely=0.184, height=39, width=128)
                self.Label3_5.configure(activebackground="#f9f9f9")
                self.Label3_5.configure(activeforeground="black")
                self.Label3_5.configure(background="#595959")
                self.Label3_5.configure(disabledforeground="#a3a3a3")
                self.Label3_5.configure(font="-family {BadBlocks TT} -size 28 -weight bold")
                self.Label3_5.configure(foreground="#82ffff")
                self.Label3_5.configure(highlightbackground="#d9d9d9")
                self.Label3_5.configure(highlightcolor="black")
                self.Label3_5.configure(text='''DATA''')
                self.Label3_5.configure(width=128)

                self.Label9 = tk.Label(self.Frame1_1)
                self.Label9.place(relx=0.163, rely=0.437, height=65, width=154)
                self.Label9.configure(background="#d9d9d9")
                self.Label9.configure(disabledforeground="#a3a3a3")
                self.Label9.configure(foreground="#000000")
                photo_location = os.path.join(prog_location,"./bg4.png")
                self._img0 = tk.PhotoImage(file=photo_location)
                self.Label9.configure(image=self._img0)
                self.Label9.configure(text='''Label''')

                self.btn_update = tk.Button(self.Frame1_1)
                self.btn_update.place(relx=0.204, rely=0.736, height=44, width=137)
                self.btn_update.configure(activebackground="#ececec")
                self.btn_update.configure(activeforeground="#000000")
                self.btn_update.configure(background="#3045ff")
                self.btn_update.configure(borderwidth="10")
                self.btn_update.configure(disabledforeground="#a3a3a3")
                self.btn_update.configure(font=font15)
                self.btn_update.configure(foreground="#f5fcff")
                self.btn_update.configure(highlightbackground="#d9d9d9")
                self.btn_update.configure(highlightcolor="black")
                self.btn_update.configure(pady="0")
                self.btn_update.configure(text='''ENTER''')
                self.btn_update.configure(width=137,command= self.updates)

                self.Frame1_2 = tk.Frame(top)
                self.Frame1_2.place(relx=0.282, rely=0.239, relheight=0.693
                        , relwidth=0.247)
                self.Frame1_2.configure(relief='ridge')
                self.Frame1_2.configure(borderwidth="10")
                self.Frame1_2.configure(relief='ridge')
                self.Frame1_2.configure(background="#595959")
                self.Frame1_2.configure(highlightbackground="#d9d9d9")
                self.Frame1_2.configure(highlightcolor="black")
                self.Frame1_2.configure(width=245)

                self.Label3_6 = tk.Label(self.Frame1_2)
                self.Label3_6.place(relx=0.163, rely=0.069, height=39, width=168)
                self.Label3_6.configure(activebackground="#f9f9f9")
                self.Label3_6.configure(activeforeground="black")
                self.Label3_6.configure(background="#595959")
                self.Label3_6.configure(disabledforeground="#a3a3a3")
                self.Label3_6.configure(font="-family {BadBlocks TT} -size 28 -weight bold")
                self.Label3_6.configure(foreground="#82ffff")
                self.Label3_6.configure(highlightbackground="#d9d9d9")
                self.Label3_6.configure(highlightcolor="black")
                self.Label3_6.configure(text='''ONLINE''')

                self.Label3_7 = tk.Label(self.Frame1_2)
                self.Label3_7.place(relx=0.163, rely=0.184, height=39, width=168)
                self.Label3_7.configure(activebackground="#f9f9f9")
                self.Label3_7.configure(activeforeground="black")
                self.Label3_7.configure(background="#595959")
                self.Label3_7.configure(disabledforeground="#a3a3a3")
                self.Label3_7.configure(font="-family {BadBlocks TT} -size 28 -weight bold")
                self.Label3_7.configure(foreground="#82ffff")
                self.Label3_7.configure(highlightbackground="#d9d9d9")
                self.Label3_7.configure(highlightcolor="black")
                self.Label3_7.configure(text='''SUPPORT''')

                self.Label8 = tk.Label(self.Frame1_2)
                self.Label8.place(relx=0.204, rely=0.414, height=88, width=154)
                self.Label8.configure(background="#d9d9d9")
                self.Label8.configure(disabledforeground="#a3a3a3")
                self.Label8.configure(foreground="#000000")
                photo_location = os.path.join(prog_location,"./bg3.png")
                self._img1 = tk.PhotoImage(file=photo_location)
                self.Label8.configure(image=self._img1)
                self.Label8.configure(text='''Label''')

                self.btn_online = tk.Button(self.Frame1_2)
                self.btn_online.place(relx=0.204, rely=0.736, height=44, width=137)
                self.btn_online.configure(activebackground="#ececec")
                self.btn_online.configure(activeforeground="#000000")
                self.btn_online.configure(background="#3045ff")
                self.btn_online.configure(borderwidth="10")
                self.btn_online.configure(disabledforeground="#a3a3a3")
                self.btn_online.configure(font="-family {Sky Skunk} -size 12")
                self.btn_online.configure(foreground="#f5fcff")
                self.btn_online.configure(highlightbackground="#d9d9d9")
                self.btn_online.configure(highlightcolor="black")
                self.btn_online.configure(pady="0")
                self.btn_online.configure(text='''ENTER''',command= self.onlines)

                self.Frame1_2 = tk.Frame(top)
                self.Frame1_2.place(relx=0.534, rely=0.239, relheight=0.693
                        , relwidth=0.237)
                self.Frame1_2.configure(relief='ridge')
                self.Frame1_2.configure(borderwidth="10")
                self.Frame1_2.configure(relief='ridge')
                self.Frame1_2.configure(background="#595959")
                self.Frame1_2.configure(highlightbackground="#d9d9d9")
                self.Frame1_2.configure(highlightcolor="black")
                self.Frame1_2.configure(width=235)

                self.Label3_8 = tk.Label(self.Frame1_2)
                self.Label3_8.place(relx=0.255, rely=0.069, height=39, width=118)
                self.Label3_8.configure(activebackground="#f9f9f9")
                self.Label3_8.configure(activeforeground="black")
                self.Label3_8.configure(background="#595959")
                self.Label3_8.configure(disabledforeground="#a3a3a3")
                self.Label3_8.configure(font="-family {BadBlocks TT} -size 28 -weight bold")
                self.Label3_8.configure(foreground="#82ffff")
                self.Label3_8.configure(highlightbackground="#d9d9d9")
                self.Label3_8.configure(highlightcolor="black")
                self.Label3_8.configure(text='''TAX''')
                self.Label3_8.configure(width=118)

                self.Label3_9 = tk.Label(self.Frame1_2)
                self.Label3_9.place(relx=0.085, rely=0.184, height=39, width=198)
                self.Label3_9.configure(activebackground="#f9f9f9")
                self.Label3_9.configure(activeforeground="black")
                self.Label3_9.configure(background="#595959")
                self.Label3_9.configure(disabledforeground="#a3a3a3")
                self.Label3_9.configure(font=font13)
                self.Label3_9.configure(foreground="#82ffff")
                self.Label3_9.configure(highlightbackground="#d9d9d9")
                self.Label3_9.configure(highlightcolor="black")
                self.Label3_9.configure(text='''CALCULATOR''')
                self.Label3_9.configure(width=198)

                self.Label7 = tk.Label(self.Frame1_2)
                self.Label7.place(relx=0.17, rely=0.414, height=97, width=154)
                self.Label7.configure(background="#d9d9d9")
                self.Label7.configure(disabledforeground="#a3a3a3")
                self.Label7.configure(foreground="#000000")
                photo_location = os.path.join(prog_location,"./bg2.png")
                self._img2 = tk.PhotoImage(file=photo_location)
                self.Label7.configure(image=self._img2)
                self.Label7.configure(text='''Label''')

                self.btn_calculator = tk.Button(self.Frame1_2)
                self.btn_calculator.place(relx=0.213, rely=0.736, height=44, width=137)
                self.btn_calculator.configure(activebackground="#ececec")
                self.btn_calculator.configure(activeforeground="#000000")
                self.btn_calculator.configure(background="#3045ff")
                self.btn_calculator.configure(borderwidth="10")
                self.btn_calculator.configure(disabledforeground="#a3a3a3")
                self.btn_calculator.configure(font="-family {Sky Skunk} -size 12")
                self.btn_calculator.configure(foreground="#f5fcff")
                self.btn_calculator.configure(highlightbackground="#d9d9d9")
                self.btn_calculator.configure(highlightcolor="black")
                self.btn_calculator.configure(pady="0")
                self.btn_calculator.configure(text='''ENTER''',command= self.calculators)

                self.Frame1_2 = tk.Frame(top)
                self.Frame1_2.place(relx=0.776, rely=0.239, relheight=0.693
                        , relwidth=0.217)
                self.Frame1_2.configure(relief='ridge')
                self.Frame1_2.configure(borderwidth="10")
                self.Frame1_2.configure(relief='ridge')
                self.Frame1_2.configure(background="#595959")
                self.Frame1_2.configure(highlightbackground="#d9d9d9")
                self.Frame1_2.configure(highlightcolor="black")
                self.Frame1_2.configure(width=215)

                self.Label3_9 = tk.Label(self.Frame1_2)
                self.Label3_9.place(relx=0.093, rely=0.069, height=39, width=178)
                self.Label3_9.configure(activebackground="#f9f9f9")
                self.Label3_9.configure(activeforeground="black")
                self.Label3_9.configure(background="#595959")
                self.Label3_9.configure(disabledforeground="#a3a3a3")
                self.Label3_9.configure(font="-family {BadBlocks TT} -size 28 -weight bold")
                self.Label3_9.configure(foreground="#82ffff")
                self.Label3_9.configure(highlightbackground="#d9d9d9")
                self.Label3_9.configure(highlightcolor="black")
                self.Label3_9.configure(text='''CONNECT''')
                self.Label3_9.configure(width=178)

                self.Label3_9 = tk.Label(self.Frame1_2)
                self.Label3_9.place(relx=0.186, rely=0.161, height=39, width=118)
                self.Label3_9.configure(activebackground="#f9f9f9")
                self.Label3_9.configure(activeforeground="black")
                self.Label3_9.configure(background="#595959")
                self.Label3_9.configure(disabledforeground="#a3a3a3")
                self.Label3_9.configure(font="-family {BadBlocks TT} -size 28 -weight bold")
                self.Label3_9.configure(foreground="#82ffff")
                self.Label3_9.configure(highlightbackground="#d9d9d9")
                self.Label3_9.configure(highlightcolor="black")
                self.Label3_9.configure(text='''TO''')

                self.Label3_9 = tk.Label(self.Frame1_2)
                self.Label3_9.place(relx=0.186, rely=0.253, height=39, width=118)
                self.Label3_9.configure(activebackground="#f9f9f9")
                self.Label3_9.configure(activeforeground="black")
                self.Label3_9.configure(background="#595959")
                self.Label3_9.configure(disabledforeground="#a3a3a3")
                self.Label3_9.configure(font="-family {BadBlocks TT} -size 28 -weight bold")
                self.Label3_9.configure(foreground="#82ffff")
                self.Label3_9.configure(highlightbackground="#d9d9d9")
                self.Label3_9.configure(highlightcolor="black")
                self.Label3_9.configure(text='''WORLD''')

                self.Label6 = tk.Label(self.Frame1_2)
                self.Label6.place(relx=0.14, rely=0.414, height=101, width=154)
                self.Label6.configure(background="#d9d9d9")
                self.Label6.configure(disabledforeground="#a3a3a3")
                self.Label6.configure(foreground="#000000")
                photo_location = os.path.join(prog_location,"./bg1.png")
                self._img3 = tk.PhotoImage(file=photo_location)
                self.Label6.configure(image=self._img3)
                self.Label6.configure(text='''Label''')

                self.btn_world = tk.Button(self.Frame1_2)
                self.btn_world.place(relx=0.186, rely=0.736, height=44, width=137)
                self.btn_world.configure(activebackground="#ececec")
                self.btn_world.configure(activeforeground="#000000")
                self.btn_world.configure(background="#3045ff")
                self.btn_world.configure(borderwidth="10")
                self.btn_world.configure(disabledforeground="#a3a3a3")
                self.btn_world.configure(font="-family {Sky Skunk} -size 12")
                self.btn_world.configure(foreground="#f5fcff")
                self.btn_world.configure(highlightbackground="#d9d9d9")
                self.btn_world.configure(highlightcolor="black")
                self.btn_world.configure(pady="0")
                self.btn_world.configure(text='''ENTER''',command= self.worlds);

                self.Label4 = tk.Label(top)
                self.Label4.place(relx=0.685, rely=0.048, height=28, width=184)
                self.Label4.configure(background="#939393")
                self.Label4.configure(disabledforeground="#a3a3a3")
                self.Label4.configure(font=font14)
                self.Label4.configure(foreground="#0e4705")
                self.Label4.configure(text='''Village : Kalamwadi''')

                self.Label4_10 = tk.Label(top)
                self.Label4_10.place(relx=0.716, rely=0.096, height=28, width=154)
                self.Label4_10.configure(activebackground="#f9f9f9")
                self.Label4_10.configure(activeforeground="black")
                self.Label4_10.configure(background="#939393")
                self.Label4_10.configure(disabledforeground="#a3a3a3")
                self.Label4_10.configure(font="-family {Serifa BT} -size 14 -weight bold")
                self.Label4_10.configure(foreground="#0e4705")
                self.Label4_10.configure(highlightbackground="#d9d9d9")
                self.Label4_10.configure(highlightcolor="black")
                self.Label4_10.configure(text='''District : Sangli''')
                self.Label4_10.configure(width=154)

                self.TSeparator1 = ttk.Separator(top)
                self.TSeparator1.place(relx=0.877, rely=0.032, relheight=0.127)
                self.TSeparator1.configure(orient="vertical")

                self.Label5 = tk.Label(top)
                self.Label5.place(relx=0.917, rely=0.048, height=44, width=44)
                self.Label5.configure(background="#d9d9d9")
                self.Label5.configure(disabledforeground="#a3a3a3")
                self.Label5.configure(foreground="#000000")
                photo_location = os.path.join(prog_location,"./login3.png")
                self._img4 = tk.PhotoImage(file=photo_location)
                self.Label5.configure(image=self._img4)
                self.Label5.configure(text='''Label''')

                self.Label4_11 = tk.Label(top)
                self.Label4_11.place(relx=0.907, rely=0.127, height=28, width=64)
                self.Label4_11.configure(activebackground="#f9f9f9")
                self.Label4_11.configure(activeforeground="black")
                self.Label4_11.configure(background="#939393")
                self.Label4_11.configure(disabledforeground="#a3a3a3")
                self.Label4_11.configure(font="-family {Serifa BT} -size 14 -weight bold")
                self.Label4_11.configure(foreground="#0e4705")
                self.Label4_11.configure(highlightbackground="#d9d9d9")
                self.Label4_11.configure(highlightcolor="black")
                self.Label4_11.configure(text='''User''')
                self.Label4_11.configure(width=64)

                self.btn_exit = tk.Button(top)
                self.btn_exit.place(relx=0.847, rely=0.939, height=34, width=107)
                self.btn_exit.configure(activebackground="#ececec")
                self.btn_exit.configure(activeforeground="#000000")
                self.btn_exit.configure(background="#ff4f19")
                self.btn_exit.configure(borderwidth="10")
                self.btn_exit.configure(disabledforeground="#a3a3a3")
                self.btn_exit.configure(font="-family {Sky Skunk} -size 12")
                self.btn_exit.configure(foreground="#f5fcff")
                self.btn_exit.configure(highlightbackground="#d9d9d9")
                self.btn_exit.configure(highlightcolor="black")
                self.btn_exit.configure(pady="0")
                self.btn_exit.configure(text='''EXIT''',command=self.exits)

                self.Label10 = tk.Label(top)
                self.Label10.place(relx=0.03, rely=0.939, height=21, width=84)
                self.Label10.configure(background="#939393")
                self.Label10.configure(disabledforeground="#a3a3a3")
                self.Label10.configure(foreground="#000000")
                self.Label10.configure(text='''e-TAX 2019''')
                self.Label10.configure(width=84)

                self.Label10_9 = tk.Label(top)
                self.Label10_9.place(relx=0.03, rely=0.971, height=11, width=84)
                self.Label10_9.configure(activebackground="#f9f9f9")
                self.Label10_9.configure(activeforeground="black")
                self.Label10_9.configure(background="#939393")
                self.Label10_9.configure(disabledforeground="#a3a3a3")
                self.Label10_9.configure(foreground="#000000")
                self.Label10_9.configure(highlightbackground="#d9d9d9")
                self.Label10_9.configure(highlightcolor="black")
                self.Label10_9.configure(text='''V 1 . 0 . 2''')
                self.Label10_9.configure(width=84)

                self.btn_back = tk.Button(top)
                self.btn_back.place(relx=0.716, rely=0.939, height=34, width=107)
                self.btn_back.configure(activebackground="#ececec")
                self.btn_back.configure(activeforeground="#000000")
                self.btn_back.configure(background="#387738")
                self.btn_back.configure(borderwidth="10")
                self.btn_back.configure(disabledforeground="#a3a3a3")
                self.btn_back.configure(font="-family {Sky Skunk} -size 12")
                self.btn_back.configure(foreground="#f5fcff")
                self.btn_back.configure(highlightbackground="#d9d9d9")
                self.btn_back.configure(highlightcolor="black")
                self.btn_back.configure(pady="0")
                self.btn_back.configure(text='''BACK''')
                self.btn_back.configure(width=107,command=self.backs)
        if __name__ == '__main__':
            vp_start_gui()













































































    class Main:
        def newdatas(self):
            root.destroy()
            newdata1()
        def openworkspaces(self):
            root.destroy()
            openworkspace1()
        def viewentrys(self):
            root.destroy()
            viewentry1()
        def extras(self):
            root.destroy()
            print("Hi User WELCOME to eTAX-2019 extra\n\n\n")
            print("*"*10)
            print("\n Please Enter your security pin before entering in this module \n")
            print("\n\n\n\n\n\n Write Your Password Here :::::\n")
            pinn=str(input())
            if pinn != "Pass@123":
                msg= tkinter.messagebox.showinfo("eTAX-2019",'Wrong Password')
                os._exit(1)
            extra1()
        def exits(self):
            msg=tkinter.messagebox.askyesno("Main","Do You Want To Exit ?")
            if msg:
                root.destroy()
                os._exit(1)
        def logouts(self):
            tkinter.messagebox.showinfo("Main", "There was some error while logging out. Please restart your application. You will be logged out")
            os._exit(1)
            
        def __init__(self, top=None):
            '''This class configures and populates the toplevel window.
               top is the toplevel containing window.'''
            _bgcolor = '#d9d9d9'  # X11 color: 'gray85'
            _fgcolor = '#000000'  # X11 color: 'black'
            _compcolor = '#d9d9d9' # X11 color: 'gray85'
            _ana1color = '#d9d9d9' # X11 color: 'gray85' 
            _ana2color = '#ececec' # Closest X11 color: 'gray92' 
            font11 = "-family {Arial Black} -size 36 -weight bold -slant "  \
                "roman -underline 0 -overstrike 0"
            font12 = "-family Georgia -size 12 -weight normal -slant roman"  \
                " -underline 0 -overstrike 0"
            font16 = "-family Cambria -size 11 -weight normal -slant roman"  \
                " -underline 0 -overstrike 0"
            font9 = "-family {Tw Cen MT Condensed Extra Bold} -size 22 "  \
                "-weight bold -slant roman -underline 0 -overstrike 0"

            top.geometry("925x601+294+152")
            top.title("Main")
            top.configure(background="#d9d9d9")
            top.configure(highlightbackground="#d9d9d9")
            top.configure(highlightcolor="black")

            self.Canvas1 = tk.Canvas(top)
            self.Canvas1.place(relx=-0.043, rely=-0.15, relheight=1.186
                    , relwidth=1.084)
            self.Canvas1.configure(background="#faff69")
            self.Canvas1.configure(borderwidth="2")
            self.Canvas1.configure(highlightbackground="#d9d9d9")
            self.Canvas1.configure(highlightcolor="black")
            self.Canvas1.configure(insertbackground="black")
            self.Canvas1.configure(relief='ridge')
            self.Canvas1.configure(selectbackground="#c4c4c4")
            self.Canvas1.configure(selectforeground="black")
            self.Canvas1.configure(width=1003)

            self.Frame1 = tk.Frame(self.Canvas1)
            self.Frame1.place(relx=0.15, rely=0.323, relheight=0.259, relwidth=0.314)

            self.Frame1.configure(relief='groove')
            self.Frame1.configure(borderwidth="2")
            self.Frame1.configure(relief='groove')
            self.Frame1.configure(background="#5b63d8")
            self.Frame1.configure(highlightbackground="#d9d9d9")
            self.Frame1.configure(highlightcolor="black")
            self.Frame1.configure(width=315)

            self.Label1 = tk.Label(self.Frame1)
            self.Label1.place(relx=0.19, rely=0.162, height=40, width=196)
            self.Label1.configure(activebackground="#f9f9f9")
            self.Label1.configure(activeforeground="black")
            self.Label1.configure(background="#5b63d8")
            self.Label1.configure(disabledforeground="#a3a3a3")
            self.Label1.configure(font=font9)
            self.Label1.configure(foreground="#ffffff")
            self.Label1.configure(highlightbackground="#d9d9d9")
            self.Label1.configure(highlightcolor="black")
            self.Label1.configure(text='''New Data Entry''')

            self.newdata = tk.Button(self.Frame1)
            self.newdata.place(relx=0.381, rely=0.541, height=34, width=77)
            self.newdata.configure(activebackground="#ffffff")
            self.newdata.configure(activeforeground="#000000")
            self.newdata.configure(background="#ffffff")
            self.newdata.configure(disabledforeground="#a3a3a3")
            self.newdata.configure(foreground="#5b63d8")
            self.newdata.configure(highlightbackground="#d9d9d9")
            self.newdata.configure(highlightcolor="black")
            self.newdata.configure(pady="0")
            self.newdata.configure(relief='groove')
            self.newdata.configure(text='''ENTER''')
            self.newdata.configure(command=self.newdatas);

            self.Frame1_3 = tk.Frame(self.Canvas1)
            self.Frame1_3.place(relx=0.568, rely=0.323, relheight=0.259
                    , relwidth=0.314)
            self.Frame1_3.configure(relief='groove')
            self.Frame1_3.configure(borderwidth="2")
            self.Frame1_3.configure(relief='groove')
            self.Frame1_3.configure(background="#5b63d8")
            self.Frame1_3.configure(highlightbackground="#d9d9d9")
            self.Frame1_3.configure(highlightcolor="black")
            self.Frame1_3.configure(width=315)

            self.Label1_4 = tk.Label(self.Frame1_3)
            self.Label1_4.place(relx=0.095, rely=0.162, height=40, width=250)
            self.Label1_4.configure(activebackground="#f9f9f9")
            self.Label1_4.configure(activeforeground="black")
            self.Label1_4.configure(background="#5b63d8")
            self.Label1_4.configure(disabledforeground="#a3a3a3")
            self.Label1_4.configure(font=font9)
            self.Label1_4.configure(foreground="#ffffff")
            self.Label1_4.configure(highlightbackground="#d9d9d9")
            self.Label1_4.configure(highlightcolor="black")
            self.Label1_4.configure(text='''Open The Workspace''')

            self.openworkspace = tk.Button(self.Frame1_3)
            self.openworkspace.place(relx=0.381, rely=0.541, height=34, width=77)
            self.openworkspace.configure(activebackground="#ffffff")
            self.openworkspace.configure(activeforeground="#000000")
            self.openworkspace.configure(background="#ffffff")
            self.openworkspace.configure(disabledforeground="#a3a3a3")
            self.openworkspace.configure(foreground="#5b63d8")
            self.openworkspace.configure(highlightbackground="#d9d9d9")
            self.openworkspace.configure(highlightcolor="black")
            self.openworkspace.configure(pady="0")
            self.openworkspace.configure(relief='groove')
            self.openworkspace.configure(text='''ENTER''')
            self.openworkspace.configure(command=self.openworkspaces);


            self.Frame1_6 = tk.Frame(self.Frame1_3)
            self.Frame1_6.place(relx=-0.349, rely=1.676, relheight=1.0, relwidth=1.0)

            self.Frame1_6.configure(relief='groove')
            self.Frame1_6.configure(borderwidth="2")
            self.Frame1_6.configure(relief='groove')
            self.Frame1_6.configure(background="#5b63d8")
            self.Frame1_6.configure(highlightbackground="#d9d9d9")
            self.Frame1_6.configure(highlightcolor="black")
            self.Frame1_6.configure(width=315)

            self.Label1_7 = tk.Label(self.Frame1_6)
            self.Label1_7.place(relx=0.19, rely=0.162, height=40, width=196)
            self.Label1_7.configure(activebackground="#f9f9f9")
            self.Label1_7.configure(activeforeground="black")
            self.Label1_7.configure(background="#5b63d8")
            self.Label1_7.configure(disabledforeground="#a3a3a3")
            self.Label1_7.configure(font=font9)
            self.Label1_7.configure(foreground="#ffffff")
            self.Label1_7.configure(highlightbackground="#d9d9d9")
            self.Label1_7.configure(highlightcolor="black")
            self.Label1_7.configure(text='''New Data Entry''')

            self.Frame1_12 = tk.Frame(self.Canvas1)
            self.Frame1_12.place(relx=0.15, rely=0.659, relheight=0.259
                    , relwidth=0.314)
            self.Frame1_12.configure(relief='groove')
            self.Frame1_12.configure(borderwidth="2")
            self.Frame1_12.configure(relief='groove')
            self.Frame1_12.configure(background="#5b63d8")
            self.Frame1_12.configure(highlightbackground="#d9d9d9")
            self.Frame1_12.configure(highlightcolor="black")
            self.Frame1_12.configure(width=315)

            self.Label1_13 = tk.Label(self.Frame1_12)
            self.Label1_13.place(relx=0.19, rely=0.162, height=40, width=191)
            self.Label1_13.configure(activebackground="#f9f9f9")
            self.Label1_13.configure(activeforeground="black")
            self.Label1_13.configure(background="#5b63d8")
            self.Label1_13.configure(disabledforeground="#a3a3a3")
            self.Label1_13.configure(font=font9)
            self.Label1_13.configure(foreground="#ffffff")
            self.Label1_13.configure(highlightbackground="#d9d9d9")
            self.Label1_13.configure(highlightcolor="black")
            self.Label1_13.configure(text='''View All Entries''')

            self.viewentry = tk.Button(self.Frame1_12)
            self.viewentry.place(relx=0.381, rely=0.541, height=34, width=77)
            self.viewentry.configure(activebackground="#ffffff")
            self.viewentry.configure(activeforeground="#000000")
            self.viewentry.configure(background="#ffffff")
            self.viewentry.configure(disabledforeground="#a3a3a3")
            self.viewentry.configure(foreground="#5b63d8")
            self.viewentry.configure(highlightbackground="#d9d9d9")
            self.viewentry.configure(highlightcolor="black")
            self.viewentry.configure(pady="0")
            self.viewentry.configure(relief='groove')
            self.viewentry.configure(text='''ENTER''')
            self.viewentry.configure(command=self.viewentrys);


            self.Frame1_13 = tk.Frame(self.Canvas1)
            self.Frame1_13.place(relx=0.568, rely=0.659, relheight=0.259
                    , relwidth=0.314)
            self.Frame1_13.configure(relief='groove')
            self.Frame1_13.configure(borderwidth="2")
            self.Frame1_13.configure(relief='groove')
            self.Frame1_13.configure(background="#5b63d8")
            self.Frame1_13.configure(highlightbackground="#d9d9d9")
            self.Frame1_13.configure(highlightcolor="black")
            self.Frame1_13.configure(width=315)

            self.Label1_14 = tk.Label(self.Frame1_13)
            self.Label1_14.place(relx=0.222, rely=0.162, height=40, width=175)
            self.Label1_14.configure(activebackground="#f9f9f9")
            self.Label1_14.configure(activeforeground="black")
            self.Label1_14.configure(background="#5b63d8")
            self.Label1_14.configure(disabledforeground="#a3a3a3")
            self.Label1_14.configure(font=font9)
            self.Label1_14.configure(foreground="#ffffff")
            self.Label1_14.configure(highlightbackground="#d9d9d9")
            self.Label1_14.configure(highlightcolor="black")
            self.Label1_14.configure(text='''Extra Features''')

            self.extra = tk.Button(self.Frame1_13)
            self.extra.place(relx=0.381, rely=0.541, height=34, width=77)
            self.extra.configure(activebackground="#ffffff")
            self.extra.configure(activeforeground="#000000")
            self.extra.configure(background="#ffffff")
            self.extra.configure(disabledforeground="#a3a3a3")
            self.extra.configure(foreground="#5b63d8")
            self.extra.configure(highlightbackground="#d9d9d9")
            self.extra.configure(highlightcolor="black")
            self.extra.configure(pady="0")
            self.extra.configure(relief='groove')
            self.extra.configure(text='''ENTER''')
            self.extra.configure(command=self.extras);


            self.Label2 = tk.Label(self.Canvas1)
            self.Label2.place(relx=0.349, rely=0.168, height=74, width=170)
            self.Label2.configure(background="#faff69")
            self.Label2.configure(disabledforeground="#a3a3a3")
            self.Label2.configure(font=font11)
            self.Label2.configure(foreground="#ff351f")
            self.Label2.configure(text='''eTAX''')
            self.Label2.configure(width=170)

            self.Label2_1 = tk.Label(self.Canvas1)
            self.Label2_1.place(relx=0.499, rely=0.168, height=74, width=170)
            self.Label2_1.configure(activebackground="#f9f9f9")
            self.Label2_1.configure(activeforeground="black")
            self.Label2_1.configure(background="#faff69")
            self.Label2_1.configure(disabledforeground="#a3a3a3")
            self.Label2_1.configure(font=font11)
            self.Label2_1.configure(foreground="#201ad8")
            self.Label2_1.configure(highlightbackground="#d9d9d9")
            self.Label2_1.configure(highlightcolor="black")
            self.Label2_1.configure(text='''2019''')

            self.Label3 = tk.Label(self.Canvas1)
            self.Label3.place(relx=0.877, rely=0.154, height=44, width=44)
            self.Label3.configure(background="#d9d9d9")
            self.Label3.configure(disabledforeground="#a3a3a3")
            self.Label3.configure(foreground="#000000")
            self._img1 = tk.PhotoImage(file="./login3.png")
            self.Label3.configure(image=self._img1)
            self.Label3.configure(text='''Label''')

            self.Label4 = tk.Label(self.Canvas1)
            self.Label4.place(relx=0.857, rely=0.224, height=24, width=96)
            self.Label4.configure(background="#faff69")
            self.Label4.configure(disabledforeground="#a3a3a3")
            self.Label4.configure(font=font12)
            self.Label4.configure(foreground="#000000")
            self.Label4.configure(text='''Hello user !!!''')

            self.Label5 = tk.Label(self.Canvas1)
            self.Label5.place(relx=0.05, rely=0.912, height=21, width=59)
            self.Label5.configure(background="#faff69")
            self.Label5.configure(disabledforeground="#a3a3a3")
            self.Label5.configure(foreground="#000000")
            self.Label5.configure(text='''e-tax 2019''')
            self.Label5.configure(width=59)

            self.Label5_2 = tk.Label(self.Canvas1)
            self.Label5_2.place(relx=0.05, rely=0.94, height=21, width=59)
            self.Label5_2.configure(activebackground="#f9f9f9")
            self.Label5_2.configure(activeforeground="black")
            self.Label5_2.configure(anchor='w')
            self.Label5_2.configure(background="#faff69")
            self.Label5_2.configure(disabledforeground="#a3a3a3")
            self.Label5_2.configure(foreground="#000000")
            self.Label5_2.configure(highlightbackground="#d9d9d9")
            self.Label5_2.configure(highlightcolor="black")
            self.Label5_2.configure(text='''v 1.0.2''')

            self.logout = tk.Button(self.Canvas1)
            self.logout.place(relx=0.07, rely=0.154, height=34, width=77)
            self.logout.configure(activebackground="#ffffff")
            self.logout.configure(activeforeground="#000000")
            self.logout.configure(background="#5b63d8")
            self.logout.configure(disabledforeground="#a3a3a3")
            self.logout.configure(font=font16)
            self.logout.configure(foreground="#ffffff")
            self.logout.configure(highlightbackground="#d9d9d9")
            self.logout.configure(highlightcolor="black")
            self.logout.configure(pady="0")
            self.logout.configure(relief='groove')
            self.logout.configure(text='''LOG OUT''')
            self.logout.configure(command=self.logouts);        # function defined at  line 


            self.exit = tk.Button(self.Canvas1)
            self.exit.place(relx=0.07, rely=0.224, height=34, width=77)
            self.exit.configure(activebackground="#ffffff")
            self.exit.configure(activeforeground="#000000")
            self.exit.configure(background="#5b63d8")
            self.exit.configure(cursor="fleur")
            self.exit.configure(disabledforeground="#a3a3a3")
            self.exit.configure(font=font16)
            self.exit.configure(foreground="#ffffff")
            self.exit.configure(highlightbackground="#d9d9d9")
            self.exit.configure(highlightcolor="black")
            self.exit.configure(pady="0")
            self.exit.configure(relief='groove')
            self.exit.configure(text='''EXIT''')
            self.exit.configure(command=self.exits);       # function defined at line 


    if __name__ == '__main__':
        vp_start_gui()






























def vp_start_gui():
    '''Starting point when module is the main routine.'''
    global val, w, root
    root = tk.Tk()
    top = loginpage (root)
    unknown_support.init(root, top)
    root.mainloop()

w = None
def create_loginpage(root, *args, **kwargs):
    '''Starting point when module is imported by another program.'''
    global w, w_win, rt
    rt = root
    w = tk.Toplevel (root)
    top = loginpage (w)
    unknown_support.init(w, top, *args, **kwargs)
    return (w, top)

def destroy_loginpage():
    global w
    w.destroy()
    w = None

class loginpage:
    def cancellogin(self):
        msg=tkinter.messagebox.askyesno("Login Page","Are You Sure , Do you Want to Cancel Login ?");         # Showing messagebox that asks user to cancel login
        if msg:
            os._exit(1)
    def loginuser(self):
        name=self.Entry1.get();               # taking username input
        password1=self.Entry2.get();           # taking password input
        try :
            u='\''+name+'\''
            mydb=mysql.connector.connect(host='remotemysql.com',user='NhpDbfVMT3',password='NXWSVJNu55',database='NhpDbfVMT3',port='3306')
            mycursor=mydb.cursor()
            query = ("SELECT password FROM passwords WHERE username ="+u)
            mycursor.execute(query)
            for(password) in mycursor:
                s="{}".format(password)
                s=str(s[2:len(s)-3])
                if s==password1:
                    messagebox.showinfo('etax-2019','Access Granted')
                    root.destroy()
                    main()
                else:
                    messagebox.showerror("eTAX-2019",'Incorrect Password')
        except:
            msg=tkinter.messagebox.showerror("Login Page","Server Connection Failed. Please contact the administrator");
            os._exit(1)
    def __init__(self, top=None):
        '''This class configures and populates the toplevel window.
           top is the toplevel containing window.'''
        _bgcolor = '#d9d9d9'  # X11 color: 'gray85'
        _fgcolor = '#000000'  # X11 color: 'black'
        _compcolor = '#d9d9d9' # X11 color: 'gray85'
        _ana1color = '#d9d9d9' # X11 color: 'gray85' 
        _ana2color = '#ececec' # Closest X11 color: 'gray92' 
        font10 = "-family {MS UI Gothic} -size 14 -weight bold -slant "  \
            "roman -underline 0 -overstrike 0"
        font11 = "-family Rockwell -size 17 -weight bold -slant roman "  \
            "-underline 0 -overstrike 0"
        font15 = "-family {Rockwell Extra Bold} -size 23 -weight bold "  \
            "-slant roman -underline 0 -overstrike 0"
        font9 = "-family Rockwell -size 14 -weight bold -slant roman "  \
            "-underline 0 -overstrike 0"

        top.geometry("763x561+439+167")
        top.title("New Toplevel")
        top.configure(background="#ffff30")

        self.Frame1 = tk.Frame(top)
        self.Frame1.place(relx=0.406, rely=0.374, relheight=0.651
                , relwidth=0.688)
        self.Frame1.configure(relief='groove')
        self.Frame1.configure(borderwidth="2")
        self.Frame1.configure(relief='groove')
        self.Frame1.configure(background="#5b2af9")
        self.Frame1.configure(width=525)

        self.Entry1 = tk.Entry(self.Frame1)
        self.Entry1.place(relx=0.381, rely=0.438,height=20, relwidth=0.408)
        self.Entry1.configure(background="white")
        self.Entry1.configure(disabledforeground="#a3a3a3")
        self.Entry1.configure(font="TkFixedFont")
        self.Entry1.configure(foreground="#000000")
        self.Entry1.configure(insertbackground="black")
        self.Entry1.configure(width=214)

        self.Entry2 = tk.Entry(self.Frame1)
        self.Entry2.place(relx=0.381, rely=0.575,height=20, relwidth=0.408)
        self.Entry2.configure(background="white")
        self.Entry2.configure(disabledforeground="#a3a3a3")
        self.Entry2.configure(font="TkFixedFont")
        self.Entry2.configure(foreground="#000000")
        self.Entry2.configure(insertbackground="black")
        self.Entry2.configure(width=214)

        self.Label1 = tk.Label(self.Frame1)
        self.Label1.place(relx=0.133, rely=0.438, height=21, width=114)
        self.Label1.configure(background="#5b2af9")
        self.Label1.configure(disabledforeground="#5b2af9")
        self.Label1.configure(font=font9)
        self.Label1.configure(foreground="#000000")
        self.Label1.configure(text='''Username :''')
        self.Label1.configure(width=114)

        self.Label2 = tk.Label(self.Frame1)
        self.Label2.place(relx=0.133, rely=0.548, height=29, width=109)
        self.Label2.configure(background="#5b2af9")
        self.Label2.configure(disabledforeground="#a3a3a3")
        self.Label2.configure(font=font9)
        self.Label2.configure(foreground="#000000")
        self.Label2.configure(text='''Password :''')

        self.Button1 = tk.Button(self.Frame1)
        self.Button1.place(relx=0.61, rely=0.712, height=24, width=87)
        self.Button1.configure(activebackground="#ececec")
        self.Button1.configure(activeforeground="#000000")
        self.Button1.configure(background="#3bfff8")
        self.Button1.configure(disabledforeground="#a3a3a3")
        self.Button1.configure(font=font10)
        self.Button1.configure(foreground="#000000")
        self.Button1.configure(highlightbackground="#d9d9d9")
        self.Button1.configure(highlightcolor="black")
        self.Button1.configure(pady="0")
        self.Button1.configure(text='''LOGIN''')
        self.Button1.configure(width=87)                            # function defined at line1453
        self.Button1.configure(command=self.loginuser);             # Giving command to button1

        self.Button1_1 = tk.Button(self.Frame1)
        self.Button1_1.place(relx=0.4, rely=0.712, height=24, width=87)
        self.Button1_1.configure(activebackground="#ececec")
        self.Button1_1.configure(activeforeground="#000000")
        self.Button1_1.configure(background="#3bfff8")
        self.Button1_1.configure(disabledforeground="#a3a3a3")
        self.Button1_1.configure(font=font10)
        self.Button1_1.configure(foreground="#000000")
        self.Button1_1.configure(highlightbackground="#d9d9d9")
        self.Button1_1.configure(highlightcolor="black")
        self.Button1_1.configure(pady="0")
        self.Button1_1.configure(text='''CANCEL''')                 # function defined at line 1449
        self.Button1_1.configure(command=self.cancellogin);         # Giving command to button1_1

        self.Frame2 = tk.Frame(top)
        self.Frame2.place(relx=-0.013, rely=-0.053, relheight=0.651
                , relwidth=0.609)
        self.Frame2.configure(relief='groove')
        self.Frame2.configure(borderwidth="2")
        self.Frame2.configure(relief='groove')
        self.Frame2.configure(background="#1fa53a")
        self.Frame2.configure(width=465)

        self.Label3 = tk.Label(self.Frame2)
        self.Label3.place(relx=0.086, rely=0.192, height=51, width=154)
        self.Label3.configure(background="#1fa53a")
        self.Label3.configure(disabledforeground="#a3a3a3")
        self.Label3.configure(font=font11)
        self.Label3.configure(foreground="#083a06")
        self.Label3.configure(text='''WELCOME''')
        self.Label3.configure(width=154)

        self.Label3_2 = tk.Label(self.Frame2)
        self.Label3_2.place(relx=0.194, rely=0.466, height=41, width=114)
        self.Label3_2.configure(activebackground="#f9f9f9")
        self.Label3_2.configure(activeforeground="black")
        self.Label3_2.configure(background="#1fa53a")
        self.Label3_2.configure(disabledforeground="#a3a3a3")
        self.Label3_2.configure(font=font15)
        self.Label3_2.configure(foreground="#d12613")
        self.Label3_2.configure(highlightbackground="#d9d9d9")
        self.Label3_2.configure(highlightcolor="black")
        self.Label3_2.configure(text='''eTAX''')
        self.Label3_2.configure(width=114)

        self.Label3_3 = tk.Label(self.Frame2)
        self.Label3_3.place(relx=0.43, rely=0.466, height=41, width=94)
        self.Label3_3.configure(activebackground="#f9f9f9")
        self.Label3_3.configure(activeforeground="black")
        self.Label3_3.configure(background="#1fa53a")
        self.Label3_3.configure(disabledforeground="#a3a3a3")
        self.Label3_3.configure(font=font15)
        self.Label3_3.configure(foreground="#2c18ad")
        self.Label3_3.configure(highlightbackground="#d9d9d9")
        self.Label3_3.configure(highlightcolor="black")
        self.Label3_3.configure(text='''2019''')
        self.Label3_3.configure(width=94)

        self.Label3_4 = tk.Label(self.Frame2)
        self.Label3_4.place(relx=0.215, rely=0.329, height=31, width=94)
        self.Label3_4.configure(activebackground="#f9f9f9")
        self.Label3_4.configure(activeforeground="black")
        self.Label3_4.configure(background="#1fa53a")
        self.Label3_4.configure(disabledforeground="#a3a3a3")
        self.Label3_4.configure(font=font11)
        self.Label3_4.configure(foreground="#083a06")
        self.Label3_4.configure(highlightbackground="#d9d9d9")
        self.Label3_4.configure(highlightcolor="black")
        self.Label3_4.configure(text='''TO''')
        self.Label3_4.configure(width=94)

        self.Frame3 = tk.Frame(top)
        self.Frame3.place(relx=-0.092, rely=0.499, relheight=0.704
                , relwidth=0.557)
        self.Frame3.configure(relief='groove')
        self.Frame3.configure(borderwidth="2")
        self.Frame3.configure(relief='groove')
        self.Frame3.configure(background="#ff2d0d")
        self.Frame3.configure(width=425)

        self.Label4 = tk.Label(self.Frame3)
        self.Label4.place(relx=0.306, rely=-0.025, height=182, width=298)
        self.Label4.configure(background="#d9d9d9")
        self.Label4.configure(disabledforeground="#a3a3a3")
        self.Label4.configure(foreground="#000000")
        self._img1 = tk.PhotoImage(file="./login1.png")
        self.Label4.configure(image=self._img1)
        self.Label4.configure(text='''Label''')
        self.Label4.configure(width=298)

        self.menubar = tk.Menu(top,font="TkMenuFont",bg=_bgcolor,fg=_fgcolor)
        top.configure(menu = self.menubar)

        self.Label5 = tk.Label(top)
        self.Label5.place(relx=0.734, rely=0.428, height=94, width=104)
        self.Label5.configure(background="#d9d9d9")
        self.Label5.configure(disabledforeground="#a3a3a3")
        self.Label5.configure(foreground="#000000")
        self._img2 = tk.PhotoImage(file="./login2.png")
        self.Label5.configure(image=self._img2)
        self.Label5.configure(text='''Label''')
        self.Label5.configure(width=104)

if __name__ == '__main__':
    vp_start_gui()





